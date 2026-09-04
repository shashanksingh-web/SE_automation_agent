#!/usr/bin/env python3
"""
SE/DC Data Normalization Agent + SE Daily Plan Generator.

Implements the pipeline in SE_DC_Data_Normalization_Agent_Prompt.docx:
    raw feeds (6 sources) -> normalized data + exceptions report
        -> BO1-BO5 scoring -> SE Daily Plan

Source coverage:
    Source 2 (DC Master + Rank)        -> local file  DC_RAnk.csv
    Source 5 (Config Master)           -> local files  BO_Configuration_Sheet_v3 - *.csv
    Source 6 (AOP & Target Data)       -> local file   Niyojan Q2-FY_26_27 Dashboard - Planning.csv
    Source 1 (SE/ABM Master + Tasks)   -> live, via Metabase REST API (Redshift db_id=41,
    Source 3 (Transactional/Field)        input-backend Postgres db_id=31 -- both confirmed
    Source 4 (Active Roster/History)      live against this Metabase instance) + kheti
                                           Postgres db_id=4 for the hyperlocal_order proxy.

Sources 1/3/4 require METABASE_URL and METABASE_API_KEY (Metabase Admin -> API Keys) to
be set as environment variables -- this is the standalone equivalent of the live Metabase
access used to confirm the schema this file queries. Without them the pipeline still runs
end-to-end on Sources 2/5/6 alone; every table and the run summary say so explicitly
rather than silently producing partial output as if it were complete (Section 8 guardrail).

This agent normalizes and quarantines; it does not invent values for genuine data gaps.
The SE Daily Plan section (Section 11 below) is an explicit extension beyond the
normalization agent's own documented scope ("prepares data, does not generate plans"),
built to the output shape the doc's own Section 10 ("Downstream Outcome") specifies.
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import re
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Set, Tuple

try:
    import requests
except ImportError:  # pragma: no cover - degrade gracefully, see MetabaseClient
    requests = None

try:
    import openpyxl
except ImportError:  # pragma: no cover - degrade gracefully, see load_top_dc_allowlist
    openpyxl = None

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent / ".env")  # REDSHIFT_*/METABASE_* -- gitignored
except ImportError:  # pragma: no cover - fine to run with plain env vars instead
    pass

Table = List[Dict[str, Any]]

BASE_DIR = Path(__file__).resolve().parent
logger = logging.getLogger("se_daily_plan_agent")


# =====================================================================================
# 0. CONFIG -- environment, file locations, confirmed database ids
# =====================================================================================

REDSHIFT_DB_ID = int(os.environ.get("SE_AGENT_REDSHIFT_DB_ID", "41"))          # "Redshift"
INPUT_BACKEND_DB_ID = int(os.environ.get("SE_AGENT_INPUT_BACKEND_DB_ID", "31"))  # "input-backend"
KHETI_DB_ID = int(os.environ.get("SE_AGENT_KHETI_DB_ID", "4"))                 # "kheti"

DC_MASTER_CSV = Path(os.environ.get("SE_AGENT_DC_MASTER_CSV", BASE_DIR / "DC_RAnk.csv"))
# Added 2026-09-04, explicit user request -- an independent allowlist on top of
# DC_RAnk.csv's own Rank<=6000 eligibility (see BusinessConstants.max_eligible_rank),
# not a replacement for it. Confirmed live: this file has NO Rank/Cohort/Score columns
# at all (just Partner Id/Name/Sales Rep/Node/State/Dehaat Club Scheme Slab/Prize/
# Sales) -- it restricts WHICH DCs are eligible, DC_RAnk.csv still decides HOW they're
# ranked among themselves. 3,143 rows / 2,517 unique Partner Ids as of 2026-09-04 (626
# duplicate rows, harmless for a set-membership check).
TOP_DC_LIST_XLSX = Path(os.environ.get("SE_AGENT_TOP_DC_LIST_XLSX", BASE_DIR / "updated TOP DC list.xlsx"))
# Moved into a subfolder 2026-08-06 -- note the trailing space in the folder name, that's
# literal (confirmed via `ls`), not a typo to "fix". All 5 BO_Configuration_Sheet_v3 files
# live here now; DC_RAnk.csv (Source 2) and the AOP dashboard (Source 6) stayed at BASE_DIR.
CONFIG_DIR = Path(os.environ.get("SE_AGENT_CONFIG_DIR", BASE_DIR / "config and parameter "))
CONFIG_ALL_PARAMS_CSV = Path(
    # Bumped to (3) 2026-08-12 -- the (2) export was replaced on disk, not just
    # supplemented; pointing at a deleted file silently degraded every run to
    # Config_File_Missing since the upload, not just a cosmetic mismatch.
    os.environ.get("SE_AGENT_CONFIG_CSV", CONFIG_DIR / "BO_Configuration_Sheet_v3 - All Parameters (3).csv")
)
CONFIG_SE_INCENTIVE_CSV = Path(
    os.environ.get(
        "SE_AGENT_INCENTIVE_CSV", CONFIG_DIR / "BO_Configuration_Sheet_v3 - SE Incentive Policy (FY26-27) (3).csv"
    )
)
# "Open Questions (Sec 9).csv" was retired in the 2026-08-06 re-sync -- superseded by
# "Agent-Determined Parameters.csv" (same 7-row dynamic-parameter list, restructured).
# This path is kept for backward compatibility only; it will correctly report
# Config_File_Missing if ever loaded, which is honest -- the file is gone, not renamed.
CONFIG_OPEN_QUESTIONS_CSV = Path(
    os.environ.get("SE_AGENT_OPEN_QUESTIONS_CSV", CONFIG_DIR / "BO_Configuration_Sheet_v3 - Open Questions (Sec 9).csv")
)
# These 5 constants are defined but never read by load_config() (see below) or anything
# else in this file -- CONFIG_ALL_PARAMS_CSV/CONFIG_SE_INCENTIVE_CSV are the only two
# actually loaded. Kept in sync with the current pitch_config-style re-export anyway
# (bumped (1)->(2) 2026-08-12, same day and same reason as CONFIG_ALL_PARAMS_CSV's
# (2)->(3) bump above) so they don't point at a guaranteed-404 path if ever wired up.
CONFIG_AGENT_DETERMINED_CSV = Path(
    os.environ.get("SE_AGENT_AGENT_DETERMINED_CSV", CONFIG_DIR / "BO_Configuration_Sheet_v3 - Agent-Determined Parameters (2).csv")
)
CONFIG_TASK_FORMULA_CSV = Path(
    os.environ.get("SE_AGENT_TASK_FORMULA_CSV", CONFIG_DIR / "BO_Configuration_Sheet_v3 - Daily Task Assignment Formula (2).csv")
)
CONFIG_GUARDRAILS_CSV = Path(
    os.environ.get("SE_AGENT_GUARDRAILS_CSV", CONFIG_DIR / "BO_Configuration_Sheet_v3 - Guardrails (2).csv")
)
CONFIG_VISIT_PURPOSE_MAPPING_CSV = Path(
    os.environ.get("SE_AGENT_VISIT_PURPOSE_MAPPING_CSV", CONFIG_DIR / "BO_Configuration_Sheet_v3 - Visit Type & Purpose Mapping (2).csv")
)
CONFIG_VISIT_PURPOSE_SYSTEM_CSV = Path(
    os.environ.get("SE_AGENT_VISIT_PURPOSE_SYSTEM_CSV", CONFIG_DIR / "BO_Configuration_Sheet_v3 - Visit Type to Purpose (System) (2).csv")
)
AOP_CSV = Path(
    os.environ.get("SE_AGENT_AOP_CSV", BASE_DIR / "Niyojan Q2-FY_26_27 Dashboard - Planning.csv")
)

NULL_PLACEHOLDERS = {"", "(blank)", "-", "#N/A", "#DIV/0!"}
LOOKBACK_DAYS = int(os.environ.get("SE_AGENT_LOOKBACK_DAYS", "90"))  # Config 0.2
# Guards against a single slow/stuck Redshift query hanging the whole pipeline forever
# (observed 2026-08-09: load_live_sources sat >20 min on an ESTABLISHED connection with
# no query-level timeout and no per-query logging, making the hang undiagnosable). Redshift
# enforces this server-side per statement, so a stuck query now fails fast with a clear
# psycopg2 error instead of hanging -- load_live_sources already quarantines per-query
# failures (Live_Pull_Failed) rather than crashing the run.
REDSHIFT_STATEMENT_TIMEOUT_MS = int(os.environ.get("SE_AGENT_REDSHIFT_STATEMENT_TIMEOUT_MS", "120000"))
# Shared backoff schedule for both connecting AND querying Redshift (RedshiftDirectClient._connect
# and .execute_sql) -- a query that dies mid-run deserves the same multi-attempt resilience as one
# that fails to connect in the first place, not a single unretried reconnect (see 2026-08-09 fix).
REDSHIFT_RETRY_BACKOFF_SECONDS = [2, 6]

# Mirrors planning.models.DCVisitStreak.ESCALATION_THRESHOLD -- duplicated, not
# imported, because this file must stay usable standalone (bare `python
# se_daily_plan_agent.py`, no Django) while DCVisitStreak is a Django model this file
# can't import. Used only for generate_se_daily_plan()'s Critical flag (confirmed
# 2026-08-18); keep the two values in sync if the threshold ever changes.
DC_VISIT_ESCALATION_THRESHOLD = 3


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# =====================================================================================
# 1. METABASE CLIENT -- real-time sync for Sources 1 / 3 / 4
# =====================================================================================

class MetabaseNotConfigured(RuntimeError):
    """Raised when a live source is requested but METABASE_URL/METABASE_API_KEY are unset."""


class MetabaseClient:
    """Thin wrapper over the Metabase REST API (POST /api/dataset, native query).

    This is the standalone-script equivalent of the live Metabase access already used to
    confirm database ids 41 (Redshift), 31 (input-backend) and 4 (kheti) against this
    instance. Auth is an API key (Metabase Admin > API Keys), passed via the
    x-api-key header.
    """

    def __init__(self, base_url: Optional[str] = None, api_key: Optional[str] = None):
        self.base_url = (base_url or os.environ.get("METABASE_URL", "")).rstrip("/")
        self.api_key = api_key or os.environ.get("METABASE_API_KEY", "")

    @property
    def configured(self) -> bool:
        return bool(self.base_url and self.api_key and requests is not None)

    def execute_sql(self, database_id: int, sql: str) -> Table:
        if not self.configured:
            reason = "requests not installed" if requests is None else "METABASE_URL/METABASE_API_KEY not set"
            raise MetabaseNotConfigured(f"Metabase source unavailable ({reason})")
        resp = requests.post(
            f"{self.base_url}/api/dataset",
            headers={"x-api-key": self.api_key, "Content-Type": "application/json"},
            json={"type": "native", "native": {"query": sql}, "database": database_id},
            timeout=180,
        )
        resp.raise_for_status()
        payload = resp.json()
        data = payload.get("data", {})
        cols = [c["name"] for c in data.get("cols", [])]
        return [dict(zip(cols, row)) for row in data.get("rows", [])]

    def close(self) -> None:
        """No persistent connection to release -- each execute_sql() is a standalone
        HTTP request. No-op kept so callers can treat both client types identically."""
        pass


class RedshiftDirectClient:
    """Direct psycopg2 connection to the Redshift cluster behind Metabase's 'Redshift'
    (db_id 41) and 'input-backend' (db_id 31) sources -- bypasses Metabase's REST/MCP API
    entirely. Confirmed live 2026-08-04: this one cluster hosts 3 databases --
    "dev" (pathik_report, input_partner_details, input_se_node_mapping, coupon_analysis,
    dc_mapping_club_scheme, dc_club_slabs, hyperlocal_order -- i.e. everything under
    Metabase db_id 41), "input_backend_db" (task_management_*, users_user, sale_orderrequest,
    customer_management_customer, attendance_attendance, payments_paymenttransaction --
    i.e. Metabase db_id 31's tables), and "locus" (unrelated, not used here).

    KNOWN GAP, confirmed live: customer_management_input_outstanding does NOT exist on
    this cluster in either database -- it appears to be Postgres-only (reachable via
    Metabase db_id 31, not this Redshift replica). Queries against it will raise; callers
    MUST catch that separately from other Source 3d queries (orders/credit_on_hold still
    work fine here) rather than losing both to one try/except.

    Also confirmed live: this cluster's Postgres dialect does NOT support `DISTINCT ON`
    (raises FeatureNotSupported) -- use ROW_NUMBER() OVER (PARTITION BY ... ORDER BY ...)
    instead, same as any query written against it elsewhere in this codebase.

    This is a warehouse replica of the OLTP input-backend Postgres DB, not the OLTP DB
    itself -- treat it as possibly lagged by some sync interval, not guaranteed
    instant-real-time, even though it's queried live on every call.

    Connection reuse: one connection is kept open per database (dev / input_backend_db)
    and reused across every execute_sql() call on this instance, instead of opening and
    closing a fresh connection per query. A single request (run_pipeline() or
    generate_plan_for_scope()) fires 13-16+ sequential queries through one client
    instance (get_client() constructs a new one per top-level call, so there's no
    cross-request state to worry about) -- per-query connection setup was real,
    measured overhead against this cluster, not opened lazily/pooled before. Call
    close() when done with this client (both call sites do, in a finally block).
    """

    def __init__(self):
        self.host = os.environ.get("REDSHIFT_HOST", "")
        self.port = os.environ.get("REDSHIFT_PORT", "5439")
        self.user = os.environ.get("REDSHIFT_USER", "")
        self.password = os.environ.get("REDSHIFT_PASSWORD", "")
        self.db_dev = os.environ.get("REDSHIFT_DB_DEV", "dev")
        self.db_input_backend = os.environ.get("REDSHIFT_DB_INPUT_BACKEND", "input_backend_db")
        self._connections: Dict[str, Any] = {}  # dbname -> open psycopg2 connection

    @property
    def configured(self) -> bool:
        return bool(self.host and self.user and self.password)

    def _db_name_for(self, database_id: int) -> str:
        if database_id == REDSHIFT_DB_ID:
            return self.db_dev
        if database_id in (INPUT_BACKEND_DB_ID, KHETI_DB_ID):
            # KHETI_DB_ID (hyperlocal_order) is also present under "dev" on this cluster;
            # input-backend tables live under input_backend_db.
            return self.db_input_backend if database_id == INPUT_BACKEND_DB_ID else self.db_dev
        raise ValueError(f"No known Redshift database mapping for Metabase db_id {database_id}")

    def _connect(self, dbname: str):
        import psycopg2

        # Retry only the connection step, with backoff, and only on connection-level
        # failures (OperationalError -- host unreachable, timeout, refused) -- a flaky
        # network shouldn't kill a whole run, but a bad query (ProgrammingError etc.)
        # retrying would just fail the same way 3x slower, so those propagate immediately.
        last_error: Optional[Exception] = None
        for attempt, delay in enumerate([0] + REDSHIFT_RETRY_BACKOFF_SECONDS):
            if delay:
                logger.warning("Redshift connect failed (attempt %d), retrying in %ds: %s", attempt, delay, last_error)
                time.sleep(delay)
            try:
                conn = psycopg2.connect(
                    host=self.host, port=self.port, user=self.user, password=self.password,
                    dbname=dbname, connect_timeout=15,
                    # Observed 2026-08-09: queries that sit longer than ~5 minutes waiting on
                    # results (client socket idle, server still working) get silently dropped
                    # with "SSL connection has been closed unexpectedly" -- failures landed at
                    # 309s/367s, right around AWS NAT Gateway's default 350s idle-connection
                    # timeout. No keepalives were being sent, so the gateway saw a dead-looking
                    # connection and killed it mid-query. TCP keepalives every 10s (starting
                    # after 20s idle) keep the mapping alive through long-running queries.
                    keepalives=1, keepalives_idle=20, keepalives_interval=10, keepalives_count=6,
                )
                # connect_timeout only bounds the TCP handshake -- once connected, a slow
                # or stuck query has no client-side limit at all. Set it server-side so a
                # bad/slow query fails with a clear OperationalError instead of hanging the
                # whole pipeline indefinitely (see REDSHIFT_STATEMENT_TIMEOUT_MS).
                with conn.cursor() as cur:
                    cur.execute(f"SET statement_timeout = {REDSHIFT_STATEMENT_TIMEOUT_MS}")
                conn.commit()
                return conn
            except psycopg2.OperationalError as e:
                last_error = e
        raise last_error

    def _get_connection(self, dbname: str):
        conn = self._connections.get(dbname)
        if conn is not None and conn.closed == 0:
            return conn
        conn = self._connect(dbname)
        self._connections[dbname] = conn
        return conn

    def execute_sql(self, database_id: int, sql: str) -> Table:
        if not self.configured:
            raise MetabaseNotConfigured("REDSHIFT_HOST/REDSHIFT_USER/REDSHIFT_PASSWORD not set")
        import psycopg2

        dbname = self._db_name_for(database_id)
        # Same backoff schedule as _connect() -- a query that dies mid-run (idle-connection
        # drop, transient server hiccup) previously got exactly one immediate, unretried
        # reconnect attempt, which 2026-08-09's live failures showed isn't enough when the
        # underlying condition (NAT idle-timeout-adjacent drops) hits more than once in a
        # row. A bad query (ProgrammingError etc.) is not an OperationalError and still
        # propagates immediately, unretried -- only connection-level failures loop here.
        last_error: Optional[Exception] = None
        for attempt, delay in enumerate([0] + REDSHIFT_RETRY_BACKOFF_SECONDS):
            if delay:
                logger.warning("Redshift query on %s failed (attempt %d), retrying in %ds: %s", dbname, attempt, delay, last_error)
                time.sleep(delay)
            conn = self._get_connection(dbname)
            try:
                cur = conn.cursor()
                cur.execute(sql)
                cols = [d[0] for d in cur.description]
                return [dict(zip(cols, row)) for row in cur.fetchall()]
            except psycopg2.errors.QueryCanceled:
                # QueryCanceled is a subclass of OperationalError, but it means the query
                # hit REDSHIFT_STATEMENT_TIMEOUT_MS -- not a dead connection. Retrying just
                # reruns the same slow query up to 3x, turning a 120s timeout into a ~6min
                # stall and defeating the whole point of the 2026-08-09 fail-fast fix. Roll
                # back (the transaction is aborted) and propagate immediately, unretried.
                conn.rollback()
                raise
            except psycopg2.OperationalError as e:
                last_error = e
                logger.warning("Redshift connection to %s died mid-run", dbname)
                self._connections.pop(dbname, None)
            except Exception:
                # Any other failure (bad SQL, permission denied, etc.) leaves this
                # connection's transaction aborted -- every later query reusing the same
                # connection (see _get_connection) would otherwise fail with
                # InFailedSqlTransaction even though nothing is wrong with it
                # individually. Confirmed live 2026-08-12: a single permission-denied
                # query on dc_datamart took 3 unrelated, working queries down with it on
                # every run since. Roll back so the connection is usable again, then
                # propagate this query's own real error unretried, same as before.
                conn.rollback()
                raise
        raise last_error

    def close(self) -> None:
        for conn in self._connections.values():
            try:
                conn.close()
            except Exception:
                pass
        self._connections.clear()


def get_client():
    """Prefer a direct Redshift connection (REDSHIFT_HOST/USER/PASSWORD) over the Metabase
    REST API (METABASE_URL/METABASE_API_KEY) when both could apply -- direct access needs
    no Metabase API key and was confirmed working end-to-end 2026-08-04. Falls back to
    MetabaseClient so existing METABASE_URL/METABASE_API_KEY setups keep working unchanged."""
    redshift_client = RedshiftDirectClient()
    if redshift_client.configured:
        return redshift_client
    return MetabaseClient()


# =====================================================================================
# 1b. PRODUCT COHORT CLIENT -- Focus Product Campaign Targeting (Product _cohort/,
# Postman collection "Product Cohort -- -sps-v1-fp", saas-platform-service.api.dehat.net)
# =====================================================================================
# A DIFFERENT, real capability from the DC Card's "Crop Type/Style" gap (see the pitch_
# config Focus Product Targeting CSV) -- product-first, not DC-first: given a Focus
# Product (materialId) and a Node, produces (2A) its historical weekly sales pattern,
# (2B) an LLM-read crop-seasonality (buildup/peak/closure weeks), and (3) a target DC
# cohort to push that product to.
#
# Auth (Postman collection's own "01 . Auth" folder, PRODUCT_COHORT_AUTH.md): a Go Admin
# sign-in (email + password) against {base_url}/admin/signin produces an X-FP-Session
# value and a go_admin_session cookie, which every other call needs. This agent does NOT
# perform that sign-in and never will, regardless of who supplies the credentials or how
# the request is framed -- entering a password to authenticate on someone's behalf is
# outside what this agent does, full stop (same boundary already documented in the pitch
# config CSV's "Status" row for this feature). What this client DOES do: accept an
# already-issued session/cookie pair via environment variables, so a human who ran the
# Postman "Sign in" request themselves (or whatever internal login flow issues the same
# X-FP-Session token) can hand this pipeline a working session without ever handing this
# agent a password. Until PRODUCT_COHORT_SESSION/PRODUCT_COHORT_GO_ADMIN_SESSION are set,
# every method below raises ProductCohortNotConfigured -- same fail-loud convention as
# MetabaseNotConfigured, never a silent empty result.
#
# Response schemas for Step 2B and Step 3 were EMPTY in the saved Postman collection (no
# example response was ever captured) -- this client still returns their raw parsed JSON
# unmodified rather than reshaping it into a typed structure. Step 3's shape IS now
# confirmed live (2026-08-14): {"status": "Success", "results": {"materialId", "nodeId",
# "totalDCs", "dcs": [{"partnerId" (matches this codebase's DC_ID format), "name",
# "state", "district", "phone", "reasons": [...], "relatedProductsBought": [...],
# "relatedProductsBreakdown": [{"productName", "qty", "value"}, ...], "totalQty",
# "totalValue"}, ...]}}. Step 2B's response is still genuinely unconfirmed -- callers
# should treat only that one as opaque until a real response is seen.

PRODUCT_COHORT_URL = os.environ.get("PRODUCT_COHORT_URL", "https://saas-platform-service.api.dehaat.net")


class ProductCohortNotConfigured(RuntimeError):
    """Raised when a Product Cohort API call is attempted but PRODUCT_COHORT_SESSION /
    PRODUCT_COHORT_GO_ADMIN_SESSION are unset -- this agent never signs in on its own to
    obtain them (see the section docstring above)."""


class ProductCohortClient:
    """Thin wrapper over the confirmed /sps/v1/fp/focus-product/* endpoints. Every method
    maps 1:1 to one request in the Postman collection; none of them guess at a payload
    shape beyond what that collection's saved sample bodies show."""

    # Same backoff schedule as Redshift's retry (REDSHIFT_RETRY_BACKOFF_SECONDS) -- this
    # is the newest live external dependency in the pipeline and, unlike every other live
    # client here, previously had zero resilience to a transient network blip: one failed
    # call meant the whole Step (2A/2B/3) failed for that run.
    RETRY_BACKOFF_SECONDS = [2, 6]

    def __init__(self, base_url: Optional[str] = None, session_token: Optional[str] = None, go_admin_session: Optional[str] = None):
        self.base_url = (base_url or PRODUCT_COHORT_URL).rstrip("/")
        self.session_token = session_token or os.environ.get("PRODUCT_COHORT_SESSION", "")
        self.go_admin_session = go_admin_session or os.environ.get("PRODUCT_COHORT_GO_ADMIN_SESSION", "")
        # Reused across every call this client instance makes (Steps 2A/2B/3 in one
        # get_focus_product_campaign_targets() run) -- pools the TCP/TLS connection
        # instead of paying full handshake cost on each of up to 4 sequential requests.
        self._session = requests.Session() if requests is not None else None

    @property
    def configured(self) -> bool:
        return bool(self.base_url and self.session_token and self.go_admin_session and requests is not None)

    def _request(self, method: str, path: str, params: Optional[Dict[str, Any]] = None, json_body: Optional[Dict[str, Any]] = None) -> Any:
        if not self.configured:
            reason = "requests not installed" if requests is None else "PRODUCT_COHORT_SESSION/PRODUCT_COHORT_GO_ADMIN_SESSION not set"
            raise ProductCohortNotConfigured(f"Product Cohort source unavailable ({reason}) -- see Product _cohort/PRODUCT_COHORT_AUTH.md for how to obtain a session yourself; this agent will not sign in on your behalf")
        url = f"{self.base_url}{path}"
        headers = {
            "Content-Type": "application/json",
            "X-FP-Session": self.session_token,
            "Cookie": f"go_admin_session={self.go_admin_session}",
        }
        last_error: Optional[Exception] = None
        for attempt, delay in enumerate([0] + self.RETRY_BACKOFF_SECONDS):
            if delay:
                logger.warning("Product Cohort request to %s failed (attempt %d), retrying in %ds: %s", path, attempt, delay, last_error)
                time.sleep(delay)
            try:
                resp = self._session.request(method, url, headers=headers, params=params, json=json_body, timeout=60)
                resp.raise_for_status()
                return resp.json()
            except requests.exceptions.HTTPError as e:
                # A 4xx (bad/expired session, bad request) will fail identically on
                # retry -- only a 5xx (server-side, plausibly transient) is worth it.
                if e.response is not None and e.response.status_code < 500:
                    raise
                last_error = e
            except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
                last_error = e
        raise last_error

    # -- Step 1: filter options --------------------------------------------------------
    def get_options(self) -> Any:
        """States + categories (Postman: 'Options root')."""
        return self._request("GET", "/sps/v1/fp/focus-product/options")

    def get_nodes(self, state_id: str) -> Any:
        return self._request("GET", "/sps/v1/fp/focus-product/options/nodes", params={"stateId": state_id})

    def get_sub_categories(self, category_id: str) -> Any:
        return self._request("GET", "/sps/v1/fp/focus-product/options/sub-categories", params={"categoryId": category_id})

    def get_products(self, category_id: str, sub_category_id: Optional[str] = None, search: Optional[str] = None) -> Any:
        params: Dict[str, Any] = {"categoryId": category_id}
        if sub_category_id is not None:
            params["subCategoryId"] = sub_category_id
        if search is not None:
            params["search"] = search
        return self._request("GET", "/sps/v1/fp/focus-product/options/products", params=params)

    def save_filter_selection(self, selection: Dict[str, Any]) -> Any:
        """selection shape per the Postman sample: stateId, nodeId, categoryId,
        categoryName, subCategoryId, subCategoryName, materialId, productName."""
        return self._request("POST", "/sps/v1/fp/focus-product/filter-selections", json_body=selection)

    def list_filter_selections(self) -> Any:
        return self._request("GET", "/sps/v1/fp/focus-product/filter-selections")

    # -- Step 2A: historical purchase pattern --------------------------------------------
    def get_ib_weekly_sales(self, material_id: str, node_id: str, years: int = 4) -> Any:
        return self._request("GET", "/sps/v1/fp/focus-product/analysis/ib-weekly-sales", params={"materialId": material_id, "nodeId": node_id, "years": years})

    def get_ib_raw_records(self, material_id: str, node_id: str, years: int = 4) -> Any:
        return self._request("GET", "/sps/v1/fp/focus-product/analysis/ib-raw-records", params={"materialId": material_id, "nodeId": node_id, "years": years})

    # -- Step 2B: LLM-inferred crop-seasonality read -------------------------------------
    def step2b_crop_seasonality(
        self, node_id: str, material_id: str,
        outer_from_week: int, outer_to_week: int,
        buildup_from_week: int, buildup_to_week: int,
        peak_week: int, closure_from_week: int, closure_to_week: int,
    ) -> Any:
        """Field names/ordering match the Postman sample body exactly. NOTE (genuinely
        unconfirmed, flag rather than assume): the sample body has buildup/peak/closure
        weeks as INPUT even though the collection describes Step 2B as producing an
        'LLM-inferred' read of those same three windows -- whether they're a prior the
        LLM refines, or the caller is expected to already know them, isn't answered
        anywhere in the saved collection (response schema is empty). Callers should treat
        whatever they pass here as a best-guess seed, not a confirmed methodology."""
        body = {
            "nodeId": node_id, "materialId": material_id,
            "outerFromWeek": outer_from_week, "outerToWeek": outer_to_week,
            "buildupFromWeek": buildup_from_week, "buildupToWeek": buildup_to_week,
            "peakWeek": peak_week,
            "closureFromWeek": closure_from_week, "closureToWeek": closure_to_week,
        }
        return self._request("POST", "/sps/v1/fp/focus-product/analysis/step2b", json_body=body)

    # -- Step 3: target DC cohort --------------------------------------------------------
    def step3_dc_cohort(self, material_id: str, node_id: str, crop_districts: List[str], related_product_names: List[str]) -> Any:
        """crop_districts/related_product_names: same genuinely-unconfirmed situation as
        Step 2B's seed weeks -- the Postman sample hardcodes them rather than showing
        where they come from (Step 2A/2B output? a separate lookup?). Passed through
        as-given, not derived here."""
        body = {"materialId": material_id, "nodeId": node_id, "cropDistricts": crop_districts, "relatedProductNames": related_product_names}
        return self._request("POST", "/sps/v1/fp/focus-product/analysis/step3-cohort", json_body=body)


# =====================================================================================
# 1c. DC FARMER MAPPING CLIENT -- per-DC product recommendations
# (saas-platform-service-farmer-dc-mapping.api.dehat.co, confirmed live 2026-09-02)
# =====================================================================================
# A SEPARATE service from the Product Cohort client above -- different domain (dehat.co,
# single-a, not dehaat.net), separate Go Admin login, separate session/cookie jar. Given a
# DC (dcId + sapPartnerId) and a radius, returns a ranked list of recommended products for
# that DC with per-product recommendation evidence (tags: past_purchase, geo_peers,
# disease_scan, crop_season -- each carries its own supporting sales/context data).
#
# Auth: same Go Admin pattern as Product Cohort (X-FP-Session header + go_admin_session
# cookie), but this agent does NOT sign in on its own here either, for the identical reason
# documented on ProductCohortClient above -- entering a password on someone's behalf is
# outside what this agent does, full stop. A human authenticates at
# {base_url}/dc-farmer-mapping themselves and hands this client an already-issued session
# via DC_FARMER_MAPPING_SESSION/DC_FARMER_MAPPING_GO_ADMIN_SESSION.

DC_FARMER_MAPPING_URL = os.environ.get("DC_FARMER_MAPPING_URL", "https://saas-platform-service-farmer-dc-mapping.api.dehat.co")


class DCFarmerMappingNotConfigured(RuntimeError):
    """Raised when a DC Farmer Mapping API call is attempted but DC_FARMER_MAPPING_SESSION /
    DC_FARMER_MAPPING_GO_ADMIN_SESSION are unset -- this agent never signs in on its own to
    obtain them (see the section docstring above)."""


class DCFarmerMappingClient:
    """Thin wrapper over the confirmed /sps/v1/fp/dc-product-recommendations endpoint."""

    # Same backoff schedule as ProductCohortClient/Redshift's retry.
    RETRY_BACKOFF_SECONDS = [2, 6]

    def __init__(self, base_url: Optional[str] = None, session_token: Optional[str] = None, go_admin_session: Optional[str] = None):
        self.base_url = (base_url or DC_FARMER_MAPPING_URL).rstrip("/")
        self.session_token = session_token or os.environ.get("DC_FARMER_MAPPING_SESSION", "")
        self.go_admin_session = go_admin_session or os.environ.get("DC_FARMER_MAPPING_GO_ADMIN_SESSION", "")
        self._session = requests.Session() if requests is not None else None

    @property
    def configured(self) -> bool:
        return bool(self.base_url and self.session_token and self.go_admin_session and requests is not None)

    def _request(self, method: str, path: str, params: Optional[Dict[str, Any]] = None, json_body: Optional[Dict[str, Any]] = None) -> Any:
        if not self.configured:
            reason = "requests not installed" if requests is None else "DC_FARMER_MAPPING_SESSION/DC_FARMER_MAPPING_GO_ADMIN_SESSION not set"
            raise DCFarmerMappingNotConfigured(f"DC Farmer Mapping source unavailable ({reason}) -- authenticate at {DC_FARMER_MAPPING_URL}/dc-farmer-mapping yourself and set the env vars; this agent will not sign in on your behalf")
        url = f"{self.base_url}{path}"
        headers = {
            "Content-Type": "application/json",
            "X-FP-Session": self.session_token,
            "Cookie": f"go_admin_session={self.go_admin_session}",
            # This host's WAF silently stalls (no response, no error) requests whose
            # User-Agent doesn't look like a browser -- confirmed live 2026-09-02: the
            # default `python-requests/x.x` UA hung for 60s+ on every attempt, while an
            # identical request with a browser UA succeeded in ~13s. Referer included to
            # match the real browser request this was reverse-engineered from.
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/151.0.0.0 Safari/537.36",
            "Referer": f"{self.base_url}/dc-farmer-mapping",
        }
        last_error: Optional[Exception] = None
        for attempt, delay in enumerate([0] + self.RETRY_BACKOFF_SECONDS):
            if delay:
                logger.warning("DC Farmer Mapping request to %s failed (attempt %d), retrying in %ds: %s", path, attempt, delay, last_error)
                time.sleep(delay)
            try:
                resp = self._session.request(method, url, headers=headers, params=params, json=json_body, timeout=60)
                resp.raise_for_status()
                return resp.json()
            except requests.exceptions.HTTPError as e:
                # A 4xx (bad/expired session, bad request) will fail identically on
                # retry -- only a 5xx (server-side, plausibly transient) is worth it.
                if e.response is not None and e.response.status_code < 500:
                    raise
                last_error = e
            except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
                last_error = e
        raise last_error

    def get_dc_product_recommendations(self, dc_id: str, sap_partner_id: str, radius_km: float = 7) -> Any:
        """Confirmed live response shape (2026-09-02): {"status": "Success", "results":
        {"dc": {...}, "radiusKm", "preferDehaat", "products": [{"materialId",
        "productName", "productCategory", "productSubCategory", "technicalContent",
        "tags": [...], "context": {<tag>: {...evidence...}}}, ...]}}."""
        return self._request("GET", "/sps/v1/fp/dc-product-recommendations", params={"dcId": dc_id, "radiusKm": radius_km, "sapPartnerId": sap_partner_id})


# =====================================================================================
# 2. SHARED NORMALIZATION HELPERS (Section 3 / 4 of the doc)
# =====================================================================================

def clean_null(value: Any) -> Any:
    """Normalize the DC Master's four null placeholders (Section 4) to a true None."""
    if isinstance(value, str) and value.strip() in NULL_PLACEHOLDERS:
        return None
    return value


def parse_number(value: Any) -> Optional[float]:
    """Strip thousands-separator commas before casting to numeric (Section 4)."""
    value = clean_null(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("%", "")
    if text in NULL_PLACEHOLDERS or text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_mobile(value: Any) -> Optional[str]:
    """Strip whitespace/country-code prefix so 1a<->1b can join on mobile_number."""
    value = clean_null(value)
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value))
    if len(digits) > 10:
        digits = digits[-10:]  # drop country-code prefix
    return digits or None


def normalize_id(value: Any) -> Optional[str]:
    """Bug fixed 2026-08-25: this used to .lstrip("0") every ID, which collides two
    genuinely different IDs whenever one happens to be zero-padded (normalize_id("0012345")
    == normalize_id("12345")) -- confirmed live that every real DC_ID in this system is a
    fixed 10-digit string starting with "1" (never "0"), so stripping leading zeros was
    pure defensive cruft against a case that doesn't occur here, while creating a real
    collision risk. No longer strips anything from the string form.

    Also fixed: a float/Decimal-typed ID (e.g. 1000016754.0, from a DB driver that returns
    a numeric column as a native number rather than a string) used to stringify with a
    trailing ".0" that a plain string ID like "1000016754" would never carry -- silently
    breaking every cross-source join on that ID with no exception raised on either side.
    Whole-number floats/Decimals are cast through int() first so both representations
    converge on the same string. A genuinely fractional numeric ID (which should never
    happen for an ID field) is left as str(value) rather than silently truncated."""
    value = clean_null(value)
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal) and value == value.to_integral_value():
        return str(int(value))
    text = str(value).strip()
    return text or None


def standardize_date(value: Any) -> Optional[str]:
    """Standardize to YYYY-MM-DD (Section 3). Keeps distinct date fields distinct upstream.

    Known, documented limitation (made explicit 2026-08-25, not fixed -- see below for
    why): a bare "DD-MM-YYYY"-shaped string is genuinely ambiguous with "MM-DD-YYYY"
    whenever both components are <=12 (e.g. "01-02-2026" could mean 1-Feb or 2-Jan), and
    this function always resolves that ambiguity as DD-MM-YYYY with no exception raised
    either way -- silently, since this is a low-level utility with no Exceptions object
    threaded through it to flag against. This is a deliberate, confirmed convention
    choice, not a guess: every live SQL source in this pipeline returns real datetime/date
    objects (handled by the isinstance() branch below, never hits the string-guessing
    path at all), so the ambiguous branch only fires for the local CSV sources (DC_RAnk,
    Config, AOP), which are India-only business exports -- DD-MM-YYYY is the correct,
    confirmed regional convention there (consistent with this codebase's own April-start
    Indian fiscal year elsewhere, e.g. _fiscal_year_start()). If a source ever legitimately
    used MM-DD-YYYY, every date where day<=12 would silently transpose with zero signal --
    flagging that live would need an Exceptions object threaded into every one of this
    function's ~15 call sites, a larger structural change than the ambiguity itself
    warrants without evidence any caller actually feeds it a non-Indian-convention date."""
    value = clean_null(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[: len(fmt) + 2], fmt).date().isoformat()
        except ValueError:
            continue
    return text[:10] if len(text) >= 10 else text


ZERO_COORD_EPSILON_DEGREES = 1e-3  # ~111m at the equator

def is_zero_coord(lat: Any, long: Any) -> bool:
    """(0, 0) means 'not yet checked out', never a real location (Section 4).

    Bug fixed 2026-08-25: the epsilon was 1e-9, which only catches an EXACT (0, 0) --
    real device GPS chips are known to emit near-zero-but-not-exactly-zero coordinates
    during a cold-start/no-fix state (a common default before a real fix is acquired),
    which slipped straight past this check and got treated as a real, valid location.
    Widened to ZERO_COORD_EPSILON_DEGREES (~111m) -- safe for this system specifically
    because every real DC/SE in this network operates inside India (roughly 8-37N,
    68-97E), nowhere near true (0, 0) (a point in the Gulf of Guinea); no real field
    coordinate could ever legitimately fall this close to Null Island, so anything this
    close is definitionally a GPS glitch here, not a real nearby location that a tighter
    epsilon would need to protect."""
    lat_f, long_f = parse_number(lat), parse_number(long)
    return lat_f is not None and long_f is not None and abs(lat_f) < ZERO_COORD_EPSILON_DEGREES and abs(long_f) < ZERO_COORD_EPSILON_DEGREES


def photo_logged_from_task_details(task_details_json: Any) -> Optional[bool]:
    """Photo_Logged must be derived from the task_details image array, never a flag column.

    Bug fixed 2026-08-25: `blob.get("images") or blob.get("image")` treated a genuinely
    present-but-empty "images": [] as falsy and silently fell through to check a second
    key -- a task_details blob confirmed to have zero images could still read Photo_Logged
    from an unrelated "image" key if one happened to exist too. Key presence is now
    checked explicitly (`"images" in blob`), not truthiness, so an explicit empty array is
    honored as "no photo, confirmed," not overridden."""
    if task_details_json is None:
        return None
    try:
        blob = json.loads(task_details_json) if isinstance(task_details_json, str) else task_details_json
    except (json.JSONDecodeError, TypeError):
        return None
    if not isinstance(blob, dict):
        return None
    if "images" in blob:
        images = blob["images"]
    elif "image" in blob:
        images = blob["image"]
    else:
        images = None
        for key in blob:
            if isinstance(blob[key], list) and "image" in key.lower():
                images = blob[key]
                break
    return bool(images)


class Exceptions:
    """Accumulates the Exceptions_Report table (Section 7): Record_ID, Source, Reason_Code,
    Detail, Run_Timestamp. Every check runs and logs pass/fail even on a clean run (Section 5)."""

    def __init__(self, run_timestamp: str):
        self.run_timestamp = run_timestamp
        self.rows: Table = []
        self.check_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: {"pass": 0, "fail": 0})

    def flag(self, record_id: Any, source: str, reason_code: str, detail: str) -> None:
        self.rows.append(
            {
                "Record_ID": record_id,
                "Source": source,
                "Reason_Code": reason_code,
                "Detail": detail,
                "Run_Timestamp": self.run_timestamp,
            }
        )
        self.check_counts[reason_code]["fail"] += 1

    def ok(self, reason_code: str) -> None:
        self.check_counts[reason_code]["pass"] += 1

    def summary(self) -> Dict[str, Dict[str, int]]:
        return dict(self.check_counts)


# =====================================================================================
# 3. SOURCE 5 -- CONFIG MASTER (local, fully resolved)
# =====================================================================================

@dataclass
class BusinessConstants:
    """Numeric constants actually needed by scoring/planning code below, mirroring the
    Configured Values read live from BO_Configuration_Sheet_v3 as of this build. Every
    business rule still *lives* in Source 5 (Config_Normalized) per the Section 8
    guardrail -- these are a typed, code-usable projection of it, not a replacement for it.
    Re-derive from Config_Normalized if the sheet changes."""

    # Section 1 (BO1) grade bands, % of PL_Expected
    bo1_grade_a: float = 0.80
    bo1_grade_b: float = 0.60
    bo1_grade_c: float = 0.40
    # Section 2 (BO2)
    bo2_valid_visit_min_minutes: float = 10.0
    bo2_coverage_grade_a: float = 1.0
    bo2_coverage_grade_b: float = 0.7
    bo2_coverage_grade_c: float = 0.4
    bo2_priority_grade_a: float = 1.5
    bo2_priority_grade_b: float = 1.0
    bo2_priority_grade_c: float = 0.5
    # Section 3 (BO3) aging-bucket risk weights
    bo3_aging_weights: Dict[str, float] = field(
        default_factory=lambda: {"0_30": 0.2, "31_60": 0.5, "61_90": 0.8, "90_plus": 1.0}
    )
    bo3_grade_a: float = 1.00
    bo3_grade_b: float = 0.75
    bo3_grade_c: float = 0.50
    bo3_ratio_cap_low: float = 0.50
    bo3_ratio_cap_high: float = 1.50
    # Section 5 (BO5)
    bo5_mega_meeting_min_farmers: int = 50
    bo5_regular_meeting_min_farmers: int = 10
    bo5_meeting_target_per_month: int = 2
    bo5_onboarding_target_per_month: int = 2
    bo5_weight_meeting: float = 0.50
    bo5_weight_onboarding: float = 0.50
    # Grade cutoffs -- Source 5 never defines these for BO5 at all; reusing BO1's bands
    # for directional consistency only (see score_bo5_long_term() docstring).
    bo5_grade_a: float = 0.80
    bo5_grade_b: float = 0.60
    bo5_grade_c: float = 0.40
    # 8.11 Layer 0 (FM_Urgency) -- Pacing_Buffer_Days. NOT a guessed default like the
    # grade-cutoff reuses above -- confirmed directly by the user 2026-08-06 ("at least
    # one [meeting] in 15 days"). See compute_fm_urgency().
    pacing_buffer_days: int = 15
    # Section 6 -- DC exclusion (6.2, 6.3, 6.4 confirmed defaults)
    min_days_since_last_visit: int = 5
    seasonal_skip_enabled: bool = False
    # User-confirmed 2026-08-13, DISABLED for eligibility 2026-09-04 (see apply_dc_
    # exclusion_rules docstring -- the Top-DC-list is now the sole eligibility gate,
    # fail-closed on load failure, no Rank fallback). Kept defined only in case of
    # reintroduction. DC_RAnk.csv's Rank/Cohort data is still used elsewhere (Step 5
    # selection ordering) -- only this eligibility role is gone.
    max_eligible_rank: float = 6000.0
    # Section 7 -- daily ranking weights (top-3 objectives)
    rank1_weight: float = 0.40
    rank2_weight: float = 0.35
    rank3_weight: float = 0.25
    # Section 8 -- capacity
    max_objectives_per_day: int = 3
    max_long_cycle_per_day: int = 1
    field_minutes_per_day: int = 7 * 60
    call_minutes_per_day: int = 60
    call_duration_min: int = 10
    visit_duration_min: int = 45
    visit_duration_max: int = 60
    meeting_duration_min: int = 180
    onboarding_duration_min: int = 120
    daily_cap_visits: int = 6
    daily_cap_outstanding: int = 4
    daily_cap_pl: int = 5
    weekly_cap_long_term: int = 2
    daily_travel_cap_km: float = 80.0
    monthly_travel_cap_km: float = 1600.0  # 8.9
    total_capacity_min: int = 480  # 8.2 confirmed: 60 min calls + 420 min field = 480 (8hr day)
    # 8.6 confirmed Win_Definition per objective (Outstanding = payment received ONLY,
    # promise-to-pay removed as a qualifying win in the v3 configured value)
    win_definitions: Dict[str, str] = field(
        default_factory=lambda: {
            "PL": "Order booked",
            "Visits": "Visit completed",
            "Outstanding": "Payment received",
            "Sales": "Order placed",
            "Long-Term": "Meeting held / DC onboarded",
            # No confirmed win condition exists for this -- Source 3d tags the underlying
            # liquidation proxy data itself Provisional pending a business definition.
            # User-added as a 6th ranked objective (2026-08-04); do not invent a formula.
            "Liquidation": "Config_Ambiguous -- no confirmed win definition (Source 3d Provisional)",
        }
    )
    # 8.7 confirmed: contact-fatigue -- max 2 attempts per rolling 3-day window,
    # 2-day wait after a failed call, -30% priority once attempts >= 2
    contact_fatigue_max_attempts: int = 2
    contact_fatigue_window_days: int = 3
    contact_fatigue_priority_cut: float = 0.30
    wait_after_failed_call_days: int = 2
    # Explicit user request 2026-09-04: a DC with real overdue aged past 90 days should
    # jump the queue -- rank/get selected ahead of every other DC in the SE's day,
    # regardless of the other three objectives' gap magnitudes. Deliberately does NOT
    # touch the Outstanding score_pct/grade shown in BO Scores/BO_Rank (user confirmed
    # that display stays as-is) -- this only affects Step 5's priority_score, which
    # decides Sr_No ordering AND what the Routing Agent's optimizer picks under capacity.
    # An out-of-band additive constant (not a multiplier) guarantees the boosted DC always
    # outranks an unboosted one: max possible unboosted priority_score is the sum of the
    # top-3 objective weights (0.40+0.35+0.25=1.0) times a gap of at most ~1.0 each, so
    # anything meaningfully larger than 1.0 dominates unconditionally.
    overdue_90_plus_priority_boost: float = 10.0
    # SE Incentive Policy (FY26-27)
    wps_weight_revenue: float = 0.25
    wps_weight_collection: float = 0.30
    wps_weight_product_mix: float = 0.25
    wps_weight_farmer_activity: float = 0.20
    wps_revenue_cliff_pct: float = 0.60  # below this, Revenue AND Product Mix -> 0
    wps_od_cap_pct: float = 0.10         # OD > 10% of trailing-12mo sales -> payout capped
    wps_od_capped_payout_pct: float = 0.70
    # Section 9 -- this quarter's top priority (tie-break 7.3)
    quarter_top_priority: str = "Outstanding"
    # 7.3 default tie-break order -- USER-SUPPLIED, not a live-computed default. Updated
    # 2026-08-06 per the "All Parameters (1)"/2026-08-05 sheet re-sync: Outstanding
    # Payments, Private Label, DC Visits, Long-Term Growth, Overall Sales. The source
    # sheet's own 7.3 row left Overall Sales (BO4) unplaced ("position TBD") -- the user
    # confirmed BO4 goes last, matching the pattern of the already-confirmed 7.4(b) all-D
    # reorder. Liquidation (user-added 6th objective, no confirmed scoring formula
    # anywhere in Source 5) stays last of all -- see win_definitions for the honest-gap
    # handling.
    default_objective_priority: Tuple[str, ...] = ("Outstanding", "PL", "Visits", "Long-Term", "Sales", "Liquidation")
    # 7.4 rule (b), confirmed (not just default-applied like rule (a)): if every BO is
    # graded D, switch to this order instead of the normal ranking.
    all_d_override_order: Tuple[str, ...] = ("Visits", "PL", "Sales", "Outstanding", "Long-Term")
    # 4.4 BO4 growth target multipliers -- REPLACED 2026-08-06 (GR-25): the flat 1.05
    # figure previously here matched nothing in Source 5 and has been retired. Real
    # category-specific values from the "All Parameters" sheet's confirmed answer (4.4).
    # Field Crop deliberately absent -- the sheet itself says "depends on seasonality (no
    # fixed multiplier given)"; GR-20 requires this be flagged provisional, not guessed.
    bo4_category_multipliers: Dict[str, float] = field(
        default_factory=lambda: {"cattle feed": 1.20, "crop nutrition": 1.15, "crop protection": 1.20}
    )
    # 4.2's Total_Working_Days_In_Period -- no confirmed non-working-day calendar exists
    # anywhere in this pipeline, so calendar days is used as a documented simplification.
    bo4_momentum_period_days: int = 30
    # 4.5 grade cut-offs -- Source 5 says "keep current cut-offs" without enumerating real
    # numbers ("engineering should confirm the exact live A/B/C/D values"). Reuses BO1's
    # confirmed 0.80/0.60/0.40 bands for directional consistency only (same treatment
    # BO3's live-proxy already got) -- not an independently confirmed BO4 cutoff.
    bo4_grade_a: float = 0.80
    bo4_grade_b: float = 0.60
    bo4_grade_c: float = 0.40
    # --- New Daily Task Assignment Formula (8.9-8.12) + Guardrails, synced 2026-08-06 ---
    max_daily_tasks: int = 5  # 8.10 -- hard cap, supersedes the old per-objective caps
    # 8.5 qualification thresholds (overridden values from the 2026-08-05 sheet)
    qualify_visits_days_since: int = 14
    qualify_outstanding_balance: float = 20000.0
    # 8.5 also specifies "overdue >= 15 days" -- NOT independently computable from
    # dc_datamart's confirmed schema (no per-DC days-overdue field, only aggregate aging
    # buckets / weighted_avg_repayment_days). Balance leg only; flagged in Safety_Flags
    # rather than silently dropped or guessed at (GR-21 fail-safe pattern).
    qualify_outstanding_days_overdue: int = 15
    qualify_pl_max_orders_30d: int = 3  # not computable live yet -- see Qualify_PL docstring
    qualify_longterm_pl_lookback_days: int = 90
    # 8.12 bundling weights reuse 7.2's rank1/2/3 weights (0.40/0.35/0.25) applied per-DC
    # instead of per-SE-objective, per Layer 3 of the new formula.
    fm_min_meetings_per_month: int = 2  # 5.3/8.11 -- no live Farmer_Meetings source yet


# Param_Key -> (BusinessConstants attribute, regex to pull the expected number out of the
# cell, caster). A deliberately small, hand-picked subset: only keys whose Configured_Value
# cell contains exactly one number tied to that single constant, so a mismatch is a
# trustworthy signal rather than parser noise. Most Source 5 cells are free-text/multi-value
# formulas (8.5, 8.7, 4.4, ...) -- extracting those reliably would risk manufacturing false
# "drift" flags, which is worse than the silent gap this replaces.
CONFIG_DRIFT_CHECKS: List[Tuple[str, str, str, Any]] = [
    ("6.2", "min_days_since_last_visit", r"(\d+)\s*days?", int),
    ("8.9", "monthly_travel_cap_km", r"([\d,]+)\s*km", lambda s: float(s.replace(",", ""))),
    ("8.10", "max_daily_tasks", r"(\d+)\s*tasks?", int),
]


def check_business_constants_against_config(constants: "BusinessConstants", config_rows: Table, exc: Exceptions) -> None:
    """Cross-checks the CONFIG_DRIFT_CHECKS subset of BusinessConstants against the live
    Source 5 CSV. BusinessConstants's own docstring claims it "mirrors" Source 5 and should
    be "re-derived if the sheet changes" -- nothing previously verified that; a business
    change in the sheet had no way to surface itself before a human happened to notice.
    Flags Config_Drift when the sheet and the hardcoded value disagree."""
    config_index = {r["Param_Key"]: r["Configured_Value"] for r in config_rows}
    for param_key, attr, pattern, cast in CONFIG_DRIFT_CHECKS:
        cell = config_index.get(param_key)
        if not cell:
            exc.flag(param_key, "Source5", "Config_Drift_Check_Unavailable", f"No Configured_Value found for {param_key} to check {attr} against")
            continue
        match = re.search(pattern, str(cell))
        if not match:
            exc.flag(param_key, "Source5", "Config_Drift_Check_Unparseable", f"Could not extract a number matching {pattern!r} from {attr}'s Configured_Value: {cell!r}")
            continue
        try:
            sheet_value = cast(match.group(1))
        except ValueError:
            exc.flag(param_key, "Source5", "Config_Drift_Check_Unparseable", f"Extracted text {match.group(1)!r} for {attr} is not a valid number")
            continue
        code_value = getattr(constants, attr)
        if sheet_value != code_value:
            exc.flag(param_key, "Source5", "Config_Drift", f"{attr}={code_value!r} hardcoded in BusinessConstants but Source 5 now says {sheet_value!r} (cell: {cell!r})")
        else:
            exc.ok("Config_Drift")


def load_config(
    all_params_csv: Path = CONFIG_ALL_PARAMS_CSV,
    se_incentive_csv: Path = CONFIG_SE_INCENTIVE_CSV,
) -> Tuple[Table, Exceptions]:
    """Source 5 -> Config_Normalized: flat key-value, one row per BO parameter."""
    exc = Exceptions(utc_now_iso())
    rows: Table = []
    section = None
    for path, kind in ((all_params_csv, "all_parameters"), (se_incentive_csv, "se_incentive_policy")):
        if not path.exists():
            exc.flag(str(path), "Source5", "Config_File_Missing", f"{kind} config file not found")
            continue
        with path.open(newline="", encoding="utf-8") as f:
            reader = list(csv.reader(f))
        header = None
        for raw in reader:
            if not any(c.strip() for c in raw):
                continue
            if raw[0].strip().startswith("Section "):
                section = raw[0].strip()
                continue
            if raw[0].strip() in ("Section", "Ref#"):
                header = [h.strip() for h in raw]
                continue
            if header is None:
                continue
            record = dict(zip(header, raw))
            key = record.get("Q#") or record.get("Ref#") or ""
            if not key:
                continue
            configured_value = record.get("Configured Value (Business Decision)") or record.get("Formula / Slabs")
            rows.append(
                {
                    "Kind": kind,
                    "Section": section,
                    "Param_Key": key,
                    "Parameter": record.get("Parameter / Topic") or record.get("Parameter"),
                    "Configured_Value": clean_null(configured_value),
                    "Formula": record.get("Calculation Formula / Logic") or record.get("Formula / Slabs"),
                    "Status": record.get("Status"),
                }
            )
    blank = [r for r in rows if not r["Configured_Value"]]
    for row in blank:
        exc.flag(row["Param_Key"], "Source5", "Config_Value_Blank", f"{row['Parameter']} has no resolved value")
    exc.ok("Config_Completeness") if not blank else None
    if not rows:
        exc.flag("ALL", "Source5", "Config_Empty", "No config rows parsed from Source 5 CSVs")
    return rows, exc


# =====================================================================================
# 4. SOURCE 2 -- DC MASTER WITH RANK (local, DC_RAnk.csv)
# =====================================================================================

COHORT_BANDS = {
    "Strategic": (1, 1000),
    "Growth": (1001, 3000),
    "Opportunity": (3001, 5545),
}

# Source 3g -- DeHaat Club Scheme 2026-27, confirmed 2026-08-13 from the signed business
# policy (Dehaat Club /*.csv, 3 documents + a State->Zone mapping, signed by Kamlesh
# Sharma, President - Agri Input, DeHaat). Supersedes the generic dc_club_slabs schema
# alone per the source doc's own instruction -- this table is the authoritative tier
# structure, not the live Redshift table (still pulled, for cross-validation only, see
# normalize_dc_club()). Ordered highest-to-lowest so club_tier_for_turnover() can walk
# it top-down and stop at the first threshold met.
DC_CLUB_TIER_TABLE = [
    # (tier, min_qualifying_turnover, tod_percent, tour_reward_west, tour_reward_east)
    ("Director's", 10_000_001, None, "Mahindra Thar (Ex-Showroom - Base Model)", "Mahindra Thar (Ex-Showroom - Base Model)"),
    ("Diamond", 6_000_000, 2.25, "Budapest + Vienna Single (4N5D) or Phuket + Krabi Couple (3N4D)", "Budapest + Vienna Single (4N5D) or Phuket + Krabi Couple (3N4D)"),
    ("Platinum", 3_500_000, 2.00, "Hong Kong + Macau Single (3N4D) or Sri Lanka Couple (3N4D)", "Hong Kong + Macau Single (3N4D) or Sri Lanka Couple (3N4D)"),
    ("Gold", 2_400_000, 1.75, "Dubai + Abu Dhabi Single (3N4D)", "Dubai + Abu Dhabi Single (3N4D)"),
    ("Silver", 1_600_000, 1.25, "Phuket + Krabi Single (3N4D)", "Phuket + Krabi Single (3N4D)"),
    ("Bronze", 900_000, 1.00, "Kerala Single (2N3D)", "Sikkim Single (2N3D)"),
    ("Copper", 400_000, 0.50, "Gold Voucher (₹16K)", "Gold Voucher (₹16K)"),
]

# State->Zone, confirmed (Dehaat Club /Zone definition - Sheet1.csv) -- covers all 11
# states DC_Master/Geo_Mapping currently carry. Affects which zone's tour reward a DC's
# club tier maps to, not the turnover threshold itself (thresholds are zone-independent).
STATE_TO_ZONE = {
    "Bihar": "East", "Chhattisgarh": "East", "Jharkhand": "East", "Orissa": "East",
    "Uttar Pradesh": "East", "West Bengal": "East",
    "Gujarat": "West", "Haryana": "West", "Madhya Pradesh": "West",
    "Maharashtra": "West", "Rajasthan": "West",
}


def club_tier_for_turnover(qualifying_turnover: Optional[float]) -> Optional[Tuple[str, Optional[float], str, str]]:
    """Returns (tier_name, tod_percent, tour_reward_west, tour_reward_east) for the
    highest tier whose threshold is met, or None if turnover is unknown or below every
    tier's entry point (Copper's 400,000). Widened 2026-08-18 to carry the tour reward
    alongside TOD% -- both are real per-tier benefits from DC_CLUB_TIER_TABLE, and
    "what tier" alone is a weaker pitch point than "what tier gets you.\""""
    if qualifying_turnover is None:
        return None
    for tier, threshold, tod_percent, reward_west, reward_east in DC_CLUB_TIER_TABLE:
        if qualifying_turnover >= threshold:
            return tier, tod_percent, reward_west, reward_east
    return None


def dc_club_participation_text(club: Optional[Dict[str, Any]]) -> str:
    """DailyTaskRow.DC_Club_Participation -- was just the enrollment-proxy string
    (Enrollment_Basis) before 2026-08-13, which never actually showed the real
    Club_Tier/Zone/TOD_Percent normalize_dc_club() now computes even though those
    fields existed in DC_Club_Normalized.json (a real gap: rich normalization output
    that never made it to the one field DailyTask/the API actually expose). Doesn't
    repeat "unconfirmed proxy" per row -- that caveat is already recorded once per run
    via the Club_Enrollment_Flag_Unconfirmed exception, same convention as every other
    run-level (not per-DC) caveat in this pipeline."""
    if not club or not club.get("Is_Club_Enrolled"):
        return "Not enrolled"
    tier = club.get("Club_Tier")
    if tier:
        bits = [tier]
        if club.get("Zone"):
            bits.append(f"{club['Zone']} zone")
        if club.get("TOD_Percent") is not None:
            bits.append(f"{club['TOD_Percent']:.2f}% TOD")
        # Bug fixed 2026-08-19: an already-tiered DC never had its own reward shown,
        # only DCs still working towards one (Eligible_Tier_Reward_If_Cleared below) --
        # see normalize_dc_club's Reward field for the fix.
        if club.get("Reward"):
            bits.append(club["Reward"])
        return "Enrolled -- " + ", ".join(bits)
    copper_threshold = DC_CLUB_TIER_TABLE[-1][1]  # lowest tier, table ordered highest-to-lowest
    if club.get("Outstanding_Cleared") is False:
        eligible = club.get("Eligible_Tier_If_Outstanding_Cleared")
        if eligible:
            benefit_bits = []
            if club.get("Eligible_Tier_TOD_Percent_If_Cleared") is not None:
                benefit_bits.append(f"{club['Eligible_Tier_TOD_Percent_If_Cleared']:.2f}% TOD")
            if club.get("Eligible_Tier_Reward_If_Cleared"):
                benefit_bits.append(club["Eligible_Tier_Reward_If_Cleared"])
            benefit_note = f" -- {', '.join(benefit_bits)}" if benefit_bits else ""
            return f"Enrolled -- no tier yet (outstanding not cleared; would be {eligible} if cleared{benefit_note})"
        # Bug fixed 2026-08-19: outstanding not cleared isn't necessarily the ONLY
        # blocker -- a DC with no/below-threshold turnover has nothing to become
        # eligible for even once outstanding clears, but this used to say only
        # "outstanding not cleared," silently hiding the second, often-bigger gap and
        # implying clearing outstanding alone would unlock a tier when it wouldn't.
        turnover = club.get("Qualifying_Turnover")
        if turnover is None:
            return "Enrolled -- no tier yet (outstanding not cleared; no qualifying turnover this scheme year either)"
        return f"Enrolled -- no tier yet (outstanding not cleared; turnover ₹{turnover:,.0f} is also below Copper's ₹{copper_threshold:,.0f} entry threshold)"
    if club.get("Qualifying_Turnover") is None:
        return "Enrolled -- no tier yet (no qualifying turnover this scheme year)"
    return f"Enrolled -- no tier yet (turnover ₹{club['Qualifying_Turnover']:,.0f} is below Copper's ₹{copper_threshold:,.0f} entry threshold)"


def load_dc_master(path: Path = DC_MASTER_CSV) -> Tuple[Table, Exceptions]:
    exc = Exceptions(utc_now_iso())
    out: Table = []
    if not path.exists():
        exc.flag(str(path), "Source2", "DC_Master_Missing", "DC_RAnk.csv not found")
        return out, exc

    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        seen_ids = set()
        for i, row in enumerate(reader):
            dc_id = normalize_id(row.get("Partner Id"))
            if dc_id is None:
                exc.flag(i, "Source2", "DC_ID_Missing", "Row has no Partner Id")
                continue
            if dc_id in seen_ids:
                exc.flag(dc_id, "Source2", "Duplicate_DC_ID", "Partner Id repeats within DC Master")
            seen_ids.add(dc_id)

            cohort = (clean_null(row.get("Cohort")) or "").strip()
            raw_rank = clean_null(row.get("Rank"))
            rank_numeric = parse_number(raw_rank) if cohort != "Long Tail" else None
            if cohort in COHORT_BANDS and rank_numeric is not None:
                lo, hi = COHORT_BANDS[cohort]
                if not (lo <= rank_numeric <= hi):
                    exc.flag(dc_id, "Source2", "Cohort_Rank_Mismatch", f"Rank {rank_numeric} outside {cohort} band {lo}-{hi}")
                else:
                    exc.ok("Cohort_Rank_Consistency")
            elif cohort == "Long Tail" and raw_rank not in (None, "Long Tail"):
                exc.flag(dc_id, "Source2", "Cohort_Rank_Mismatch", "Long Tail cohort should carry text placeholder Rank")

            total_score = parse_number(row.get("TOTAL SCORE"))
            unscored = cohort == "Long Tail" or total_score == 0.0

            # "TOP SE PL State" is a mislabeled column: it holds the assigned SE's email,
            # not a state. Read by position/content, not header name (Section 2 note).
            assigned_se_email = clean_null(row.get("TOP SE PL State"))
            has_assigned_se = assigned_se_email is not None
            if not has_assigned_se:
                exc.flag(dc_id, "Source2", "Unassigned_DC", "DC has no SE assignment; excluded from any daily plan")
            else:
                exc.ok("Unassigned_DC_Check")

            out.append(
                {
                    "DC_ID": dc_id,
                    "DC_Name": clean_null(row.get("Partner Name")),
                    "Node": clean_null(row.get("node_name")),
                    "State": clean_null(row.get("state_name")),
                    "Rank": rank_numeric if rank_numeric is not None else clean_null(raw_rank),
                    "Cohort": cohort or None,
                    "Total_Score": None if unscored else total_score,
                    "Total_Score_Unscored": unscored,
                    "Assigned_SE_Email": assigned_se_email,
                    "Has_Assigned_SE": has_assigned_se,
                    "NRV_FY2526": parse_number(row.get("NRV FY-25-26")),
                    "GM_FY2526": parse_number(row.get("GM FY-25-26")),
                    "GM_Percent": parse_number(row.get("GM%")),
                    "PL_Percent": parse_number(row.get("PL%")),
                    "Avg_Repayment_Days": parse_number(row.get("Avg Repayment days")),
                    "Credit_Score": parse_number(row.get("Credit Score")),
                    "In_Scope_Flag": None,  # resolved in apply_dc_exclusion_rules() once Source 4 status is known
                    "Latitude": None,       # filled from Geo_Mapping_Normalized (1c) join
                    "Longitude": None,
                    "DC_Status": None,      # filled from Source 4 active-status cross-check
                }
            )
    if not any(r for r in exc.rows if r["Reason_Code"] == "Duplicate_DC_ID"):
        exc.ok("Duplicate_DC_ID")
    return out, exc


def load_top_dc_allowlist(path: Path = TOP_DC_LIST_XLSX) -> Tuple[Optional[Set[str]], Exceptions]:
    """Added 2026-09-04, explicit user request -- restricts the DC universe to only the
    Partner Ids listed in 'updated TOP DC list.xlsx', independent of (on top of, not
    instead of) DC_RAnk.csv's own Rank<=6000 eligibility -- see apply_dc_exclusion_
    rules()'s new top_dc_allowlist check.

    Returns (None, exc) -- NOT an empty set -- when the file is missing or fails to
    parse, per direct instruction ("fail open"): apply_dc_exclusion_rules() treats None
    as "no allowlist restriction this run" rather than "everyone fails," so one missing/
    renamed/corrupted file can't silently zero out DC selection for the whole network.
    A loud DC_Top_List_Missing/DC_Top_List_Unreadable exception is still flagged either
    way, so the degraded run is never silent about it."""
    exc = Exceptions(utc_now_iso())
    if openpyxl is None:
        exc.flag("openpyxl", "Source2b", "DC_Top_List_Unreadable", "openpyxl not installed -- Top DC allowlist restriction skipped this run, every Rank-eligible DC treated as eligible")
        return None, exc
    if not path.exists():
        exc.flag(str(path), "Source2b", "DC_Top_List_Missing", f"{path.name} not found -- Top DC allowlist restriction skipped this run, every Rank-eligible DC treated as eligible")
        return None, exc
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(min_row=2, values_only=True)
        allowlist: Set[str] = set()
        for row in rows:
            dc_id = normalize_id(row[0]) if row and row[0] is not None else None
            if dc_id:
                allowlist.add(dc_id)
        wb.close()
    except Exception as e:
        exc.flag(str(path), "Source2b", "DC_Top_List_Unreadable", f"{type(e).__name__}: {e} -- Top DC allowlist restriction skipped this run, every Rank-eligible DC treated as eligible")
        return None, exc
    if not allowlist:
        exc.flag(str(path), "Source2b", "DC_Top_List_Empty", f"{path.name} parsed with zero valid Partner Ids -- Top DC allowlist restriction skipped this run")
        return None, exc
    exc.ok("DC_Top_List_Loaded")
    return allowlist, exc


def apply_dc_exclusion_rules(
    dc_master: Table, exc: Exceptions, constants: BusinessConstants, last_visit_by_dc: Dict[str, str], today: str,
    top_dc_allowlist: Optional[Set[str]] = None,
) -> None:
    """Section 6: rules that remove a DC from consideration entirely.
    6.4 (Agent-Determined): always block Legal_Hold; Credit_Blocked/Blacklisted are
    evaluated live elsewhere (see resolve_full_block), not auto-blocked here.
    6.2 (confirmed default, RESTORED 2026-09-04 after a same-day removal-then-revert --
    explicit user request both times): exclude a DC entirely if visited within the last
    min_days_since_last_visit (5) days. Tied to a live config-drift check (Param_Key
    "6.2" in Source 5).
    Rank<=6000 (BusinessConstants.max_eligible_rank) is DISABLED for eligibility as of
    2026-09-04, explicit user request -- DC_RAnk.csv's Rank column no longer gates
    eligibility in ANY capacity, including as a fallback. Kept defined on
    BusinessConstants only in case of reintroduction, same pattern as completion_
    multiplier/min_days_since_last_visit's own remove-then-restore history. DC_RAnk.csv
    itself is untouched -- it's still Source 2's DC Master (names/node/Cohort/financials)
    and still drives Step 5's Cohort/Total_Score selection ordering; only its Rank
    column's eligibility role is gone.

    top_dc_allowlist (2026-09-04, explicit user request, "use that list only" then "go
    to first updated TOP DC list file"): the ONLY DC-selection eligibility gate now,
    beyond Has_Assigned_SE/not-Legal_Hold/not-too-recent. FAIL-CLOSED, not fail-open: if
    the allowlist itself fails to load this run (None), every DC is excluded rather than
    falling back to Rank -- a load failure is now loud (zero tasks network-wide that run)
    rather than silently substituting a different, unconfirmed-for-this-purpose rule."""
    today_dt = datetime.fromisoformat(today)
    for dc in dc_master:
        legal_hold = dc.get("DC_Status") == "Legal_Hold"
        last_visit = last_visit_by_dc.get(dc["DC_ID"])
        days_since_visit = None
        if last_visit:
            days_since_visit = (today_dt - datetime.fromisoformat(last_visit)).days
        too_recent = days_since_visit is not None and days_since_visit < constants.min_days_since_last_visit
        top_list_eligible = top_dc_allowlist is not None and dc["DC_ID"] in top_dc_allowlist
        if top_dc_allowlist is None:
            exc.flag(
                dc["DC_ID"], "Source2b", "Top_DC_List_Unavailable",
                "'updated TOP DC list.xlsx' failed to load this run -- excluded from all agents' "
                "DC selection (fail-closed, no Rank<=6000 fallback)",
            )
        elif not top_list_eligible:
            exc.flag(
                dc["DC_ID"], "Source2b", "DC_Not_In_Top_List",
                "DC_ID not present in 'updated TOP DC list.xlsx' -- excluded from all agents' DC selection",
            )
        else:
            exc.ok("DC_Not_In_Top_List")
        in_scope = dc["Has_Assigned_SE"] and not legal_hold and not too_recent and top_list_eligible
        dc["In_Scope_Flag"] = in_scope
        dc["Days_Since_Last_Visit"] = days_since_visit
        dc["Last_Visit_Date"] = last_visit
        if legal_hold:
            exc.flag(dc["DC_ID"], "Source2", "DC_Legal_Hold_Blocked", "DC under Legal_Hold, excluded from all lists")


SQL_LIVE_DC_ROSTER = """
SELECT sap_partner_id, partner_name, node_name, state_name, sales_rep_email, lat_2, long_2
FROM input_partner_details
WHERE active = 'true' AND sap_partner_id IS NOT NULL
"""

SQL_CANONICAL_NODE_MAPPING = """
SELECT DISTINCT state, node
FROM input_node_mapping
WHERE state <> 'State' AND state IS NOT NULL AND state <> ''
"""


def supplement_dc_master_from_live(dc_master: Table, client, exc: Exceptions) -> None:
    """Confirmed live 2026-08-09: DC_RAnk.csv (Source 2's local export) is missing real,
    active DCs that exist live in input_partner_details -- 20 entire nodes' worth across
    9 states, confirmed against the canonical node-state master (input_node_mapping,
    113 nodes/11 states, matches Metabase exactly) and cross-checked DC-by-DC (e.g. Agra
    has 2 real live DCs, zero in the local export). This silently excluded those DCs
    from every NODE/STATE-scoped plan.

    Originally scoped to ONLY those ~20 fully-missing CANONICAL nodes, deliberately
    excluding the other 9,096 of 19,069 active DCs (48%) missing within nodes DC_Master
    already covers -- flagged as too large a behavior change at the time (nearly doubles
    DC_Master's size). Extended network-wide 2026-08-12 by user decision, after a real
    STATE=Rajasthan plan run silently excluded 14 of 24 actually-assigned live SEs (DCs
    existed live, in nodes DC_Master already had, but not the specific DC rows) -- the
    within-node gap isn't a hypothetical, it visibly cuts real SEs out of real plans, so
    the full fix is now in scope.

    "Real node" is still judged against input_node_mapping (the canonical 113-node
    master, same table Metabase's own count comes from), NOT simply "not already in
    DC_Master" -- a first version used the latter and it was wrong: input_partner_
    details.node_name carries a lot of non-canonical junk (virtual/cluster labels like
    'AHMEDNAGAR_VIRTUAL', 'Do not Use', 'Frontier market_Varanasi_Pindra') that were
    never real nodes to begin with. Cross-referencing the canonical list first is what
    correctly admits every real missing DC (whether its node is entirely new to
    DC_Master or already partially covered) while still rejecting the noise.

    Appends each qualifying live-active DC with real Node/State/Assigned_SE_Email/geo
    from input_partner_details, but Rank/Cohort/Total_Score/NRV/GM/PL%/Avg_Repayment_
    Days/Credit_Score all left None -- DC_RAnk.csv is the only source for those, so a
    supplemented DC is honestly unscored (Total_Score_Unscored=True), never a fabricated
    rank. Every added DC is individually flagged DC_Supplemented_From_Live so it's
    traceable in the exceptions report, not silent. Mutates dc_master in place (appends)
    -- called after DC_RAnk.csv's own load_dc_master() and before apply_dc_exclusion_
    rules(), so a supplemented DC still goes through the same Has_Assigned_SE /
    Legal_Hold / recent-visit scoping as every other DC."""
    existing_ids = {dc["DC_ID"] for dc in dc_master}
    existing_nodes = {dc["Node"].strip().upper() for dc in dc_master if dc.get("Node")}
    try:
        canonical_rows = client.execute_sql(REDSHIFT_DB_ID, SQL_CANONICAL_NODE_MAPPING)
        canonical_nodes = {r["node"].strip().upper() for r in canonical_rows if r.get("node")}
        live_rows = client.execute_sql(REDSHIFT_DB_ID, SQL_LIVE_DC_ROSTER)
    except Exception as e:
        exc.flag("ALL", "Source2", "DC_Live_Supplement_Failed", f"{type(e).__name__}: {e}")
        return

    missing_canonical_nodes = canonical_nodes - existing_nodes

    added = 0
    for row in live_rows:
        dc_id = normalize_id(row.get("sap_partner_id"))
        node = clean_null(row.get("node_name"))
        node_key = node.strip().upper() if node else None
        if dc_id is None or dc_id in existing_ids:
            continue
        if not node_key or node_key not in canonical_nodes:
            continue  # only real canonical nodes qualify -- rejects junk/virtual node labels
        existing_ids.add(dc_id)
        lat, lon = parse_number(row.get("lat_2")), parse_number(row.get("long_2"))
        assigned_se_email = clean_null(row.get("sales_rep_email"))
        dc_master.append(
            {
                "DC_ID": dc_id,
                "DC_Name": clean_null(row.get("partner_name")),
                "Node": clean_null(row.get("node_name")),
                "State": clean_null(row.get("state_name")),
                "Rank": None,
                "Cohort": None,
                "Total_Score": None,
                "Total_Score_Unscored": True,
                "Assigned_SE_Email": assigned_se_email,
                "Has_Assigned_SE": assigned_se_email is not None,
                "NRV_FY2526": None,
                "GM_FY2526": None,
                "GM_Percent": None,
                "PL_Percent": None,
                "Avg_Repayment_Days": None,
                "Credit_Score": None,
                "In_Scope_Flag": None,  # resolved in apply_dc_exclusion_rules(), same as every other DC
                "Latitude": lat,
                "Longitude": lon,
                "DC_Status": None,
            }
        )
        if node_key in missing_canonical_nodes:
            reason = (f"Node '{node}' entirely missing from DC_RAnk.csv (local export) but active in "
                      "input_partner_details (live) -- added unscored (no Rank/Cohort/financials available "
                      "outside DC_RAnk.csv)")
        else:
            reason = (f"DC missing from DC_RAnk.csv (local export) within node '{node}', which DC_Master "
                      "otherwise covers, but active in input_partner_details (live) -- added unscored "
                      "(no Rank/Cohort/financials available outside DC_RAnk.csv)")
        exc.flag(dc_id, "Source2", "DC_Supplemented_From_Live", reason)
        added += 1
    if added:
        logger.info("  Supplemented %d DC(s) missing from DC_RAnk.csv with live input_partner_details data", added)


# =====================================================================================
# 5. SOURCE 6 -- AOP & TARGET DATA (local, Niyojan dashboard export)
# =====================================================================================

def load_aop_targets(path: Path = AOP_CSV) -> Tuple[Table, Exceptions]:
    """Source 6 is a wide, messy dashboard export with no confirmed field structure
    (doc: 'Fields: SE_ID/DC_ID, Month, Metric, AOP_Target [FILL IN exact structure]',
    'not yet live-verified'). Confirmed live 2026-08-09: the real 106-column export has
    NO DC/SE dimension anywhere -- finest grain is Node x Material/SKU x Month. This
    loader locates the real header row defensively and, as of 2026-08-09, also captures
    each row's Segment (needed to isolate PRIVATE LABEL rows) and its real GMV (AOP)
    value per month -- previously only Node/Metric/Status metadata was captured, with no
    usable number anywhere in the output. Every row still tagged Provisional -- it must
    not be trusted at the same confidence as the confirmed sources until the business
    confirms the exact layout. See aop_pl_target_by_node() for how a per-DC figure gets
    derived from this Node-level data (an allocation estimate, not a confirmed target)."""
    exc = Exceptions(utc_now_iso())
    out: Table = []
    if not path.exists():
        exc.flag(str(path), "Source6", "AOP_File_Missing", "Niyojan planning file not found")
        return out, exc

    with path.open(newline="", encoding="utf-8") as f:
        reader = list(csv.reader(f))
    header_idx = next((i for i, r in enumerate(reader) if r and r[0].strip() == "Node Key"), None)
    if header_idx is None:
        exc.flag(str(path), "Source6", "AOP_Structure_Unrecognized", "Could not locate 'Node Key' header row")
        return out, exc

    def _norm_header(h: str) -> str:
        # Raw headers carry embedded newlines/extra whitespace (e.g. 'GMV \nAug 26
        # (AOP)') -- collapse before matching by name, so a column-order reshuffle in a
        # future re-export (this file has already broken the agent's default path once
        # via a filename-suffix change, see AGENT_OPERATING_PROMPTS.md) doesn't silently
        # misread values via a hardcoded index.
        return re.sub(r"\s+", " ", h).strip()

    header = [_norm_header(h) for h in reader[header_idx]]
    # Only Q2 FY26-27 (Jul/Aug/Sep 2026) months exist as real AOP columns in this export;
    # aop_pl_target_by_node() falls back to the quarterly total /3 for any plan_date
    # outside this range.
    gmv_aop_cols = {7: "GMV July 26 (AOP)", 8: "GMV Aug 26 (AOP)", 9: "GMV Sept 26 (AOP)"}
    for i, raw in enumerate(reader[header_idx + 1 :], start=header_idx + 1):
        if not any(c.strip() for c in raw):
            continue
        record = dict(zip(header, raw))
        row = {
            "Node_Key": clean_null(record.get("Node Key")),
            "State": clean_null(record.get("State")),
            "Node": clean_null(record.get("Node")),
            "Zone": clean_null(record.get("Zone")),
            "RBM": clean_null(record.get("RBM")),
            "Segment": clean_null(record.get("Segment")),
            "Metric": clean_null(record.get("Material Name") or record.get("Category")),
            "AOP_Version": "unverified_local_export",
            "Last_Adjusted_Date": None,
            "Status": "Provisional",
        }
        for month_num, col_name in gmv_aop_cols.items():
            row[f"GMV_AOP_Month_{month_num}"] = parse_number(record.get(col_name))
        row["GMV_AOP_Q2"] = parse_number(record.get("GMV Q2 26 (AOP)"))
        out.append(row)
    exc.flag("ALL", "Source6", "AOP_Freshness_Unknown", "Last_Adjusted_Date not derivable from this export; treat AOP figures as Provisional")
    return out, exc


def aop_pl_target_by_node(aop_targets: Table, plan_date: str) -> Dict[str, float]:
    """Real Node-level PL (Private Label) AOP target for plan_date's month, summed across
    every PRIVATE LABEL-segment row for that Node -- NOT per-DC/per-SE, the source has no
    finer grain (see load_aop_targets docstring). Callers needing a per-DC figure must
    allocate this down themselves (e.g. by trailing PL sales share within the Node) and
    must label the result an estimate, not a confirmed per-DC AOP target -- see
    planning/services.py's PL scoring block for the live allocation. Node names are
    normalized (stripped + uppercased) before summing since the raw export carries the
    same real node under inconsistent casing (both 'Alwar' and 'ALWAR' rows exist).
    Falls back to the quarterly total /3 for a plan_date outside Jul-Sep 2026, the only
    months this export actually covers -- a coarser approximation, not a missing value."""
    month_num = datetime.fromisoformat(plan_date).date().month
    in_quarter = month_num in (7, 8, 9)
    totals: Dict[str, float] = {}
    for row in aop_targets:
        if (row.get("Segment") or "").strip().upper() != "PRIVATE LABEL":
            continue
        node = (row.get("Node") or "").strip().upper()
        if not node:
            continue
        value = row.get(f"GMV_AOP_Month_{month_num}") if in_quarter else row.get("GMV_AOP_Q2")
        if value is None:
            continue
        value = value if in_quarter else value / 3.0
        totals[node] = totals.get(node, 0.0) + value
    return totals


# =====================================================================================
# 6. SOURCES 1 / 3 / 4 -- LIVE METABASE QUERIES (real-time sync)
# =====================================================================================
# SQL below mirrors the confirmed join chains, table/column names and quirks documented
# in the agent prompt (no FILTER clause on Redshift -> CASE WHEN; explicit bool casts).
# Column names below are CONFIRMED LIVE against the direct-Redshift connection
# (RedshiftDirectClient, 2026-08-04) via information_schema.columns -- several corrected
# real assumptions from the doc / earlier sessions:
#   pathik_report has no plan_execution_date column at all -- it's transaction_date.
#   attendance_attendance's geo columns are check_in_latitude/longitude (not _lat/_long),
#     and it also carries total_distance_travelled/osrm_match_coordinates/
#     osrm_trip_coordinates/google_polyline -- real road-distance data, an alternative
#     to haversine sequencing the doc left undecided (see sequence_with_distance()).
#   Node/Block/District/State and ABM/RBM/ZBM/Growth-Manager names+emails all live
#     directly on input_partner_details (block_name, district_name, node_name, state_name,
#     abm, "abm email id", rbm, "rbm email id", zbm, "zbm email id", "growth manager name",
#     "growth manager email id") -- input_se_node_mapping does NOT carry node/block/district
#     at all (only zone, state, "p&l node"); it's a separate table keyed by
#     sales_rep_email that only adds employee CODES (emp id se, abm e code, rbm e code,
#     zbm e code, growth manager e code), joined in for that reason alone.

def _lookback_clause(date_column: str, days: int = LOOKBACK_DAYS) -> str:
    return f"{date_column} >= CURRENT_DATE - INTERVAL '{days} days'"


SQL_TASK_NODES_1B = f"""
SELECT
    t.id AS task_id, t.plan_id, t.partner_id AS dc_id, t.status AS task_status,
    t.type AS task_type, t.visit_type_id, t.visit_purpose_id,
    p.user_id AS se_user_id, p.plan_execution_date, p.status AS plan_status, p.created_by_id,
    u.mobile AS mobile_number, u.email AS se_email, u.is_active AS se_is_active
FROM task_management_task t
JOIN task_management_plan p ON p.id = t.plan_id
JOIN users_user u ON u.id = p.user_id
WHERE {_lookback_clause('p.plan_execution_date')}
"""

SQL_GEO_MAPPING_1C = """
WITH se_map AS (
    SELECT DISTINCT
        sales_rep_email,
        "emp id se" AS emp_id_se,
        "abm e code" AS abm_e_code,
        "rbm e code" AS rbm_e_code,
        "zbm e code" AS zbm_e_code,
        "growth manager e code" AS growth_manager_e_code
    FROM input_se_node_mapping
)
SELECT
    ipd.sap_partner_id AS dc_id, ipd.odoo_partner_id, ipd.ib_partner_id,
    ipd.lat_2 AS latitude, ipd.long_2 AS longitude, ipd.active, ipd.sales_rep_email,
    sm.emp_id_se, sm.abm_e_code, sm.rbm_e_code, sm.zbm_e_code, sm.growth_manager_e_code,
    ipd.abm, ipd."abm email id" AS abm_email,
    ipd.rbm, ipd."rbm email id" AS rbm_email,
    ipd.zbm, ipd."zbm email id" AS zbm_email,
    ipd."growth manager name" AS growth_manager_name, ipd."growth manager email id" AS growth_manager_email,
    ipd.node_name AS node, ipd.block_name AS block, ipd.district_name AS district, ipd.state_name AS state
FROM input_partner_details ipd
LEFT JOIN se_map sm ON sm.sales_rep_email = ipd.sales_rep_email
WHERE ipd.is_dc = true AND ipd.active = 'true'
"""

SQL_ATTENDANCE_3A = f"""
SELECT user_id AS se_user_id, check_in_time, check_out_time,
       check_in_latitude AS check_in_lat, check_in_longitude AS check_in_long,
       check_out_latitude AS check_out_lat, check_out_longitude AS check_out_long,
       google_distance, total_distance_travelled
FROM attendance_attendance
WHERE {_lookback_clause('check_in_time')}
"""

SQL_ACTIVE_ROSTER_4 = f"""
-- Dedup CTEs, confirmed necessary live 2026-08-14: both attendance_attendance (138 real
-- (user_id, date) groups with >1 row in a 90-day window) and task_management_taskdetails
-- (53 real task_id groups with >1 row) can have more than one row per join key, which
-- silently fanned a single task_management_task row out into multiple output rows --
-- ~580 of the ~10,675 Duplicate_Visit flags in a typical run were this artifact, not
-- genuine duplicate visits (confirmed by sampling: same task_id, same visit_check_in_time,
-- only punch_in_time differed -- i.e. the SAME task joined against two different
-- attendance rows). ROW_NUMBER() picks a single deterministic row per key (earliest) so
-- each task_management_task row can only ever produce exactly one output row here.
WITH attendance_dedup AS (
    SELECT *, ROW_NUMBER() OVER (PARTITION BY user_id, check_in_time::date ORDER BY check_in_time ASC) AS rn
    FROM attendance_attendance
),
taskdetails_dedup AS (
    SELECT *, ROW_NUMBER() OVER (PARTITION BY task_id ORDER BY created_at ASC) AS rn
    FROM task_management_taskdetails
)
SELECT
    t.id AS task_id, p.user_id AS se_user_id, u.is_active AS se_is_active,
    cc.partner_id AS dc_id, cc.active AS dc_active,  -- cc.id is the internal PK, NOT sap_partner_id -- confirmed live 2026-08-04
    a.check_in_time AS punch_in_time, a.check_out_time AS punch_out_time,
    a.check_in_latitude AS punch_in_lat, a.check_in_longitude AS punch_in_long,
    a.check_out_latitude AS punch_out_lat, a.check_out_longitude AS punch_out_long,
    a.total_distance_travelled,
    td.created_at AS visit_check_in_time, td.latitude AS visit_check_in_lat,
    td.longitude AS visit_check_in_long, td.task_details,
    t.visit_type_id, t.status AS task_status, p.plan_execution_date
FROM task_management_task t
JOIN task_management_plan p ON p.id = t.plan_id
JOIN users_user u ON u.id = p.user_id
LEFT JOIN customer_management_customer cc ON cc.id = t.partner_id
LEFT JOIN attendance_dedup a ON a.user_id = p.user_id AND a.check_in_time::date = p.plan_execution_date AND a.rn = 1
LEFT JOIN taskdetails_dedup td ON td.task_id = t.id AND td.rn = 1
WHERE {_lookback_clause('p.plan_execution_date')}
"""

SQL_SALES_TRANSACTIONS_3D = f"""
SELECT sap_order_date, order_value, cancellations_flag, business_segment, order_request_id
FROM coupon_analysis
WHERE {_lookback_clause('sap_order_date')}
"""

# SUPERSEDES customer_management_input_outstanding entirely -- confirmed live 2026-08-04
# that table does not exist anywhere reachable on this Redshift cluster (searched all 32
# databases the readonly credential can see), and no separate input-backend Postgres
# credential was available. dc_datamart (dev, Redshift -- same cluster/db as everything
# else in Source 1c/3d) is a denormalized per-DC datamart with the same information and
# genuinely fresh data (confirmed live: MAX(last_invoice_date) = 2026-07-30, i.e. days
# old, not the ~2-year-stale snapshots customer_management_input_outstanding had).
# Grain: one row per sap_partner_id, except ~100 rows (of ~14,862) have a NULL
# sap_partner_id -- filtered out below, not a real per-DC duplicate.
# total_outstanding/total_overdue map directly to the old current_os/current_od; the
# aging split here (current_month_os/os_1_to_90/os_90_plus) is coarser-grained than the
# old os_60/overdue_in_7_days but from fresh, reliable data instead of stale text fields.
# is_active filter added 2026-09-01 (explicit user request) -- dc_datamart's own
# is_active/is_blocked are varchar columns storing literal 'true'/'false' strings, not a
# real boolean type (confirmed live). Only ~30% of rows (7,197/23,536) are is_active=
# 'true'. is_active is pulled as a column (not filtered in SQL) so
# normalize_sales_transactions() can flag DC_Datamart_Inactive_Outstanding_Unavailable
# for each excluded DC instead of silently omitting it -- this codebase's own
# convention (every data gap gets a real Exceptions_Report entry, never just a silent
# blank field), confirmed after a 2026-09-01 review found ~22% of currently-eligible
# (Rank<=6000) DCs lose Outstanding/Overdue coverage from this filter.
SQL_OUTSTANDING_3D = """
SELECT sap_partner_id AS dc_id, total_outstanding, total_overdue, current_month_os,
       os_1_to_90, os_90_plus, weighted_avg_repayment_days, last_invoice_date, is_mismatch,
       is_active
FROM dc_datamart
WHERE sap_partner_id IS NOT NULL
"""

# Full 23-column schema confirmed (supersedes the earlier 2-column partial view).
# status='processed' is the only status that should count as a real order (Source 3d note).
# credit_on_hold/credit_on_hold_reason feed the 6.4 Credit_Blocked case-by-case check.
# NEW TRAP, confirmed live (not in the doc): sale_orderrequest.partner_id is the INTERNAL
# customer_management_customer.id, NOT sap_partner_id directly -- same indirection as
# task_management_task.partner_id (Source 4). A direct sap_partner_id filter on this table
# silently returns zero rows. Bridge through customer_management_customer, same as 1b/4.
SQL_ORDERS_3D = f"""
SELECT o.id, o.amount_total, o.created_at, o.request_date, o.status, cc.partner_id AS dc_id,
       o.created_by_id, o.sales_channel, o.payment_mode, o.credit_on_hold,
       o.credit_on_hold_reason, o.partner_finance_status
FROM sale_orderrequest o
JOIN customer_management_customer cc ON cc.id = o.partner_id
WHERE {_lookback_clause('o.created_at')}
"""

# Replaced 2026-08-06 -- hyperlocal_order (the original Source 3d liquidation proxy) has
# NO DC/partner join key at all (columns: id, full_display_address, created_at, source,
# status; confirmed live), so it was pulled every run and silently discarded, never
# usable. invoice_liquidation_with_pog is real, live (max invoice_date = today,
# confirmed), and joins directly on partner_id = sap_partner_id (no bridge table needed,
# unlike sale_orderrequest/payments_paymenttransaction) -- confirmed live via a join to
# input_partner_details. Per GR-24 (Guardrails sheet): a confirmed data source is not a
# confirmed formula -- this feeds normalize_liquidation() for visibility only, NEVER a
# score/grade, until a Liquidation_Rate formula is signed off.
SQL_LIQUIDATION_3D = f"""
SELECT partner_id, invoice_date, billed_qty, liquidated_qty, business_category, net_billed_amount
FROM invoice_liquidation_with_pog
WHERE {_lookback_clause('invoice_date')}
"""

# Source 3f -- Payments (new). Join key customer_id -> partner_id is NOT yet proven to be
# the same ID space; every derived Last_Payment_Date must be flagged Join_Key_Unconfirmed.
# Join key CONFIRMED live 2026-08-06: payments_paymenttransaction.customer_id matches
# customer_management_customer.id (the internal PK), NOT partner_id/sap_partner_id
# directly -- 521,345 of 538,039 rows (96.9%) matched via .id, 0 matched via .partner_id.
# Same bridging pattern as sale_orderrequest.partner_id (SQL_ORDERS_3D) -- bridge through
# customer_management_customer to get sap_partner_id, exactly like Orders.
SQL_PAYMENTS_3F = f"""
SELECT cc.partner_id AS dc_id, p.id, p.status, p.created_at, p.reference_number,
       p.payment_reference_id, p.pg_payment_mode, p.is_va_payment
FROM payments_paymenttransaction p
JOIN customer_management_customer cc ON cc.id = p.customer_id
WHERE {_lookback_clause('p.created_at')}
"""

# Source 3j -- Promise To Pay (added 2026-09-04, per SE_DC_Data_Normalization_Agent_
# Prompt.docx update). Primary, confirmed-live source: task_management_visitpurposedetails
# joined to task_management_task (input-backend Postgres, database_id 31) --
# visit_purpose_id=4 is the confirmed 'Promise To Pay / Collection' id. Supersedes the
# stale public.promise_to_pay Redshift table (frozen since 2024-12-31 per the doc, not
# used here) for "current" commitment tracking.
#
# Dialect trap, confirmed live: this Redshift replica has no native json/jsonb type --
# the doc's own ->> operator syntax fails (UndefinedFunction). Use JSON_EXTRACT_PATH_TEXT
# instead, same "this cluster's Postgres dialect differs from real Postgres" pattern as
# ROW_NUMBER-instead-of-DISTINCT-ON elsewhere in this file.
#
# Only the MOST RECENT promise per DC is kept (ROW_NUMBER, rn=1) -- older promises are
# superseded, per direct instruction. paid_on_time is computed server-side: ANY real
# SUCCESS payment landing between when the promise was logged (promise_created_at) and
# its committed date (promise_date) counts -- not required to cover the full promised
# amount, per direct instruction. Known trap, confirmed live (matches the doc's own
# finding): ~32% of records have promise_amount = 0 (an SE logged the purpose without a
# specific number) -- still carried through and still checked for a qualifying payment,
# per direct instruction, not excluded as junk.
SQL_PROMISE_TO_PAY_3J = f"""
WITH latest_promise AS (
    SELECT vpd.id AS record_id, cc.partner_id AS dc_id,
           JSON_EXTRACT_PATH_TEXT(vpd.visit_purpose_details, 'amount') AS promise_amount_raw,
           TIMESTAMP 'epoch' + CAST(JSON_EXTRACT_PATH_TEXT(vpd.visit_purpose_details, 'date') AS BIGINT) * INTERVAL '1 second' AS promise_date,
           vpd.created_at AS promise_created_at,
           ROW_NUMBER() OVER (PARTITION BY cc.partner_id ORDER BY vpd.created_at DESC) AS rn
    FROM task_management_visitpurposedetails vpd
    JOIN task_management_task t ON t.id = vpd.task_id
    JOIN customer_management_customer cc ON cc.id = t.partner_id
    WHERE vpd.visit_purpose_id = 4 AND {_lookback_clause('vpd.created_at')}
)
SELECT lp.dc_id, lp.promise_amount_raw, lp.promise_date, lp.promise_created_at,
       EXISTS (
           SELECT 1 FROM payments_paymenttransaction p
           JOIN customer_management_customer cc2 ON cc2.id = p.customer_id
           WHERE cc2.partner_id = lp.dc_id AND p.status = 'SUCCESS'
             AND p.created_at >= lp.promise_created_at AND p.created_at <= lp.promise_date
       ) AS paid_on_time
FROM latest_promise lp
WHERE lp.rn = 1
"""

# Source 3g -- DC Club Scheme (new). dc_mapping_club_scheme's presence-by-partner_id is a
# PLAUSIBLE, not confirmed, proxy for enrollment (no explicit is_member/status flag exists).
# dc_club_slabs defines tier rules (turnover band -> club), not a per-DC assignment --
# actual tier must be computed downstream against each DC's turnover from Source 3d.
SQL_DC_CLUB_MAPPING_3G = """
SELECT partner_id AS dc_id, partner_name, node, state, sales_rep_email, abm, rbm
FROM dc_mapping_club_scheme
"""

SQL_DC_CLUB_SLABS_3G = """
SELECT club, turnover_ll, turnover_ul, tour, tod_percent, score
FROM dc_club_slabs
"""

# Qualifying_Turnover, confirmed 2026-08-13 (see DC_CLUB_TIER_TABLE) -- the scheme's own
# definition: Gross Business/Sales (order_value) within the scheme's calendar-year
# validity (2026-01-01 to 2026-12-31, confirmed in Terms & Conditions item 1), status=
# 'confirmed' only -- checked coupon_analysis's real distinct status values live rather
# than reusing SQL_ORDERS_3D's status='processed' rule (a DIFFERENT table's convention,
# confirmed live to not even exist as a value here): coupon_analysis's own statuses are
# cancelled/failed/draft/confirmed, so 'confirmed' is this table's real-order filter.
# coupon_analysis's coupon_applied_flag is separately confirmed unusable for anything --
# always 'true' regardless of whether a coupon actually applied.
#
# HONEST PARTIAL EXCLUSION, not the full T&C list -- checked the real product_category/
# product_sub_category/product_name values live before writing this filter (never guess
# spelling from prose): only 3 of the T&C's 5 excluded categories are reliably
# identifiable in this schema:
#   - Crop Nutrition -> WSF sub-category: exact category+sub_category match, confident.
#   - Tools & Machinery: exact category match (both its sub-categories), confident.
#   - Cattle Feed Khurak/Chokar: Dairy Input/Cattle Feed sub-category has no finer-grained
#     category field, but product_name values are a consistent, matchable pattern
#     ("DeHaat Khurak ...", "Dehaat chokar ...") -- confident via ILIKE.
# NOT excluded, and NOT reliably computable from this schema: "OP/Certified" field-crop
# varieties (Paddy/Wheat/Soybean/Pulses/Groundnut/Forages) -- product names carry no
# consistent OP-vs-Hybrid marker ("Hy Paddy", "Res Paddy", "Certified Wheat" all appear
# for what should be different varieties, live-checked, not distinguishable by a safe
# pattern without risking both false inclusions and false exclusions); and "Open
# Marketing Liquidation/Clearance Sales" -- no category, sub-category, or coupon field in
# the confirmed schema corresponds to this at all. Qualifying_Turnover below is therefore
# an honest OVER-estimate against the full T&C definition, flagged as such in
# normalize_dc_club() (Club_Turnover_Partial_Exclusion), not silently treated as exact --
# per GR-24 (a confirmed, DC-joinable source is not a confirmed formula).
# Scheme's own confirmed calendar-year validity (Terms & Conditions item 1). Named here,
# not just inlined in the SQL below, so dc_club_scheme_window_expired() can't drift out of
# sync with the query's WHERE clause -- once DC_CLUB_SCHEME_WINDOW_END passes, the query
# below silently returns zero rows for every DC; that check makes the rollover visible via
# the Exceptions system instead of every Qualifying_Turnover just quietly going to None.
DC_CLUB_SCHEME_WINDOW_START = "2026-01-01"
DC_CLUB_SCHEME_WINDOW_END = "2027-01-01"  # exclusive upper bound

SQL_DC_CLUB_QUALIFYING_TURNOVER_3G = f"""
SELECT partner_id AS dc_id, SUM(order_value) AS qualifying_turnover
FROM coupon_analysis
WHERE status = 'confirmed'
  AND created_at >= '{DC_CLUB_SCHEME_WINDOW_START}' AND created_at < '{DC_CLUB_SCHEME_WINDOW_END}'
  AND NOT (
        (product_category = 'Crop Nutrition' AND product_sub_category = 'WSF')
     OR (product_category = 'Tools & Machinery')
     OR (product_sub_category = 'Cattle Feed' AND (product_name ILIKE '%khurak%' OR product_name ILIKE '%chokar%'))
  )
GROUP BY partner_id
"""


def dc_club_scheme_window_expired(as_of_date: str) -> bool:
    """True once as_of_date has moved past the DC Club scheme's confirmed calendar-year
    validity window -- see DC_CLUB_SCHEME_WINDOW_START/END above."""
    return as_of_date >= DC_CLUB_SCHEME_WINDOW_END


def load_live_sources(client: MetabaseClient) -> Tuple[Dict[str, Table], Exceptions]:
    """Sources 1, 3, 4. Returns empty tables (with an Exceptions_Report entry) per query
    that couldn't run, instead of raising -- a local-only run should still complete."""
    exc = Exceptions(utc_now_iso())
    results: Dict[str, Table] = {}
    queries = [
        # SE_Visit_Log_1a (SQL_SE_VISIT_LOG_1A) removed 2026-08-12 -- confirmed zero
        # downstream consumers (never read from `live` anywhere in this file or
        # planning/), yet the single most expensive query in the whole pipeline: 2.07M
        # pathik_report rows, ~80s/run (~40% of the entire live-query phase). If a real
        # consumer for Source 1a's visit_email/number_of_visits ever gets built, re-add
        # it filtered/aggregated server-side rather than pulling 2M raw rows again.
        ("Task_Nodes_1b", INPUT_BACKEND_DB_ID, SQL_TASK_NODES_1B),
        ("Geo_Mapping_1c", REDSHIFT_DB_ID, SQL_GEO_MAPPING_1C),
        ("Attendance_3a", INPUT_BACKEND_DB_ID, SQL_ATTENDANCE_3A),
        ("Active_Roster_4", INPUT_BACKEND_DB_ID, SQL_ACTIVE_ROSTER_4),
        ("Sales_Transactions_3d", REDSHIFT_DB_ID, SQL_SALES_TRANSACTIONS_3D),
        ("Outstanding_3d", REDSHIFT_DB_ID, SQL_OUTSTANDING_3D),  # dc_datamart, see SQL_OUTSTANDING_3D docstring
        ("Orders_3d", INPUT_BACKEND_DB_ID, SQL_ORDERS_3D),
        ("Liquidation_3d", REDSHIFT_DB_ID, SQL_LIQUIDATION_3D),
        ("Payments_3f", INPUT_BACKEND_DB_ID, SQL_PAYMENTS_3F),
        ("Promise_To_Pay_3j", INPUT_BACKEND_DB_ID, SQL_PROMISE_TO_PAY_3J),
        ("DC_Club_Mapping_3g", REDSHIFT_DB_ID, SQL_DC_CLUB_MAPPING_3G),
        ("DC_Club_Slabs_3g", REDSHIFT_DB_ID, SQL_DC_CLUB_SLABS_3G),
        ("DC_Club_Qualifying_Turnover_3g", REDSHIFT_DB_ID, SQL_DC_CLUB_QUALIFYING_TURNOVER_3G),
    ]
    if not client.configured:
        exc.flag(
            "ALL", "Source1/3/4", "Metabase_Not_Configured",
            "METABASE_URL/METABASE_API_KEY not set -- Sources 1, 3 and 4 skipped this run; "
            "normalized output covers Sources 2/5/6 (local files) only",
        )
        return {name: [] for name, _, _ in queries}, exc

    for name, db_id, sql in queries:
        # Per-query start/done logging -- previously only "4/6 loading Sources 1/3/4..."
        # was logged for all 12 queries combined, so a hang anywhere in this loop gave no
        # indication of which query was stuck (see REDSHIFT_STATEMENT_TIMEOUT_MS above).
        start = time.monotonic()
        logger.info("  4/6 querying %s...", name)
        try:
            results[name] = client.execute_sql(db_id, sql)
            exc.ok(f"Live_Pull_{name}")
            logger.info("  4/6 %s: %d rows in %.1fs", name, len(results[name]), time.monotonic() - start)
        except Exception as e:  # network/permission/SQL errors -> quarantine, don't crash the run
            exc.flag(name, "Source1/3/4", "Live_Pull_Failed", f"{type(e).__name__}: {e}")
            results[name] = []
            logger.warning("  4/6 %s FAILED after %.1fs: %s: %s", name, time.monotonic() - start, type(e).__name__, e)

    return results, exc


# =====================================================================================
# 7. NORMALIZATION -- assemble Attendance / Visits / Sales tables (Sections 4, 7 output schema)
# =====================================================================================

def normalize_attendance(raw_attendance: Table, exc: Exceptions) -> Table:
    out: Table = []
    for row in raw_attendance:
        check_in, check_out = row.get("check_in_time"), row.get("check_out_time")
        status = "Open_Shift" if check_in and not check_out else "Complete"
        if check_in and check_out and str(check_out) < str(check_in):
            exc.flag(row.get("se_user_id"), "Source3a", "Attendance_Out_Before_In", "punch_out before punch_in")
            status = "Invalid"
        else:
            exc.ok("Attendance_Consistency")
        out_zero = is_zero_coord(row.get("check_out_lat"), row.get("check_out_long"))
        out.append(
            {
                "SE_ID": row.get("se_user_id"),
                "Date": standardize_date(check_in),
                "Punch_In_Time": check_in,
                "Punch_Out_Time": check_out,
                "Punch_In_Lat": parse_number(row.get("check_in_lat")),
                "Punch_In_Long": parse_number(row.get("check_in_long")),
                "Punch_Out_Lat": None if out_zero else parse_number(row.get("check_out_lat")),
                "Punch_Out_Long": None if out_zero else parse_number(row.get("check_out_long")),
                "Attendance_Status": status,
            }
        )
    return out


def normalize_visits(active_roster: Table, task_nodes: Table, config_index: Dict[str, Any], exc: Exceptions) -> Table:
    """Visits_Normalized: check-in + type + geo, kept at per-visit grain (Section 4/7)."""
    # visit_type_id -> which fields are expected to be non-null (Section 4 known trap)
    NO_PARTNER_TYPES = {3, 4, 6}  # External Meeting, Node/Warehouse Visit, Lead Generation
    seen_keys = set()
    out: Table = []
    for row in active_roster:
        vtype = row.get("visit_type_id")
        dc_id = normalize_id(row.get("dc_id"))
        if vtype == 1 and dc_id is None:
            exc.flag(row.get("task_id"), "Source3b", "Null_Partner_Anomaly", "DC Visit with null partner_id")
        elif vtype in NO_PARTNER_TYPES:
            if dc_id is not None:
                exc.flag(row.get("task_id"), "Source3b", "Null_Partner_By_Visit_Type", "non-DC visit type carries a partner_id")
            else:
                exc.ok("Null_Partner_By_Visit_Type")

        check_in = row.get("visit_check_in_time")
        se_id = row.get("se_user_id")
        date = standardize_date(row.get("plan_execution_date"))
        key = (se_id, dc_id, date)
        if dc_id is None:
            # NO_PARTNER_TYPES (External Meeting/Node/Lead Gen) legitimately have no
            # dc_id -- an SE can log several of these on the same day, so the
            # (se_id, dc_id, date) key would collide on every one of them and flag
            # correct, routine visits as duplicates. Only DC-tied visits can be real
            # duplicates under this key.
            exc.ok("Duplicate_Visit")
        elif key in seen_keys:
            exc.flag(key, "Source3b", "Duplicate_Visit", "Repeated (SE_ID, DC_ID, Date) in Visits")
        else:
            seen_keys.add(key)
            exc.ok("Duplicate_Visit")

        photo_logged = photo_logged_from_task_details(row.get("task_details"))
        valid_visit = bool(photo_logged) if photo_logged is not None else None
        # Visit_Duration_Minutes -- RESOLVED 2026-08-12 (business decision, confirmed in
        # the updated Data Normalization Agent doc): no per-visit check-out timestamp
        # exists anywhere in the data model (task_management_taskdetails carries one
        # check-in timestamp per visit, nothing else), so every visit gets a flat
        # assumed 45 minutes, used for both planning and BO2 grading. No longer
        # Config_Ambiguous. Since 45 >= the 10-minute valid-visit threshold, BO2's
        # duration condition always passes now -- Valid_Visit_Flag above depends on
        # Photo_Logged alone, which is the confirmed, intended outcome, not a bug.
        zero_coord = is_zero_coord(row.get("visit_check_in_lat"), row.get("visit_check_in_long"))

        out.append(
            {
                "SE_ID": se_id,
                "DC_ID": dc_id,
                "Date": date,
                "Visit_Check_In_Time": check_in,
                "Visit_Duration_Minutes": 45.0,
                "Photo_Logged": photo_logged,
                "Valid_Visit_Flag": valid_visit,
                "Visit_Type_ID": vtype,
                "Visit_Check_In_Lat": None if zero_coord else parse_number(row.get("visit_check_in_lat")),
                "Visit_Check_In_Long": None if zero_coord else parse_number(row.get("visit_check_in_long")),
                "Task_Status": row.get("task_status"),
            }
        )
    return out


def normalize_sales_transactions(
    sales_raw: Table, outstanding_raw: Table, orders_raw: Table, exc: Exceptions
) -> Tuple[Table, Dict[str, Dict[str, Any]]]:
    """Sales_Transactions_Normalized: daily grain, negatives kept & tagged (Section 4/7).
    Also returns dc_financials, one row per DC, built from Orders (23 cols) and
    dc_datamart (the Outstanding replacement, see SQL_OUTSTANDING_3D -- confirmed live
    one row per sap_partner_id, no dedup-by-latest needed the way the old
    customer_management_input_outstanding required). Orders' credit_on_hold/
    credit_on_hold_reason are the confirmed data source for the 6.4 Credit_Blocked
    case-by-case check."""
    out: Table = []
    dc_financials: Dict[str, Dict[str, Any]] = {}

    for row in outstanding_raw:
        dc_id = normalize_id(row.get("dc_id"))
        if not dc_id:
            continue
        # is_active filter (2026-09-01): a DC not marked active in dc_datamart gets no
        # dc_financials entry -- flagged explicitly rather than just silently omitted, so
        # the ~22% of otherwise-eligible (Rank<=6000) DCs this affects show up in
        # Exceptions_Report instead of just a blank Present_Outstanding/Overdue column
        # with no explanation.
        if str(row.get("is_active")).lower() != "true":
            exc.flag(
                dc_id, "dc_datamart", "DC_Datamart_Inactive_Outstanding_Unavailable",
                f"DC {dc_id}: dc_datamart marks this DC is_active={row.get('is_active')!r} -- "
                f"Present_Outstanding/Overdue/aging data withheld, not fabricated",
            )
            continue
        fin = dc_financials.setdefault(dc_id, {})
        fin["Current_Outstanding"] = parse_number(row.get("total_outstanding"))
        fin["Current_Overdue"] = parse_number(row.get("total_overdue"))
        fin["Current_Month_OS"] = parse_number(row.get("current_month_os"))
        fin["OS_1_To_90"] = parse_number(row.get("os_1_to_90"))
        fin["OS_90_Plus"] = parse_number(row.get("os_90_plus"))
        fin["Weighted_Avg_Repayment_Days"] = parse_number(row.get("weighted_avg_repayment_days"))
        fin["Outstanding_Last_Invoice_Date"] = standardize_date(row.get("last_invoice_date"))
        fin["Outstanding_Is_Mismatch"] = bool(row.get("is_mismatch"))
    exc.ok("Outstanding_Datamart_Cast")

    latest_processed_order: Dict[str, Dict[str, Any]] = {}
    latest_order_any_status: Dict[str, Dict[str, Any]] = {}
    for row in orders_raw:
        dc_id = normalize_id(row.get("dc_id"))
        if not dc_id:
            continue
        created = row.get("created_at") or ""
        if dc_id not in latest_order_any_status or created > (latest_order_any_status[dc_id].get("created_at") or ""):
            latest_order_any_status[dc_id] = row
        if row.get("status") == "processed":
            if dc_id not in latest_processed_order or created > (latest_processed_order[dc_id].get("created_at") or ""):
                latest_processed_order[dc_id] = row
    for dc_id, row in latest_processed_order.items():
        fin = dc_financials.setdefault(dc_id, {})
        fin["Last_Order_Date"] = standardize_date(row.get("created_at"))
        fin["Last_Order_Value"] = parse_number(row.get("amount_total"))
    for dc_id, row in latest_order_any_status.items():
        # credit_on_hold can apply even without a processed order -- checked on the DC's
        # most recent order of any status, not just processed ones.
        fin = dc_financials.setdefault(dc_id, {})
        fin["Credit_On_Hold"] = row.get("credit_on_hold") in (True, "true", "t", 1)
        fin["Credit_On_Hold_Reason"] = row.get("credit_on_hold_reason")
        fin["Partner_Finance_Status"] = row.get("partner_finance_status")

    for row in sales_raw:
        order_value = parse_number(row.get("order_value"))
        cancelled = row.get("cancellations_flag") in (True, "true", "True", "t", 1)
        tag = None
        if order_value is not None and order_value < 0:
            tag = "Return_Credit"
            exc.ok("Negative_Billed_Amount_Tagged")
        out.append(
            {
                "Date": standardize_date(row.get("sap_order_date")),
                "GMV": order_value,
                "Cancelled": cancelled,
                "Business_Segment": row.get("business_segment"),
                "Order_Request_ID": row.get("order_request_id"),
                "Amount_Tag": tag,
            }
        )
    return out, dc_financials


def normalize_payments(payments_raw: Table, exc: Exceptions) -> Tuple[Table, Dict[str, str]]:
    """Payments_Normalized (Source 3f): Last_Payment_Date = MAX(created_at) WHERE
    status='SUCCESS', per DC. Join key CONFIRMED live 2026-08-06: SQL_PAYMENTS_3F/
    _sql_payments() now bridge payments_paymenttransaction.customer_id through
    customer_management_customer.id -> .partner_id (sap_partner_id), same pattern as
    Orders -- rows already carry a real dc_id, no longer Join_Key_Unconfirmed."""
    out: Table = []
    last_payment_by_dc: Dict[str, str] = {}
    for row in payments_raw:
        dc_id = normalize_id(row.get("dc_id"))
        if not dc_id:
            continue
        out.append(
            {
                "DC_ID": dc_id,
                "Payment_Date": standardize_date(row.get("created_at")),
                "Status": row.get("status"),
                "Join_Key_Unconfirmed": False,
            }
        )
        if row.get("status") == "SUCCESS":
            date = standardize_date(row.get("created_at"))
            if date and (dc_id not in last_payment_by_dc or date > last_payment_by_dc[dc_id]):
                last_payment_by_dc[dc_id] = date
    exc.ok("Join_Key_Confirmed") if payments_raw else None
    return out, last_payment_by_dc


def normalize_promise_to_pay(promise_raw: Table, exc: Exceptions) -> Dict[str, Dict[str, Any]]:
    """Promise_To_Pay_By_DC (Source 3j, added 2026-09-04) -- see SQL_PROMISE_TO_PAY_3J's
    own docstring for the source/dialect details. One entry per DC (already de-duped to
    the most recent promise server-side, via ROW_NUMBER rn=1): {Promise_Date (ISO date,
    the committed date only, not the timestamp), Promise_Amount (float, 0.0 for the
    confirmed-live '32% zero-amount' pattern -- never None just because the SE didn't
    give a number, that's a real recorded promise of an unspecified amount, not a
    missing promise), Paid_On_Time (bool, computed server-side)}. Does NOT compute
    Kept/Broken/Pending here -- that depends on plan_date, which this normalization
    step doesn't have (see generate_se_daily_plan._promise_status)."""
    out: Dict[str, Dict[str, Any]] = {}
    for row in promise_raw:
        dc_id = normalize_id(row.get("dc_id"))
        promise_date = standardize_date(row.get("promise_date"))
        if not dc_id or not promise_date:
            continue
        amount = parse_number(row.get("promise_amount_raw"))
        out[dc_id] = {
            "Promise_Date": promise_date,
            "Promise_Amount": amount if amount is not None else 0.0,
            "Paid_On_Time": row.get("paid_on_time") in (True, "true", "t", 1),
        }
    exc.ok("Promise_To_Pay_Normalized") if promise_raw else None
    return out


def normalize_dc_club(
    club_mapping_raw: Table, club_slabs_raw: Table, qualifying_turnover_raw: Table,
    dc_financials: Dict[str, Dict[str, Any]], exc: Exceptions,
) -> Table:
    """DC_Club_Normalized (Source 3g). Presence in dc_mapping_club_scheme by partner_id
    is a PLAUSIBLE, not confirmed, proxy for enrollment (no explicit is_member/status
    flag exists) -- unchanged from before.

    Club_Tier IS now computed here (confirmed 2026-08-13, reversing the earlier "belongs
    to the scoring/planning layer" deferral -- the signed business policy makes this a
    normalization-time fact, not a per-request one): DC_CLUB_TIER_TABLE (the confirmed
    7-tier structure, superseding dc_club_slabs' generic schema) matched against
    Qualifying_Turnover (SQL_DC_CLUB_QUALIFYING_TURNOVER_3G, an honest PARTIAL exclusion
    -- see that query's own docstring for exactly which T&C exclusion categories are and
    aren't reliably computable), gated by Outstanding_Cleared (dc_financials'
    Current_Outstanding <=0 -- per the confirmed rule "a DC hitting its turnover
    threshold is still ineligible if outstanding isn't cleared"). Zone comes from
    STATE_TO_ZONE against this row's own State (affects which zone's tour reward
    applies, not the threshold).

    Slab_Advance_Bonus_Eligible is deliberately left None/unconfirmed -- computing it
    needs a FY25-26 baseline Qualifying_Turnover, and the source documents themselves
    disagree on what "FY25-26"/"FY26-27" even mean here (the Club Slabs sheet states a
    calendar-year validity, 1-Jan-2026 to 31-Dec-2026, while the FY25-26/FY26-27 naming
    implies an April-March Indian fiscal year) -- a genuine ambiguity in the source
    paperwork itself, flagged rather than silently resolved, per the doc's own explicit
    instruction to do exactly that when the paperwork contradicts itself."""
    turnover_by_dc = {normalize_id(r.get("dc_id")): parse_number(r.get("qualifying_turnover")) for r in qualifying_turnover_raw}

    out: Table = []
    partial_exclusion_flagged = False
    for row in club_mapping_raw:
        dc_id = normalize_id(row.get("dc_id"))
        if not dc_id:
            continue
        state = row.get("state")
        zone = STATE_TO_ZONE.get(state) if state else None

        qualifying_turnover = turnover_by_dc.get(dc_id)
        current_outstanding = dc_financials.get(dc_id, {}).get("Current_Outstanding")
        outstanding_cleared = None if current_outstanding is None else current_outstanding <= 0

        # Computed unconditionally, regardless of outstanding_cleared -- confirmed
        # 2026-08-18: an SE talking to a DC that's turnover-qualified but outstanding-
        # blocked needs to know WHICH tier clearing the balance would unlock, not just
        # that they're currently blocked. Club_Tier itself stays hard-gated below (the
        # confirmed rule is turnover AND cleared outstanding, never a tiebreak) -- this
        # is purely informational, never used as if it were the real, applicable tier.
        turnover_tier_match = club_tier_for_turnover(qualifying_turnover)
        eligible_tier_if_cleared = eligible_tod_if_cleared = eligible_reward_if_cleared = None
        if turnover_tier_match:
            eligible_tier_if_cleared, eligible_tod_if_cleared, reward_west, reward_east = turnover_tier_match
            # Zone-specific -- DC_CLUB_TIER_TABLE's tour reward only actually differs
            # West vs East for Bronze (Kerala vs Sikkim); every other tier's reward is
            # identical either way, but looking it up by zone is still correct rather
            # than assuming that always holds. None (not guessed) when this DC's own
            # zone isn't resolvable (state not in STATE_TO_ZONE).
            eligible_reward_if_cleared = reward_west if zone == "West" else reward_east if zone == "East" else None

        club_tier = tod_percent = reward = None
        if outstanding_cleared and turnover_tier_match:
            club_tier, tod_percent, tier_reward_west, tier_reward_east = turnover_tier_match
            # Same zone-specific lookup as Eligible_Tier_Reward_If_Cleared below -- bug
            # fixed 2026-08-19: this tier's own reward was computed into
            # turnover_tier_match but never actually carried into the output row, so a
            # DC that had ALREADY achieved a tier could never be told what it was
            # getting, only DCs still working towards one. Confirmed live: every
            # DC_CLUB_TIER_TABLE tier has a real reward string, this wasn't a "some
            # tiers have none" gap.
            reward = tier_reward_west if zone == "West" else tier_reward_east if zone == "East" else None
        # outstanding_cleared is False or None (unconfirmed) -> Club_Tier stays None,
        # even if turnover alone would qualify -- the confirmed rule is a hard AND, not
        # "turnover, with outstanding as a tiebreak."

        if qualifying_turnover is not None:
            partial_exclusion_flagged = True

        # Only actionable (and only ever surfaced) when it differs from the real,
        # already-applicable Club_Tier -- a DC that's already tiered has nothing to gain
        # from being told what it's already getting.
        still_pending = eligible_tier_if_cleared is not None and eligible_tier_if_cleared != club_tier

        out.append(
            {
                "DC_ID": dc_id,
                "Is_Club_Enrolled": True,  # presence-based, unconfirmed interpretation -- see docstring
                "Enrollment_Basis": "Presence_In_Mapping_Table_Unconfirmed",
                "Qualifying_Turnover": qualifying_turnover,
                "Outstanding_Cleared": outstanding_cleared,
                "Club_Tier": club_tier,
                # This tier's own reward (tour voucher etc.) -- None until club_tier is
                # actually achieved, same gating as TOD_Percent below.
                "Reward": reward,
                # Informational only -- see the comment above where this is computed.
                "Eligible_Tier_If_Outstanding_Cleared": eligible_tier_if_cleared if still_pending else None,
                "Eligible_Tier_TOD_Percent_If_Cleared": eligible_tod_if_cleared if still_pending else None,
                "Eligible_Tier_Reward_If_Cleared": eligible_reward_if_cleared if still_pending else None,
                "Zone": zone,
                "TOD_Percent": tod_percent,
                "Slab_Advance_Bonus_Eligible": None,  # see docstring -- genuine source-doc date ambiguity, not computed
                "Node": row.get("node"),
                "State": state,
            }
        )
    if club_mapping_raw:
        exc.flag("ALL", "Source3g", "Club_Enrollment_Flag_Unconfirmed", "presence-in-table used as enrollment proxy; no explicit is_member/status column exists")
    if partial_exclusion_flagged:
        exc.flag(
            "ALL", "Source3g", "Club_Turnover_Partial_Exclusion",
            "Qualifying_Turnover excludes only Crop Nutrition/WSF, Tools & Machinery, and Cattle Feed "
            "Khurak/Chokar (confirmed reliably matchable in coupon_analysis) -- OP/Certified field-crop "
            "varieties and Open Marketing Liquidation/Clearance sales are NOT excluded (no reliable "
            "product-level signal exists for either in the confirmed schema), so this figure honestly "
            "over-estimates the T&C's full definition, not an exact match",
        )
    return out


def normalize_liquidation(raw: Table, exc: Exceptions) -> Table:
    """Liquidation_Normalized (Source 3d, wired 2026-08-06 from invoice_liquidation_with_pog
    -- see SQL_LIQUIDATION_3D for why this replaced the dead hyperlocal_order pull).

    PER GR-24 (Guardrails sheet): a confirmed, DC-joinable data source is NOT a confirmed
    formula. This function normalizes the raw invoice lines for visibility/future use --
    it deliberately computes no Liquidation_Rate, no grade, and feeds no dc_financials or
    BO score. Do not add scoring here without a signed-off Liquidation_Rate formula."""
    out: Table = []
    for row in raw:
        dc_id = normalize_id(row.get("partner_id"))
        if not dc_id:
            continue
        out.append(
            {
                "DC_ID": dc_id,
                "Invoice_Date": standardize_date(row.get("invoice_date")),
                "Billed_Qty": parse_number(row.get("billed_qty")),
                "Liquidated_Qty": parse_number(row.get("liquidated_qty")),
                "Business_Category": row.get("business_category"),
                "Net_Billed_Amount": parse_number(row.get("net_billed_amount")),
            }
        )
    exc.ok("Liquidation_Normalized") if raw else None
    return out


# =====================================================================================
# 8. VALIDATION / SANITY CHECKS (Section 5)
# =====================================================================================

def run_cross_source_checks(dc_master: Table, geo_mapping: Table, exc: Exceptions) -> None:
    """Geo hierarchy consistency + dual DC-active-status check (Section 5)."""
    geo_by_dc = {g["DC_ID"]: g for g in geo_mapping if g.get("DC_ID")}
    for dc in dc_master:
        geo = geo_by_dc.get(dc["DC_ID"])
        if geo is None:
            continue
        if dc.get("Node") and geo.get("Node") and str(dc["Node"]).strip().lower() != str(geo["Node"]).strip().lower():
            exc.flag(dc["DC_ID"], "CrossSource", "Geo_Mapping_Conflict", f"Node '{dc['Node']}' vs canonical '{geo['Node']}' (1c)")
        else:
            exc.ok("Geo_Hierarchy_Consistency")
        dc["Latitude"], dc["Longitude"] = geo.get("Latitude"), geo.get("Longitude")
        if dc["Latitude"] is None or dc["Longitude"] is None:
            exc.flag(dc["DC_ID"], "Source2", "Geo_Incomplete", "In-scope DC missing resolvable lat/long")
        else:
            exc.ok("Geo_Completeness")


def referential_integrity_check(visits: Table, sales: Table, dc_ids: set, se_ids: set, exc: Exceptions) -> None:
    for v in visits:
        if v["DC_ID"] and v["DC_ID"] not in dc_ids:
            exc.flag(v["DC_ID"], "Visits", "Referential_Integrity", "DC_ID not found in DC Master")
        elif v["DC_ID"]:
            exc.ok("Referential_Integrity_DC")
        if v["SE_ID"] and se_ids and v["SE_ID"] not in se_ids:
            exc.flag(v["SE_ID"], "Visits", "Referential_Integrity", "SE_ID not found in Active Roster")
        elif v["SE_ID"]:
            exc.ok("Referential_Integrity_SE")


def valid_visit_compliance(visits: Table, exc: Exceptions) -> Dict[str, float]:
    """% of check-ins failing the duration/photo rule, per SE (feeds BO2 downstream)."""
    by_se: Dict[str, List[bool]] = defaultdict(list)
    for v in visits:
        if v["Valid_Visit_Flag"] is not None:
            by_se[v["SE_ID"]].append(bool(v["Valid_Visit_Flag"]))
    compliance = {}
    for se_id, flags in by_se.items():
        rate = sum(flags) / len(flags) if flags else 0.0
        compliance[se_id] = rate
        exc.ok("Valid_Visit_Compliance_Logged")
    return compliance


# =====================================================================================
# 9. BO1-BO5 SCORING ENGINE (Config Section 1-5 confirmed formulas)
# =====================================================================================

def completion_multiplier(completion_rate_30d: Optional[float], sample_size: int, min_sample_size: int = 5) -> float:
    """DISABLED 2026-09-04, explicit user request -- no longer called anywhere in the
    scoring pipeline (planning/services.py's score_bo3_outstanding_live_proxy/
    score_bo1_private_label calls no longer pass weight_multiplier at all, so it
    defaults to 1.0/no-op). Confirmed live this "Tier-2 adaptive weighting" concept
    appears NOWHERE in SE_DC_Data_Normalization_Agent_Prompt.docx -- it was a system
    extension never validated against the actual business requirements, silently
    shrinking/inflating every DC's Outstanding/PL score by up to 30% based on the SE's
    own completion history, not the DC's own data. Left defined (not deleted), same as
    ObjectiveCompletionStats/compute_completion_stats, only in case this is ever
    reintroduced later with a confirmed formula -- do not wire this back in without
    that confirmation.

    Original docstring, for reference: maps a trailing-30d completion rate (0.0-1.0)
    linearly onto [0.7, 1.3]: objectives whose assigned tasks are actually getting done
    get a mild boost, chronically-missed ones get mildly dampened, bounded either way so
    one bad/good week can't swing scoring by more than 30%. Neutral (1.0, no-op)
    whenever sample_size is too small to be meaningful -- never reweight off a handful
    of data points."""
    if completion_rate_30d is None or sample_size < min_sample_size:
        return 1.0
    return max(0.7, min(1.3, 0.7 + 0.6 * completion_rate_30d))


def score_bo1_private_label(
    pl_value: float, pl_expected: float, c: BusinessConstants,
    weight_multiplier: float = 1.0, yoy_growth_multiplier: float = 1.0,
) -> Dict[str, Any]:
    """yoy_growth_multiplier (confirmed 2026-08-18, planning.services._yoy_pl_growth_
    multiplier): a second, independent multiplier from YoY PL growth (this fiscal-YTD
    vs the same window last fiscal year) -- a provisional business default (+/-10% band),
    not a confirmed Source 5 formula like weight_multiplier's completion-rate weighting
    is. Multiplies in alongside weight_multiplier, not instead of it -- the two answer
    different questions (recent completion rate vs longer-term growth trend)."""
    if not pl_expected:
        return {"score_pct": None, "grade": None, "reason": "PL_Expected undefined -- Config_Ambiguous"}
    weight_multiplier = max(0.7, min(1.3, weight_multiplier))
    yoy_growth_multiplier = max(0.9, min(1.1, yoy_growth_multiplier))
    pct = (pl_value / pl_expected) * weight_multiplier * yoy_growth_multiplier
    grade = "A" if pct >= c.bo1_grade_a else "B" if pct >= c.bo1_grade_b else "C" if pct >= c.bo1_grade_c else "D"
    reason = f"PL at {pct:.0%} of trailing baseline" if pct >= 0 else f"net PL negative this window ({pct:.0%} of baseline -- returns exceeding new PL billing)"
    if weight_multiplier != 1.0:
        reason += f"; completion-weighted {weight_multiplier:.2f}x"
    if yoy_growth_multiplier != 1.0:
        reason += f"; YoY growth-weighted {yoy_growth_multiplier:.2f}x"
    return {"score_pct": pct, "grade": grade, "reason": reason}


def score_bo2_visits(valid_visits: int, total_in_scope_dcs: int, c: BusinessConstants) -> Dict[str, Any]:
    if not total_in_scope_dcs:
        return {"coverage_pct": None, "grade": None, "reason": "No in-scope DCs to cover"}
    coverage = valid_visits / total_in_scope_dcs
    grade = (
        "A" if coverage >= c.bo2_coverage_grade_a else
        "B" if coverage >= c.bo2_coverage_grade_b else
        "C" if coverage >= c.bo2_coverage_grade_c else "D"
    )
    return {"coverage_pct": coverage, "grade": grade}


def score_bo3_outstanding(current_os: float, last_month_os: float, sales_growth_pct: float, c: BusinessConstants) -> Dict[str, Any]:
    """The literal 3.1-3.6 formula -- Expected_Outstanding needs last month's outstanding
    balance, and no historical/time-series Outstanding source exists anywhere in this
    pipeline (dc_datamart is a current-snapshot table only), so `last_month_os` has no
    real caller yet. Kept dormant, doc-faithful, ready for whenever that source exists --
    see score_bo3_outstanding_live_proxy() for the live-data substitute used today."""
    expected = last_month_os * (1 + sales_growth_pct) if last_month_os is not None else None
    if not expected:
        return {"ratio": None, "grade": None, "reason": "Expected_Outstanding undefined"}
    ratio = max(min(current_os / expected, c.bo3_ratio_cap_high), c.bo3_ratio_cap_low) if expected else None
    grade = "A" if ratio and ratio >= c.bo3_grade_a else "B" if ratio and ratio >= c.bo3_grade_b else "C" if ratio and ratio >= c.bo3_grade_c else "D"
    return {"ratio": ratio, "grade": grade}


def score_bo3_outstanding_live_proxy(
    current_outstanding: Optional[float], current_overdue: Optional[float],
    os_90_plus: Optional[float], c: BusinessConstants, weight_multiplier: float = 1.0,
) -> Dict[str, Any]:
    """NOT the literal 3.1-3.6 formula (see score_bo3_outstanding docstring for why) --
    a live, per-DC substitute built entirely from dc_datamart fields that are actually
    available today: Outstanding_Health_Pct = 1 - (Current_Overdue / Current_Outstanding),
    i.e. what fraction of this DC's balance is NOT overdue. A DC with nothing outstanding
    is the best case (1.0, Grade A) -- there's no payment risk to flag. Reuses the doc's
    own 3.6 grade cutoffs (A>=100%, B>=75%, C>=50%, D<50%) against this health fraction
    for directional consistency (A=best/healthiest, D=worst), not because it's the same
    ratio the doc defines. os_90_plus is surfaced as a `reason` note when present -- a
    real >90-day-overdue balance is worth flagging regardless of the overall ratio, but
    this does NOT change the grade itself (no business-approved downgrade rule for that
    exists yet -- never invented)."""
    if not current_outstanding:
        return {"score_pct": 1.0, "grade": "A", "reason": "no outstanding balance", "basis": "live_proxy_not_3_1_formula"}
    weight_multiplier = max(0.7, min(1.3, weight_multiplier))
    overdue_fraction = min((current_overdue or 0.0) / current_outstanding, 1.0)
    health_pct = max(0.0, min(1.0, (1.0 - overdue_fraction) * weight_multiplier))
    grade = "A" if health_pct >= c.bo3_grade_a else "B" if health_pct >= c.bo3_grade_b else "C" if health_pct >= c.bo3_grade_c else "D"
    reason = f"{overdue_fraction:.0%} of outstanding is overdue"
    if os_90_plus:
        reason += f"; ₹{os_90_plus:,.0f} is 90+ days overdue"
    if weight_multiplier != 1.0:
        reason += f"; completion-weighted {weight_multiplier:.2f}x"
    return {"score_pct": health_pct, "grade": grade, "reason": reason, "basis": "live_proxy_not_3_1_formula"}


def score_bo4_sales_momentum(
    momentum_this: Optional[float], momentum_last_year: Optional[float], business_category: Optional[str], c: BusinessConstants,
) -> Dict[str, Any]:
    """4.1-4.3: Momentum = Total_Sales_This_Period / Total_Working_Days_In_Period, graded
    against Baseline_Momentum x Category_Multiplier (4.4, real per-category values -- see
    BusinessConstants.bo4_category_multipliers). Wired 2026-08-06 from
    invoice_liquidation_with_pog.net_billed_amount/business_category (confirmed live, see
    GR-25). Deliberately excluded from Candidate_DCs (8.12) -- this scores a DC, it never
    selects one; caller-side wiring only stores the result, doesn't qualify against it.

    Baseline CHANGED 2026-09-04, explicit user request: momentum_last_year is the SAME
    30-day window one year ago, not last month (the original "prior 30 days" reading).
    Confirmed live why: a DC with an unusually quiet PRIOR MONTH could show 800%+
    "momentum" that was really just recovering off a temporarily depressed base, while
    its real year-over-year trend was flat or declining -- comparing against the same
    calendar window a year back is a fairer read of genuine growth than comparing
    against whatever the immediately preceding month happened to look like.

    Category multiplier lookup is case-insensitive against the confirmed 4.4 categories.
    An unmapped category (Field Crop -- seasonal factor undefined per the sheet itself,
    or anything else business_category carries that 4.4 never priced) is an honest gap,
    not a guess -- GR-20 requires exactly this "provisional" treatment for Field Crop."""
    category_key = (business_category or "").strip().lower()
    multiplier = c.bo4_category_multipliers.get(category_key)
    if multiplier is None:
        reason = (
            "Field Crop growth multiplier is seasonal/undefined in Source 5 (4.4) -- provisional, pending seasonal table"
            if category_key == "field crop"
            else f"no 4.4 growth multiplier defined for category '{business_category}'"
        )
        return {"score_pct": None, "grade": None, "reason": reason, "basis": "provisional_no_multiplier"}
    if not momentum_last_year or momentum_this is None:
        return {"score_pct": None, "grade": None, "reason": "insufficient sales history for momentum comparison", "basis": "live_from_invoice_liquidation_with_pog"}
    momentum_target = momentum_last_year * multiplier
    if not momentum_target:
        return {"score_pct": None, "grade": None, "reason": "momentum target is zero", "basis": "live_from_invoice_liquidation_with_pog"}
    pct = momentum_this / momentum_target
    grade = "A" if pct >= c.bo4_grade_a else "B" if pct >= c.bo4_grade_b else "C" if pct >= c.bo4_grade_c else "D"
    reason = f"momentum at {pct:.0%} of target ({business_category} x{multiplier}, vs. same period last year)"
    return {"score_pct": pct, "grade": grade, "reason": reason, "basis": "live_from_invoice_liquidation_with_pog"}


def score_bo5_long_term(meetings_held: int, dcs_onboarded: int, c: BusinessConstants) -> Dict[str, Any]:
    """5.5: BO5_Score = 0.50 x Meeting_Score_% + 0.50 x Onboarding_Score_%. Wired
    2026-08-06 -- meetings_held/dcs_onboarded now come from real live data (see
    planning/services.py's _sql_bo5_meetings()/_sql_bo5_first_orders() docstrings for the
    Mega-tier-only meeting-count interpretation this depends on). Grade added here reusing
    BO1's 0.80/0.60/0.40 bands -- Source 5 never defines a BO5 grade cutoff at all, same
    "reuse for directional consistency only" treatment BO3/BO4 already got."""
    meeting_pct = meetings_held / c.bo5_meeting_target_per_month
    onboarding_pct = dcs_onboarded / c.bo5_onboarding_target_per_month
    score = c.bo5_weight_meeting * meeting_pct + c.bo5_weight_onboarding * onboarding_pct
    grade = "A" if score >= c.bo5_grade_a else "B" if score >= c.bo5_grade_b else "C" if score >= c.bo5_grade_c else "D"
    return {"score": score, "grade": grade, "meeting_pct": meeting_pct, "onboarding_pct": onboarding_pct}


def compute_fm_urgency(meetings_held_mtd: int, days_left_in_month: int, c: BusinessConstants) -> Dict[str, Any]:
    """8.11 Layer 0: FM_Urgency tracks whether an SE is falling behind the >=2
    Mega-meetings/month pace (5.3) and needs one scheduled soon rather than left to
    month-end. FM_Gap = target - held_MTD; urgent when the days left, spread evenly
    across the still-needed meetings, drops to Pacing_Buffer_Days or less.

    Wired 2026-08-06 -- shipped compute-and-log only first (verified realistic,
    non-blanket results and proved task generation was unaffected via a byte-identical
    before/after diff), then flipped live the same day per direct instruction: this
    result now feeds generate_se_daily_plan()'s farmer_meeting_scheduled_today
    parameter directly (planning/services.py). Manually_Scheduled_Today (the other half
    of 8.11's FM_Scheduled_Today = FM_Urgency OR Manually_Scheduled_Today) still has no
    usable data source -- the only candidate, task_management_task.visit_purpose_name,
    is unstructured free text (confirmed live) -- and stays honestly unimplemented, not
    guessed via string matching."""
    fm_gap = c.bo5_meeting_target_per_month - meetings_held_mtd
    if fm_gap <= 0:
        return {
            "fm_gap": fm_gap, "fm_days_left": days_left_in_month, "fm_urgency": False,
            "reason": f"already met the {c.bo5_meeting_target_per_month}/month Mega-meeting target ({meetings_held_mtd} held MTD)",
        }
    pacing = days_left_in_month / max(fm_gap, 1)
    urgent = pacing <= c.pacing_buffer_days
    reason = (
        f"{meetings_held_mtd}/{c.bo5_meeting_target_per_month} Mega meetings MTD, {days_left_in_month}d left in month, "
        f"{fm_gap} still needed -- {'URGENT' if urgent else 'on pace'} "
        f"(pacing={pacing:.1f}d/meeting vs {c.pacing_buffer_days}d buffer)"
    )
    return {"fm_gap": fm_gap, "fm_days_left": days_left_in_month, "fm_urgency": urgent, "reason": reason}


def score_se_incentive_wps(
    revenue_pct: float, collection_pct: float, product_mix_pct: float, farmer_activity_pct: float,
    trailing_12mo_sales: float, current_os: float, c: BusinessConstants,
) -> Dict[str, Any]:
    """SE Incentive Policy FY26-27: 6-month Weighted Performance Score + the two
    highest-stakes cross-BO triggers (Revenue<60% cliff, OD>10% payout cap)."""
    if revenue_pct is not None and revenue_pct < c.wps_revenue_cliff_pct:
        revenue_pct, product_mix_pct = 0.0, 0.0
        cliff_triggered = True
    else:
        cliff_triggered = False

    wps = (
        (revenue_pct or 0) * c.wps_weight_revenue
        + (collection_pct or 0) * c.wps_weight_collection
        + (product_mix_pct or 0) * c.wps_weight_product_mix
        + (farmer_activity_pct or 0) * c.wps_weight_farmer_activity
    )
    od_pct = (current_os / trailing_12mo_sales) if trailing_12mo_sales else None
    od_cap_triggered = od_pct is not None and od_pct > c.wps_od_cap_pct
    payout_cap_pct = c.wps_od_capped_payout_pct if od_cap_triggered else None

    return {
        "wps": wps,
        "revenue_cliff_triggered": cliff_triggered,
        "od_pct": od_pct,
        "od_cap_triggered": od_cap_triggered,
        "payout_cap_pct": payout_cap_pct,
    }


# =====================================================================================
# 10. THE 11 AGENT-DETERMINED (DYNAMIC) PARAMETERS -- resolved live, not invented
# =====================================================================================

def resolve_dynamic_parameters(
    dc_master: Table, aop_targets: Table, se_visit_history: Dict[str, List[float]], config_rows: Table,
    constants: Optional["BusinessConstants"] = None,
) -> Dict[str, Any]:
    """Condensed resolution logic table (Section 5 of the doc), computed from real data
    at query time. Each entry is logged for auditability, per 8.5's requirement."""
    constants = constants or BusinessConstants()
    resolved: Dict[str, Any] = {}

    # 1.3 PL growth requirement -- Source 5 (1.3): "not a fixed number -- require x%
    # growth tied to the AOP target." AOP_Target_Normalized (Source 6) carries the AOP
    # target side (GMV_AOP_Month_*/Q2 per node/material) but nothing in this pipeline
    # carries the matching actual/current GMV needed to compute a real growth ratio
    # against it, so x% cannot be derived here -- Config_Ambiguous (GR-20/GR-21: flag,
    # don't guess), not silently dropped like a lookup that just found nothing.
    aop_growth = next((a for a in aop_targets if a.get("Metric")), None)
    resolved["1.3_pl_growth_requirement"] = {
        "value": None,
        "basis": (
            "Config_Ambiguous -- AOP_Target_Normalized (Source 6, Provisional) has AOP targets but no actual/current GMV to compute growth % against"
            if aop_growth else
            "Config_Ambiguous -- no AOP data available to compute growth %"
        ),
    }

    # 2.3 Effort target -- multiplier from SE's trailing 3-month visit trend, once mapping is clean
    effort_targets = {}
    for se_id, visits in se_visit_history.items():
        trend = (sum(visits[-3:]) / len(visits[-3:])) if visits else None
        effort_targets[se_id] = trend
    resolved["2.3_effort_target"] = {"per_se": effort_targets, "basis": "trailing 3-month visit trend"}

    # 2.4 Coverage/effort weighting -- bigger gap gets more weight (computed per SE at score time)
    resolved["2.4_coverage_effort_weighting"] = {"rule": "weight = 1 - min(coverage_pct, effort_pct)/max(coverage_pct, effort_pct)"}

    # 4.4 BO4 growth multipliers -- REAL category-specific values from Source 5 (GR-25),
    # replacing the flat 1.05 figure that matched nothing in the sheet.
    resolved["4.4_bo4_growth_multipliers"] = {
        "value": constants.bo4_category_multipliers,
        "basis": "Source 5 confirmed category-specific multipliers (Cattle Feed x1.20, Crop Nutrition x1.15, Crop Protection x1.20); Field Crop provisional -- seasonal factor undefined (GR-20)",
    }
    # 4.5 RESOLVED 2026-08-06: Source 5 confirms "keep current cut-offs" (Status:
    # Confirmed Default) without enumerating real numbers -- BO4 momentum scoring (wired
    # 2026-08-06 in planning/services.py) reuses BO1's 0.80/0.60/0.40 bands. Still doesn't
    # affect task selection: BO4/Sales stays out of the Daily Task Assignment Formula's
    # Candidate_DCs pool (8.12), per GR-25 -- scored and available, not selection-driving.
    resolved["4.5_bo4_grade_cutoffs"] = {"basis": "Confirmed Default -- reusing BO1's grade bands (Source 5 never enumerated real BO4 numbers)"}

    # 6.1 Excluded DC statuses -- outlier supply-chain cost relative to route/cluster.
    # No live outlier-vs-median computation is wired to this rule yet -- stated as a
    # policy description only (same treatment as the 2.4/6.4/8.8 rule-text entries
    # below), not backed by a per-node NRV computation that nothing here consumes.
    resolved["6.1_excluded_dc_statuses"] = {"basis": "Inactive/closed status + supply-chain cost outlier vs route/cluster median"}

    # 6.4 Full block conditions -- Legal_Hold always blocks; Credit_Blocked/Blacklisted case-by-case
    resolved["6.4_full_block_conditions"] = {"rule": "Legal_Hold -> always block. Credit_Blocked/Blacklisted -> check live OD + SE/DC notes"}

    # 7.3 Tie-break order -- this quarter's top priority first (Outstanding, per Section 9), then default order
    resolved["7.3_tie_break_order"] = {"order": (constants.quarter_top_priority,) + tuple(o for o in constants.default_objective_priority if o != constants.quarter_top_priority)}

    # 7.4 Override rule (a) -- Outstanding=D caps Overall Sales at B, applied by default, flagged for confirmation
    resolved["7.4_override_rule_a"] = {"rule": "Outstanding=D caps Overall_Sales at B", "status": "default_applied_pending_confirmation"}

    # 8.5 RESOLVED 2026-08-06: Source 5 confirms real fixed thresholds (Status:
    # Overridden), not a live-computed default -- Visits: not visited >14 days (=
    # default) -- Outstanding: balance >=Rs20,000 AND overdue >=15 days (overridden
    # from the Rs5,000/30-day default) -- PL: <3 orders in 30 days (= default, still
    # not computable live, see _qualify_pl docstring) -- Long-Term: had PL sales in
    # last 90 days (= default). These exact numbers are already what
    # constants.qualify_outstanding_balance/qualify_outstanding_days_overdue/
    # qualify_visits_days_since encode -- no longer asked before every activation.
    resolved["8.5_qualification_thresholds"] = {
        "qualify_visits_days_since": constants.qualify_visits_days_since,
        "qualify_outstanding_balance": constants.qualify_outstanding_balance,
        "qualify_outstanding_days_overdue": constants.qualify_outstanding_days_overdue,
        "qualify_pl_max_orders_30d": constants.qualify_pl_max_orders_30d,
        "qualify_longterm_pl_lookback_days": constants.qualify_longterm_pl_lookback_days,
        "basis": "Confirmed/Overridden fixed thresholds (Source 5), not live-computed",
    }

    # 8.8 Discount cap -- last-approved ceiling as conservative default, exceptions to human review
    resolved["8.8_discount_cap"] = {"rule": "last-approved category discount ceiling; exceptions routed to human approval"}

    return resolved


# =====================================================================================
# 11. SE DAILY PLAN GENERATION -- matches the doc's Section 10 output shape exactly
#     (extension beyond the normalization agent's own scope: "prepares data, does not
#     generate plans" -- built on top of the normalized tables at the user's request)
# =====================================================================================

# Corrected 2026-08-06 against the real system taxonomy confirmed live in the
# "Visit Type & Purpose Mapping" / "Visit Type to Purpose (System)" sheets -- "Visit" and
# "Call" are NOT real Visit Types in task_management (the real ones are DC Visit, Demo
# Visit, External Meeting, Farmer Meeting, Lead Generation, Node/Warehouse Visit; "Call"
# never existed as one -- matches the new Daily Task Formula dropping Call-type
# generation entirely already). "Private Label Product Promotion" is a FARMER MEETING
# purpose, not a DC Visit purpose -- DC Visit's real purposes are Promise To Bill (P2B),
# "Promise To Pay / Collection" (note the spacing, exact system string), Query
# Resolution, Sale, Stock at DC. PL is distinguished by product tag on a Sale-purpose
# order line, not by a separate Purpose value (per the bridge sheet's own note).
TASK_TYPE_BY_OBJECTIVE = {
    "Visits": "DC Visit", "Outstanding": "DC Visit", "PL": "DC Visit",
    "Sales": "DC Visit", "Liquidation": "DC Visit", "Long-Term": "Farmer Meeting",
}
PURPOSE_BY_OBJECTIVE = {
    "Visits": "Sale", "Outstanding": "Promise To Pay / Collection", "PL": "Sale",
    "Sales": "Sale", "Long-Term": "Farmer Meeting", "Liquidation": "Config_Ambiguous -- no confirmed purpose",
}


@dataclass
class DailyTaskRow:
    """One row per DC per day -- matches the doc's Section 10 15-column output exactly
    (columns 1-3, 5-15 populated here; column 4 Distance is filled by
    sequence_with_distance() since it depends on route order)."""
    Sr_No: int
    DC_Name: Optional[str]
    DC_ID: Optional[str]
    Distance_Km: Optional[float]
    Recommended_Task_Type: str
    Purpose_Of_Visit: str
    Reason_Of_Visit: str
    Last_Visit_Date: Optional[str]
    Days_Since_Last_Visit: Optional[int]
    Present_Outstanding: Optional[float]
    Present_Overdue: Optional[float]
    Last_Order_Date: Optional[str]
    Last_Order_Value: Optional[float]
    Last_Payment_Date: Optional[str]
    YTD_Private_Label: Optional[float]
    DC_Club_Participation: str
    # Not printed columns, kept for traceability/safety -- Objective is which BO(s) drove
    # this DC's inclusion. A DC selected under multiple objectives appears once, tagged
    # with ALL of its matched objectives, comma-joined (see the Objective=",".join(matched)
    # call site) -- not just the highest-ranked one. Downstream consumers already split on
    # comma (see objectives_used below).
    Objective: str = ""
    No_New_Orders: bool = False
    Credit_On_Hold: bool = False
    Credit_On_Hold_Reason: Optional[str] = None
    Estimated_Duration: int = 0
    Priority_Multiplier: float = 1.0
    # Full per-DC BO score dict (2026-09-03), same dc_bo_scores this DC's Priority_Score/
    # qualification already read -- {objective: {"grade": "A"/"B"/"C"/"D"/None, "reason":
    # str, ...score_pct/ratio/coverage_pct, "basis"}}. Previously only partially visible,
    # folded into Reason_Of_Visit's prose and only for objectives that MATCHED (grade_
    # notes) -- this exposes the DC's full score set as-computed, including objectives
    # that were scored but didn't qualify it (e.g. Sales/BO4, deliberately excluded from
    # selection per 8.12/GR-25 but still scored) so the underlying grades are visible
    # even when they weren't the reason this task exists. Only ever contains the
    # objectives dc_bo_scores actually computed for THIS DC (Outstanding/PL/Sales today
    # -- Visits/Long-Term are SE-level, never per-DC scored). None when no BO scores were
    # supplied at all for this DC this run (dc_bo_scores wasn't wired for this call path).
    BO_Scores: Optional[Dict[str, Dict[str, Any]]] = None
    # Numeric composite (2026-09-03, explicit user request: "numerical ranking on the
    # basis of BO scoring") -- unweighted average of whatever score_pct/ratio/
    # coverage_pct values BO_Scores actually has (None entries skipped, not treated as
    # 0). All 3 objectives scored per-DC today (Outstanding/PL/Sales) already share the
    # same "higher = healthier/better-performing" direction, so a straight average is a
    # reasonable single index -- but it IS a straight average across differently-scaled
    # ratios (Outstanding's health_pct is capped 0-1; PL/Sales momentum ratios can
    # exceed 1.0), not a weighted/normalized formula signed off by anyone. Treat as a
    # rough composite, not a confirmed score. None when BO_Scores has no usable value at
    # all for this DC.
    BO_Composite_Score: Optional[float] = None
    # 1 = lowest BO_Composite_Score (worst-performing/most in need of attention) among
    # this SE's OTHER selected tasks the same day -- computed once per day's full task
    # list (see _assign_bo_ranks), not globally across the network. Ties share the same
    # rank (dense ranking) rather than an arbitrary tiebreak. None for a task with no
    # BO_Composite_Score to rank by (e.g. Farmer Meeting tasks, or a DC with no BO
    # scores supplied this run).
    BO_Rank: Optional[int] = None
    # sale_orderrequest.partner_finance_status ("financed"/"non_financed"), from this
    # DC's own most recent order of ANY status (normalize_sales_transactions' latest_
    # order_any_status, same source as Credit_On_Hold above) -- confirmed live 2026-09-03
    # this field genuinely changes over a DC's order history (16.6% of DCs with any
    # finance-status data show at least one financed<->non_financed transition), so this
    # is the DC's CURRENT status, not a static attribute. None when the DC has no order
    # with this field populated at all, not assumed non_financed.
    Finance_Status: Optional[str] = None
    # Promise To Pay tracking (2026-09-04, explicit user request + SE_DC_Data_
    # Normalization_Agent_Prompt.docx update, Source 3j -- task_management_
    # visitpurposedetails, visit_purpose_id=4) -- this DC's MOST RECENT Promise To Pay
    # commitment (older ones are superseded, per direct instruction). Promise_To_Pay_
    # Date/Amount are the raw committed date/amount as logged; Promise_Status is
    # "Pending" (date hasn't arrived yet -- too early to judge), "Kept" (a real payment
    # landed between the promise and its date -- amount not required to match, per
    # direct instruction), "Broken" (date passed, no qualifying payment), or None (no
    # promise on record for this DC at all). See _qualify_outstanding/_promise_status
    # for how Kept/Broken feed back into whether this DC gets selected at all.
    Promise_To_Pay_Date: Optional[str] = None
    Promise_To_Pay_Amount: Optional[float] = None
    Promise_Status: Optional[str] = None
    # Confirmed live 2026-08-06 -- payments_paymenttransaction.customer_id bridges
    # through customer_management_customer.id -> .partner_id, same pattern as Orders.
    Last_Payment_Join_Key_Unconfirmed: bool = False
    # normalize_dc_club()'s full per-DC dict (confirmed 2026-08-19) -- DC_Club_Participation
    # above is a one-line prose summary of this same data; this is the structured form so
    # a caller can render current standing (Club_Tier/Zone/TOD_Percent/Reward) and
    # eligibility-if-outstanding-cleared (Eligible_Tier_If_Outstanding_Cleared/
    # ..._TOD_Percent_If_Cleared/..._Reward_If_Cleared) as distinct UI elements instead
    # of parsing prose. None when club data wasn't available this run.
    Club_Detail: Optional[Dict[str, Any]] = None
    # Real aging bucket from dc_datamart's os_1_to_90/os_90_plus split -- NOT an exact
    # "overdue since <date>" or a literal day count: checked live 2026-08-06, no
    # due-date or days-overdue column exists anywhere in dc_datamart's confirmed schema,
    # only these two aggregate amount buckets. Never invented a fake exact date/day
    # count to fill that gap -- this is the honest, real-data substitute for the doc's
    # own 3.2 aging-bucket concept (0-30/31-60/61-90/90+ collapses to what's actually
    # available: current-month / 1-90 days / 90+ days).
    Overdue_Aging_Bucket: Optional[str] = None
    # dc_datamart.weighted_avg_repayment_days -- a real, live, per-DC figure (already
    # pulled into dc_financials, see SQL_OUTSTANDING_3D), just not previously surfaced
    # here. A single DC-wide average across the whole outstanding balance, NOT a
    # per-bucket day count -- dc_datamart has no field that splits days-overdue by the
    # 0-90/90+ amount buckets (see Overdue_Aging_Bucket's docstring on why that doesn't
    # exist), so this is the one honest days-overdue figure available, not two.
    Avg_Repayment_Days: Optional[float] = None
    # Confirmed 2026-08-18 -- True when any of: this (SE, DC) pair has a live
    # DCVisitStreak.consecutive_misses >= DCVisitStreak.ESCALATION_THRESHOLD (chronic
    # non-execution, same threshold reconcile_outcomes' own escalation note uses),
    # Overdue_Aging_Bucket == "90+ days" (real overdue balance aged past 90 days), or
    # Credit_On_Hold is True. A cross-cutting "cover this one first" signal, not a new
    # objective -- doesn't affect Priority_Score/ranking, purely informational.
    Critical: bool = False
    Critical_Reasons: str = ""


def haversine_km(lat1: Optional[float], lon1: Optional[float], lat2: Optional[float], lon2: Optional[float]) -> Optional[float]:
    if None in (lat1, lon1, lat2, lon2):
        return None
    import math
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi, dlambda = math.radians(lat2 - lat1), math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlambda / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def sequence_with_distance(
    rows: List[DailyTaskRow], dc_by_id: Dict[str, Dict[str, Any]], constants: BusinessConstants,
    punch_in_coords: Optional[Tuple[float, float]] = None,
) -> Tuple[List[DailyTaskRow], bool, str]:
    """Section 10 column 4: sequential distance punch-in -> DC1 -> DC2 -> ..., not
    straight-line to each DC independently. The doc documents two implementation options
    as still undecided: (a) haversine between confirmed lat/long points [used here], or
    (b) attendance_attendance's own google_distance/total_distance_travelled/OSRM-matched
    fields, which may already reflect real road distance -- worth checking before trusting
    (a) over (b). Degrades PER-DC, not all-or-nothing: confirmed live 2026-08-06 that one
    genuinely coordinate-less DC in an otherwise-fully-geocoded 5-task list was discarding
    real distance data for the other 4 -- a single missing DC now only costs that DC's own
    Distance_Km (None), appended after the sequenced/geocoded DCs, rather than blanking
    the whole route."""
    if not rows:
        return rows, False, "list_order_geo_unavailable"

    def _coords(r: DailyTaskRow) -> Tuple[Optional[float], Optional[float]]:
        dc = dc_by_id.get(r.DC_ID, {})
        return dc.get("Latitude"), dc.get("Longitude")

    with_coords = [r for r in rows if None not in _coords(r)]
    without_coords = [r for r in rows if None in _coords(r)]

    if not with_coords:
        for i, r in enumerate(rows, 1):
            r.Sr_No = i
        return rows, False, "list_order_geo_unavailable"

    remaining = list(with_coords)
    route: List[DailyTaskRow] = []
    last_point = punch_in_coords
    total_km = 0.0
    while remaining:
        if last_point is not None:
            remaining.sort(key=lambda r: haversine_km(last_point[0], last_point[1], *_coords(r)) or 1e9)
        nxt = remaining.pop(0)
        leg = haversine_km(last_point[0], last_point[1], *_coords(nxt)) if last_point is not None else 0.0
        total_km += leg or 0.0
        nxt.Distance_Km = round(total_km, 1)
        route.append(nxt)
        last_point = _coords(nxt)

    for r in without_coords:
        r.Distance_Km = None
    route.extend(without_coords)

    for i, r in enumerate(route, 1):
        r.Sr_No = i
    basis = "haversine_from_punch_in" if punch_in_coords else "haversine_from_first_dc_no_punch_in"
    if without_coords:
        basis += f"_partial_{len(without_coords)}_of_{len(rows)}_dcs_missing_geo"
    return route, total_km > constants.daily_travel_cap_km, basis


# =====================================================================================
# 10a. ROUTING AGENT -- Models 1-3 (Routing_Agent_Configuration_Sheet_v5_FINAL, v5)
# =====================================================================================
# Pure algorithmic logic only -- persistence, guardrail filtering (GR-R1 geo, GR-R2
# Legal_Hold), and Origin_Point resolution live in planning/routing.py, which calls
# these. Each builder takes the SAME candidate shape generate_se_daily_plan() already
# produces for its route_selector branch: [{"row": DailyTaskRow, "dc": dict,
# "priority_score": float, "matched": [str]}, ...], plus an origin (lat, lon) tuple, and
# returns a common shape: {"stops": [...], "dropped": [...], "total_distance_km",
# "total_travel_min", "total_visit_min", "priority_score_captured", "feasible",
# "infeasibility_reason"}. "dropped" here only ever means "this model's own
# priority/efficiency cutoff left it out" -- guardrail-level drops (bad geo, Legal_Hold)
# already happened before these are called, per planning/routing.py's docstring.

R3_1_CIRCUITY_FACTOR = 1.4  # R3.1, FINAL: Haversine x 1.4 is the locked primary per-leg distance method.
R1_1_FIELD_MINUTES_CAP = 420  # R1.1, HARD cap (GR-R3)
# R1.2, MINIMUM FLOOR (RE-CONFIRMED 2026-08-12 by the v7 Outcome Example, which reversed
# an earlier 2026-08-12 "ceiling" reading based on a side-file that claimed a correction
# never actually written back into the Routing Agent's own sheets. v7 is decisive: its
# own header states "Min_Travel_Minutes >= 180 (R1.2)" and all 3 of its worked-example
# plans have travel times of 195/220/240 min -- every one comfortably ABOVE 180, none
# below, which is only self-consistent with a floor. Matches v5's own original R1.2/GR-R5
# wording ("cumulative floor", reason code Travel_Floor_Not_Met) too. A route is
# feasible only if SUM(Travel_Time_Leg) >= 180.
R1_2_MIN_TRAVEL_MINUTES = 180
R1_7_MAX_STOPS = 5  # R1.7, HARD cap (GR-R4)
R3_2_DEFAULT_AVG_SPEED_KMPH = 25.0  # R3.2 -- undefined in the sheet; bottom of its own suggested 25-30 km/h range


def resolve_typical_origin(points: List[Tuple[float, float, str]], buffer_km: float = 0.5) -> Optional[Dict[str, Any]]:
    """Routing Agent R0.4 Origin_Point, REWRITTEN 2026-09-04 (explicit user request,
    replacing the single-most-recent-day punch-in) -- confirmed live root cause of a
    real 300km+ routing anomaly (kanhaiya.raj1): trusting one raw GPS reading from one
    day, with zero cross-checking against the SE's actual recent pattern, meant a single
    device glitch or one-off trip produced a route starting from the wrong side of the
    state.

    Takes `points` = one (lat, lon, date) per day over the SE's last N days (N=30 per
    direct instruction), in chronological order, and groups them into location clusters
    using a straight-line (haversine, NOT the 1.4x road-circuity-adjusted distance --
    this is "is this the same physical spot," not "how far would you drive") buffer_km
    radius (500m per direct instruction): a point joins the first existing cluster whose
    running centroid is within buffer_km, else starts a new cluster; centroid is
    recomputed as the plain mean of every point in the cluster after each join (a
    single-pass greedy grouping, not full k-means/DBSCAN -- good enough for "does this
    SE consistently start from roughly one place," doesn't need to be optimal).

    Returns the DOMINANT (most days) cluster's centroid as the SE's "typical" start-of-
    day location, plus enough metadata for a caller to flag when the most recent single
    day disagreed with that pattern (most_recent_point_in_dominant_cluster=False) --
    see planning.services' Origin_Point_Outlier_Overridden exception, which is exactly
    what would have caught kanhaiya.raj1's case before it produced a bad route. None
    when points is empty (no punch-in history at all in the window)."""
    if not points:
        return None
    clusters: List[Dict[str, Any]] = []
    for lat, lon, date in points:
        target = None
        for c in clusters:
            d = haversine_km(lat, lon, c["centroid"][0], c["centroid"][1])
            if d is not None and d <= buffer_km:
                target = c
                break
        if target is None:
            target = {"points": [], "centroid": (lat, lon)}
            clusters.append(target)
        target["points"].append((lat, lon, date))
        n = len(target["points"])
        target["centroid"] = (
            sum(p[0] for p in target["points"]) / n,
            sum(p[1] for p in target["points"]) / n,
        )
    dominant = max(clusters, key=lambda c: len(c["points"]))
    dates = sorted(p[2] for p in dominant["points"])
    return {
        "lat": dominant["centroid"][0], "lon": dominant["centroid"][1],
        "days_in_cluster": len(dominant["points"]), "days_total": len(points),
        "most_recent_date_in_cluster": dates[-1], "least_recent_date_in_cluster": dates[0],
        "most_recent_point_in_dominant_cluster": points[-1] in dominant["points"],
    }


def circuity_distance_km(lat1: Optional[float], lon1: Optional[float], lat2: Optional[float], lon2: Optional[float]) -> Optional[float]:
    """R3.1, FINAL: Distance_Leg(i,j) = Haversine(i,j) x 1.4. Same None-propagation
    convention as haversine_km() -- an incomplete point yields an incomplete distance,
    never a guessed number."""
    straight = haversine_km(lat1, lon1, lat2, lon2)
    return None if straight is None else straight * R3_1_CIRCUITY_FACTOR


def travel_time_minutes(distance_km: Optional[float], avg_speed_kmph: float = R3_2_DEFAULT_AVG_SPEED_KMPH) -> Optional[float]:
    """R3.2: Travel_Time_Leg = Distance_Leg / Avg_Speed_Kmph. avg_speed_kmph has no
    confirmed value anywhere in the Routing Agent's own sheet -- flagged there as "the
    single most important undefined number in the whole sheet," recommending an ops call
    over the 25-30 km/h range for mixed rural/semi-urban roads. Defaults to 25 (the
    slower end of that range) until that call happens -- worth knowing this is no longer
    the "conservative" choice it was under the old ceiling reading: against R1.2's
    RE-CONFIRMED >=180-min FLOOR, a slower assumed speed inflates computed travel time
    for any given distance, making the floor easier to satisfy on paper than an SE's
    actual drive time might support. Neither 25 nor 30 has been re-picked with that in
    mind -- still genuinely undefined, still worth the same ops call the sheet asks for,
    just flagging the direction changed. Callers should flag this Provisional rather
    than treat it as confirmed."""
    if distance_km is None:
        return None
    return (distance_km / avg_speed_kmph) * 60.0


def _candidate_coords(c: Dict[str, Any]) -> Tuple[Optional[float], Optional[float]]:
    dc = c["dc"]
    return dc.get("Latitude"), dc.get("Longitude")


def _route_metrics(stop_candidates: List[Dict[str, Any]], origin: Tuple[float, float], avg_speed_kmph: float) -> Dict[str, Any]:
    """Given candidates already in visit order (not including the origin/return legs),
    computes the closed route (origin -> stop_1 -> ... -> stop_N -> origin)'s per-leg
    and total distance/travel-time/visit-time, and its cumulative Priority_Score
    captured. Shared by all 3 models so they can't drift on how a route's numbers add up."""
    stops: List[Dict[str, Any]] = []
    total_distance = total_travel = total_visit = priority_captured = 0.0
    last_point = origin
    cumulative_minutes = 0.0
    for c in stop_candidates:
        lat, lon = _candidate_coords(c)
        leg_km = circuity_distance_km(last_point[0], last_point[1], lat, lon) or 0.0
        leg_min = travel_time_minutes(leg_km, avg_speed_kmph) or 0.0
        visit_min = c["row"].Estimated_Duration
        total_distance += leg_km
        total_travel += leg_min
        total_visit += visit_min
        priority_captured += c["priority_score"]
        cumulative_minutes += leg_min + visit_min
        stops.append({
            "row": c["row"], "distance_from_prev_km": round(leg_km, 2),
            "travel_time_from_prev_min": round(leg_min, 1), "eta_minutes": round(cumulative_minutes, 1),
        })
        last_point = (lat, lon)
    # Closing leg back to Origin_Point (R3.3: routes are closed).
    return_km = circuity_distance_km(last_point[0], last_point[1], origin[0], origin[1]) or 0.0
    return_min = travel_time_minutes(return_km, avg_speed_kmph) or 0.0
    total_distance += return_km
    total_travel += return_min
    return {
        "stops": stops, "total_distance_km": round(total_distance, 2), "total_travel_min": round(total_travel, 1),
        "total_visit_min": round(total_visit, 1), "priority_score_captured": round(priority_captured, 4),
        "return_leg_km": round(return_km, 2), "return_leg_min": round(return_min, 1),
    }


def _within_caps(metrics: Dict[str, Any]) -> bool:
    """GR-R3 (<=420 total), GR-R4 (<=5 stops, enforced by callers via candidate-set size,
    not here), GR-R5 (>=180 travel-only floor -- RE-CONFIRMED 2026-08-12 by the v7
    Outcome Example, reversing an earlier same-day "ceiling" reading; see R1_2_MIN_TRAVEL_MINUTES)."""
    total_minutes = metrics["total_travel_min"] + metrics["total_visit_min"]
    return metrics["total_travel_min"] >= R1_2_MIN_TRAVEL_MINUTES and total_minutes <= R1_1_FIELD_MINUTES_CAP


def _two_opt(order: List[Dict[str, Any]], origin: Tuple[float, float], avg_speed_kmph: float) -> List[Dict[str, Any]]:
    """Standard 2-opt local search minimizing total route distance, treating the closed
    route's origin as a fixed anchor (only the visit-order among stops is permuted).
    Used by both Model 2 (Distance-Min) and Model 3 (Balanced) to polish their
    construction heuristics -- same improvement pass, different starting tour."""
    if len(order) < 3:
        return order

    def _tour_distance(seq: List[Dict[str, Any]]) -> float:
        return _route_metrics(seq, origin, avg_speed_kmph)["total_distance_km"]

    best = list(order)
    best_dist = _tour_distance(best)
    improved = True
    while improved:
        improved = False
        for i in range(len(best) - 1):
            for j in range(i + 1, len(best)):
                candidate = best[: i] + best[i : j + 1][::-1] + best[j + 1 :]
                dist = _tour_distance(candidate)
                if dist < best_dist - 1e-9:
                    best, best_dist = candidate, dist
                    improved = True
    return best


def _or_opt(order: List[Dict[str, Any]], origin: Tuple[float, float], avg_speed_kmph: float, max_segment: int = 3) -> List[Dict[str, Any]]:
    """R4.2 specifies "2-opt/Or-opt local-search improvement" -- only 2-opt (edge-swaps,
    see _two_opt above) was implemented; this is the missing half. Or-opt relocates a
    short contiguous segment (length 1-3) to a different position in the route, keeping
    the move only if it strictly reduces total distance -- catches improvements 2-opt's
    pure edge-reversal can't reach (e.g. a single out-of-place stop that needs moving,
    not reversing). Run after _two_opt in both Model 2 and Model 3, same "polish the
    construction heuristic" role, a different neighborhood of moves. No special-cased
    skip for "reinserting at the original spot" -- that candidate's distance always
    equals best_dist exactly, so the strict-improvement check already excludes it."""
    if len(order) < 3:
        return order

    def _tour_distance(seq: List[Dict[str, Any]]) -> float:
        return _route_metrics(seq, origin, avg_speed_kmph)["total_distance_km"]

    best = list(order)
    best_dist = _tour_distance(best)
    improved = True
    while improved:
        improved = False
        n = len(best)
        for seg_len in range(1, min(max_segment, n - 1) + 1):
            for i in range(n - seg_len + 1):
                segment = best[i : i + seg_len]
                remainder = best[:i] + best[i + seg_len :]
                for j in range(len(remainder) + 1):
                    candidate = remainder[:j] + segment + remainder[j:]
                    dist = _tour_distance(candidate)
                    if dist < best_dist - 1e-9:
                        best, best_dist = candidate, dist
                        improved = True
                        break
                if improved:
                    break
            if improved:
                break
    return best


def build_route_priority_max(
    candidates: List[Dict[str, Any]], origin: Tuple[float, float], constants: "BusinessConstants",
    avg_speed_kmph: float = R3_2_DEFAULT_AVG_SPEED_KMPH,
) -> Dict[str, Any]:
    """Model 1 (R4.1), recommended primary -- Google OR-Tools Routing Solver, Orienteering
    / Prize-Collecting formulation: single vehicle, closed route (R3.3), one Time
    Dimension capped at Field_Minutes_Available (R1.1=420, GR-R3), a Count dimension
    capped at 5 stops (R1.7, GR-R4), and AddDisjunction(node, penalty=Priority_Score(DC))
    per candidate so every DC is optionally visitable at the cost of its own priority if
    skipped -- maximizing SUM(Priority_Score * Visited) is then minimizing total
    skip-penalty, since arc cost is set to 0 (Model 1 has "No distance cap," per OQ-13 --
    time/count are hard constraints here, never part of the objective).

    R1.2's >=180-min travel floor is deliberately NOT hard-constrained inside the solver
    (unlike R1.1/R1.7 above): GR-R5 asks to "fail safe rather than force artificial
    detours" when the floor can't be met, and forcing a lower bound on the solver's own
    Time Dimension would instead make the whole problem infeasible whenever the
    priority-maximal stop set is geographically clustered -- silently discarding a
    perfectly good, high-priority route rather than showing it with an honest shortfall
    noted. So: solve for max-priority under the real hard caps first, exactly like
    Models 2/3 already do, then check the floor post-hoc via _within_caps() and flag
    (not discard) if the resulting route falls short -- same pattern, same honesty,
    across all 3 models."""
    from ortools.constraint_solver import pywrapcp, routing_enums_pb2

    if not candidates:
        return {
            "stops": [], "dropped": [], "total_distance_km": 0.0, "total_travel_min": 0.0,
            "total_visit_min": 0.0, "priority_score_captured": 0.0, "feasible": True, "infeasibility_reason": "",
        }

    with_coords = [c for c in candidates if None not in _candidate_coords(c)]
    without_coords = [c for c in candidates if None in _candidate_coords(c)]
    dropped = [{"dc_id": c["dc"]["DC_ID"], "reason": "Geo_Incomplete"} for c in without_coords]

    if not with_coords:
        return {
            "stops": [], "dropped": dropped, "total_distance_km": 0.0, "total_travel_min": 0.0,
            "total_visit_min": 0.0, "priority_score_captured": 0.0, "feasible": True,
            "infeasibility_reason": "No candidate had usable geo-coordinates",
        }

    points = [origin] + [_candidate_coords(c) for c in with_coords]
    n = len(points)

    def _travel_min(i: int, j: int) -> int:
        dist = circuity_distance_km(points[i][0], points[i][1], points[j][0], points[j][1])
        minutes = travel_time_minutes(dist, avg_speed_kmph) or 0.0
        return int(round(minutes))

    def _visit_min(i: int) -> int:
        return 0 if i == 0 else int(round(with_coords[i - 1]["row"].Estimated_Duration))

    manager = pywrapcp.RoutingIndexManager(n, 1, 0)
    routing = pywrapcp.RoutingModel(manager)

    # Total-time transit (travel + visit-at-origin-of-leg) -- caps R1.1.
    def total_time_callback(from_index, to_index):
        i, j = manager.IndexToNode(from_index), manager.IndexToNode(to_index)
        return _travel_min(i, j) + _visit_min(i)

    total_time_idx = routing.RegisterTransitCallback(total_time_callback)
    routing.SetArcCostEvaluatorOfAllVehicles(routing.RegisterTransitCallback(lambda a, b: 0))
    routing.AddDimension(total_time_idx, 0, R1_1_FIELD_MINUTES_CAP, True, "TotalTime")

    # No travel-only dimension here -- R1.2's floor is checked post-solve (see the
    # docstring above for why a hard lower-bound constraint isn't used).

    # Stop-count transit -- caps R1.7 (<=5 real stops, excludes the depot).
    def count_callback(from_index, to_index):
        i = manager.IndexToNode(from_index)
        return 0 if i == 0 else 1

    count_idx = routing.RegisterTransitCallback(count_callback)
    routing.AddDimension(count_idx, 0, R1_7_MAX_STOPS, True, "StopCount")

    # Every candidate is optional -- skipping node i costs its own Priority_Score,
    # scaled to an integer (OR-Tools requires integer costs).
    for i, c in enumerate(with_coords, start=1):
        penalty = max(1, int(round(c["priority_score"] * 1000)))
        routing.AddDisjunction([manager.NodeToIndex(i)], penalty)

    search_params = pywrapcp.DefaultRoutingSearchParameters()
    search_params.first_solution_strategy = routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC
    # GR-R9 requires Models 1-3 to always produce the same 3 plans for a fixed candidate
    # pool/date snapshot -- a wall-clock time_limit is machine-load-dependent, so local
    # search can be cut off at a different point on a re-run under contention.
    # solution_limit instead bounds the search by a deterministic COUNT of solutions
    # explored, reached almost instantly given the tiny problem size here (<=5 stops +
    # depot, R1_7_MAX_STOPS). time_limit is kept only as a generous safety net against a
    # genuinely pathological case, not as what normally governs termination.
    search_params.solution_limit = 200
    search_params.time_limit.FromSeconds(30)

    solution = routing.SolveWithParameters(search_params)
    if solution is None:
        # No feasible route at all, not even an empty one -- shouldn't happen since every
        # node is skippable, but fail safe (GR-R5) rather than raise.
        return {
            "stops": [], "dropped": dropped + [{"dc_id": c["dc"]["DC_ID"], "reason": "Capacity_Exceeded"} for c in with_coords],
            "total_distance_km": 0.0, "total_travel_min": 0.0, "total_visit_min": 0.0, "priority_score_captured": 0.0,
            "feasible": False, "infeasibility_reason": "OR-Tools found no feasible solution, including the empty route",
        }

    visited_order: List[Dict[str, Any]] = []
    index = routing.Start(0)
    while not routing.IsEnd(index):
        node = manager.IndexToNode(index)
        if node != 0:
            visited_order.append(with_coords[node - 1])
        index = solution.Value(routing.NextVar(index))

    visited_ids = {c["dc"]["DC_ID"] for c in visited_order}
    dropped += [{"dc_id": c["dc"]["DC_ID"], "reason": "Not_Selected_By_Solver"} for c in with_coords if c["dc"]["DC_ID"] not in visited_ids]

    metrics = _route_metrics(visited_order, origin, avg_speed_kmph)
    feasible = _within_caps(metrics)
    return {
        "stops": metrics["stops"], "dropped": dropped,
        "total_distance_km": metrics["total_distance_km"], "total_travel_min": metrics["total_travel_min"],
        "total_visit_min": metrics["total_visit_min"], "priority_score_captured": metrics["priority_score_captured"],
        "feasible": feasible,
        "infeasibility_reason": "" if feasible else "Travel_Floor_Not_Met: solver's own max-priority route falls short of the 180-min travel floor (or exceeds the 420-min total) -- shown as best available, not forced into artificial detours (GR-R5)",
    }


def _greedy_nearest_neighbor(candidates: List[Dict[str, Any]], origin: Tuple[float, float]) -> List[Dict[str, Any]]:
    remaining = list(candidates)
    order: List[Dict[str, Any]] = []
    last_point = origin
    while remaining:
        remaining.sort(key=lambda c: circuity_distance_km(last_point[0], last_point[1], *_candidate_coords(c)) or 1e9)
        nxt = remaining.pop(0)
        order.append(nxt)
        last_point = _candidate_coords(nxt)
    return order


def build_route_distance_min(
    candidates: List[Dict[str, Any]], origin: Tuple[float, float], constants: "BusinessConstants",
    avg_speed_kmph: float = R3_2_DEFAULT_AVG_SPEED_KMPH,
) -> Dict[str, Any]:
    """Model 2 (R4.2), recommended secondary -- Nearest-Neighbor construction + 2-opt
    local-search improvement over circuity-adjusted distances. Visit_Set =
    TOP-K(Ranked_Pool BY Priority_Score) for the largest K whose resulting closed route
    still satisfies R1.1/R1.2/R1.7 -- tries K from min(5, N) down to 0."""
    with_coords = [c for c in candidates if None not in _candidate_coords(c)]
    without_coords = [c for c in candidates if None in _candidate_coords(c)]
    dropped = [{"dc_id": c["dc"]["DC_ID"], "reason": "Geo_Incomplete"} for c in without_coords]

    by_priority = sorted(with_coords, key=lambda c: -c["priority_score"])
    max_k = min(R1_7_MAX_STOPS, len(by_priority))

    best_metrics = None
    best_k = 0
    computed_by_k: Dict[int, Dict[str, Any]] = {}
    for k in range(max_k, 0, -1):
        top_k = by_priority[:k]
        order = _or_opt(_two_opt(_greedy_nearest_neighbor(top_k, origin), origin, avg_speed_kmph), origin, avg_speed_kmph)
        metrics = _route_metrics(order, origin, avg_speed_kmph)
        computed_by_k[k] = metrics
        if _within_caps(metrics):
            best_metrics, best_k = metrics, k
            break

    if best_metrics is None and by_priority:
        # No K satisfied both R1.2's >=180-min floor and R1.1's <=420-min total cap
        # together. These need DIFFERENT fallbacks -- picking "more stops" (max_k)
        # unconditionally, as before, helps only a floor shortfall (more stops -> more
        # cumulative travel, closer to 180) and actively makes a cap breach WORSE (more
        # stops -> more total time, further past 420). GR-R3's cap is a hard "no bypass"
        # (Plan rejected/re-solved, marked Infeasible), unlike R1.2's floor, which GR-R5
        # explicitly says to fail-safe and show anyway -- so only ever fall back to a K
        # that actually respects the 420-min cap; if none does, show no route at all
        # rather than one that breaches it.
        under_cap_ks = [k for k, m in computed_by_k.items() if (m["total_travel_min"] + m["total_visit_min"]) <= R1_1_FIELD_MINUTES_CAP]
        if under_cap_ks:
            best_k = max(under_cap_ks)
            best_metrics = computed_by_k[best_k]
            feasible = False
            reason = (
                f"Travel_Floor_Not_Met: even using the largest cap-respecting candidate set (K={best_k} of {max_k}), "
                "this route falls short of the 180-min travel floor -- shown as best available (GR-R5)"
            )
        else:
            # Not even K=1 stays inside the 420-min hard cap -- GR-R3 has no bypass, so
            # this is a genuine Infeasible, not a "show the fullest attempt anyway."
            selected_ids: set = set()
            dropped += [{"dc_id": c["dc"]["DC_ID"], "reason": "Capacity_Exceeded"} for c in with_coords]
            return {
                "stops": [], "dropped": dropped, "total_distance_km": 0.0, "total_travel_min": 0.0,
                "total_visit_min": 0.0, "priority_score_captured": 0.0, "feasible": False,
                "infeasibility_reason": "Field_Time_Cap_Exceeded: even a single-stop route exceeds the 420-min field-time cap (GR-R3, hard, no bypass) -- no route shown",
            }
    elif best_metrics is None:
        return {
            "stops": [], "dropped": dropped, "total_distance_km": 0.0, "total_travel_min": 0.0,
            "total_visit_min": 0.0, "priority_score_captured": 0.0, "feasible": True, "infeasibility_reason": "",
        }
    else:
        feasible, reason = True, ""

    selected_ids = {c["dc"]["DC_ID"] for c in by_priority[:best_k]}
    dropped += [{"dc_id": c["dc"]["DC_ID"], "reason": "Capacity_Exceeded"} for c in with_coords if c["dc"]["DC_ID"] not in selected_ids]

    return {
        "stops": best_metrics["stops"], "dropped": dropped,
        "total_distance_km": best_metrics["total_distance_km"], "total_travel_min": best_metrics["total_travel_min"],
        "total_visit_min": best_metrics["total_visit_min"], "priority_score_captured": best_metrics["priority_score_captured"],
        "feasible": feasible, "infeasibility_reason": reason,
    }


def _clarke_wright_order(candidates: List[Dict[str, Any]], origin: Tuple[float, float]) -> List[Dict[str, Any]]:
    """Classic Clarke-Wright Savings construction, adapted to a single vehicle: starts
    with each candidate as its own depot-out-and-back route, merges the pair with the
    highest savings s(i,j) = d(depot,i) + d(depot,j) - d(i,j) into one path whenever
    both are still route endpoints, repeats until every candidate is one merged path,
    then that path becomes the visit order (closing the loop back to depot happens in
    _route_metrics, same as every other model)."""
    if len(candidates) <= 1:
        return list(candidates)

    dist_from_origin = {c["dc"]["DC_ID"]: circuity_distance_km(origin[0], origin[1], *_candidate_coords(c)) or 0.0 for c in candidates}
    by_id = {c["dc"]["DC_ID"]: c for c in candidates}
    ids = list(by_id.keys())

    def d(a: str, b: str) -> float:
        ca, cb = by_id[a], by_id[b]
        return circuity_distance_km(*_candidate_coords(ca), *_candidate_coords(cb)) or 0.0

    savings = sorted(
        ((dist_from_origin[a] + dist_from_origin[b] - d(a, b), a, b) for idx, a in enumerate(ids) for b in ids[idx + 1 :]),
        key=lambda t: -t[0],
    )

    routes: Dict[str, List[str]] = {i: [i] for i in ids}  # dc_id -> the path it currently belongs to (by identity of the list)
    route_of: Dict[str, List[str]] = {i: routes[i] for i in ids}
    for _, a, b in savings:
        ra, rb = route_of[a], route_of[b]
        if ra is rb:
            continue
        # Merge only when both are still endpoints of their (distinct) routes.
        if a not in (ra[0], ra[-1]) or b not in (rb[0], rb[-1]):
            continue
        if ra[-1] != a:
            ra.reverse()
        if rb[0] != b:
            rb.reverse()
        merged = ra + rb
        for node in merged:
            route_of[node] = merged

    final_route = route_of[ids[0]]
    return [by_id[i] for i in final_route]


def build_route_balanced(
    candidates: List[Dict[str, Any]], origin: Tuple[float, float], constants: "BusinessConstants",
    avg_speed_kmph: float = R3_2_DEFAULT_AVG_SPEED_KMPH,
    node_avg_travel_min: Optional[float] = None,
    node_avg_travel_range: Optional[Tuple[float, float]] = None,
    alpha_min: float = 0.4, alpha_max: float = 0.8, global_default_alpha: float = 0.6,
) -> Dict[str, Any]:
    """Model 3 (R4.3), recommended tertiary -- Clarke-Wright Savings construction + 2-opt,
    over a handful of candidate visit-sets (top-K by priority, K=1..5, same sweep as
    Model 2), each scored by Blended_Score = alpha * Normalized_Priority_Captured +
    (1-alpha) * (1 - Normalized_Travel_Time), picking whichever K maximizes it -- the
    highest-scoring FEASIBLE (R1.1/R1.2/R1.7) set wins. alpha is derived per-node from
    historical average travel time (OQ-20, CONFIRMED bounds Alpha_Min=0.4/Alpha_Max=0.8,
    higher historical travel -> LOWER alpha) when node_avg_travel_min and
    node_avg_travel_range are both supplied; falls back to Global_Default_Alpha=0.6
    otherwise -- which is what actually happens in this build today, since no historical
    per-node travel-time aggregation is wired yet (see planning/routing.py). Flagged, not
    silently pretended to be dynamic."""
    if node_avg_travel_min is not None and node_avg_travel_range and node_avg_travel_range[1] > node_avg_travel_range[0]:
        lo, hi = node_avg_travel_range
        normalized = (node_avg_travel_min - lo) / (hi - lo)
        normalized = min(1.0, max(0.0, normalized))
        alpha = alpha_min + (alpha_max - alpha_min) * (1 - normalized)
    else:
        alpha = global_default_alpha

    with_coords = [c for c in candidates if None not in _candidate_coords(c)]
    without_coords = [c for c in candidates if None in _candidate_coords(c)]
    dropped = [{"dc_id": c["dc"]["DC_ID"], "reason": "Geo_Incomplete"} for c in without_coords]

    if not with_coords:
        return {
            "stops": [], "dropped": dropped, "total_distance_km": 0.0, "total_travel_min": 0.0,
            "total_visit_min": 0.0, "priority_score_captured": 0.0, "feasible": True,
            "infeasibility_reason": "" if not candidates else "No candidate had usable geo-coordinates",
            "alpha_used": round(alpha, 3),
        }

    by_priority = sorted(with_coords, key=lambda c: -c["priority_score"])
    max_k = min(R1_7_MAX_STOPS, len(by_priority))

    # Build every K=1..max_k option first, THEN normalize priority/travel across the
    # options actually being compared (min-max, same convention the alpha formula
    # itself uses via Min/Max_Across_All_Nodes) -- normalizing against an unrelated
    # absolute scale (e.g. the full candidate pool's total priority, or the raw 180-min
    # figure) instead structurally punishes every K>1 option, since travel grows with
    # K while priority is compared against candidates that were never in contention.
    raw = []  # (k, metrics, feasible)
    for k in range(1, max_k + 1):
        top_k = by_priority[:k]
        order = _or_opt(_two_opt(_clarke_wright_order(top_k, origin), origin, avg_speed_kmph), origin, avg_speed_kmph)
        metrics = _route_metrics(order, origin, avg_speed_kmph)
        raw.append((k, metrics, _within_caps(metrics)))

    priorities = [m["priority_score_captured"] for _, m, _ in raw]
    travels = [m["total_travel_min"] for _, m, _ in raw]
    p_lo, p_hi = min(priorities), max(priorities)
    t_lo, t_hi = min(travels), max(travels)

    evaluated = []  # (blended_score, k, metrics, feasible)
    for k, metrics, feasible in raw:
        norm_priority = (metrics["priority_score_captured"] - p_lo) / (p_hi - p_lo) if p_hi > p_lo else 1.0
        norm_travel = (metrics["total_travel_min"] - t_lo) / (t_hi - t_lo) if t_hi > t_lo else 0.0
        blended = alpha * norm_priority + (1 - alpha) * (1 - norm_travel)
        evaluated.append((blended, k, metrics, feasible))

    if not evaluated:
        return {
            "stops": [], "dropped": dropped, "total_distance_km": 0.0, "total_travel_min": 0.0,
            "total_visit_min": 0.0, "priority_score_captured": 0.0, "feasible": True, "infeasibility_reason": "",
            "alpha_used": round(alpha, 3),
        }

    feasible_options = [e for e in evaluated if e[3]]
    if feasible_options:
        blended, best_k, best_metrics, feasible = max(feasible_options, key=lambda e: e[0])
        reason = ""
    else:
        # Nothing reaches R1.2's >=180-min floor together with R1.1's <=420-min total
        # cap. These need DIFFERENT fallbacks -- picking whichever K got closest to the
        # floor (highest travel time) unconditionally, as before, actively picks a WORSE
        # option when the real problem is a cap breach (more travel -> further past 420,
        # not closer to feasible). GR-R3's cap is hard/"no bypass" (Infeasible, not
        # shown), unlike R1.2's floor, which GR-R5 says to fail-safe and show anyway --
        # so only ever fall back to a K that respects the 420-min cap.
        under_cap = [e for e in evaluated if (e[2]["total_travel_min"] + e[2]["total_visit_min"]) <= R1_1_FIELD_MINUTES_CAP]
        if under_cap:
            blended, best_k, best_metrics, feasible = max(under_cap, key=lambda e: e[2]["total_travel_min"])
            reason = f"Travel_Floor_Not_Met: no cap-respecting candidate set reaches the 180-min travel floor -- shown is the closest (K={best_k} of {max_k}), per GR-R5"
        else:
            dropped += [{"dc_id": c["dc"]["DC_ID"], "reason": "Capacity_Exceeded"} for c in with_coords]
            return {
                "stops": [], "dropped": dropped, "total_distance_km": 0.0, "total_travel_min": 0.0,
                "total_visit_min": 0.0, "priority_score_captured": 0.0, "feasible": False,
                "infeasibility_reason": "Field_Time_Cap_Exceeded: even K=1 exceeds the 420-min field-time cap (GR-R3, hard, no bypass) -- no route shown",
                "alpha_used": round(alpha, 3),
            }

    selected_ids = {c["dc"]["DC_ID"] for c in by_priority[:best_k]}
    dropped += [{"dc_id": c["dc"]["DC_ID"], "reason": "Capacity_Exceeded"} for c in with_coords if c["dc"]["DC_ID"] not in selected_ids]

    return {
        "stops": best_metrics["stops"], "dropped": dropped,
        "total_distance_km": best_metrics["total_distance_km"], "total_travel_min": best_metrics["total_travel_min"],
        "total_visit_min": best_metrics["total_visit_min"], "priority_score_captured": best_metrics["priority_score_captured"],
        "feasible": feasible, "infeasibility_reason": reason, "alpha_used": round(alpha, 3),
    }


# =====================================================================================
# 10b. PLAN B -- BEAT PLANNING / CLUSTER-BASED MODEL
# (Routing_agent/Beat_Planning_Routing_Agent_Cluster_Model.xlsx, confirmed 2026-08-28)
#
# A genuinely different mode from Models 1-3 above (all "Plan A" -- sequential
# priority/distance/balanced route construction over the full candidate pool), not a
# 4th variant of the same algorithm. Inverts the framing from classic TSP/VRP (minimize
# distance, visit everyone) to Prize-Collecting/Orienteering (maximize collected BO
# score within a fixed daily travel budget) -- the source workbook states this
# explicitly as its central design tension (its Section 4).
#
# Three stages, matching the workbook's own section numbering:
#   Stage 1 (3.1-3.3): Territory Clustering -- partition candidates into density-
#     balanced clusters so no cluster is disproportionately sparse or dense.
#   Stage 2: Cumulative BO Score (CBS) per cluster.
#   Stage 3: greedy score-per-km cluster selection under the 80km/180min budget.
#
# Two honest scope notes, both flagged rather than silently glossed over (same
# convention as build_route_balanced's alpha_used flag above):
#   - Stage 1 clustering: the workbook names sklearn-style "density-aware k-means /
#     DBSCAN variants, refined using OR-Tools' distance matrix" as algorithm
#     candidates -- neither sklearn nor an OR-Tools distance-matrix refinement is wired
#     into this build. What runs here is a deterministic greedy nearest-neighbor
#     agglomeration bounded by the same two normalization goals the workbook states
#     (comparable BO-count per cluster; capped intra-cluster travel) -- functionally
#     equivalent for this purpose, not a literal DBSCAN/k-means implementation.
#   - Stage 3 selection: the workbook is explicit that ITS OWN documented algorithm is
#     "the transparent/explainable approximation," separate from a hypothetical
#     "production" Prize-Collecting VRP solve via OR-Tools that "can find combinations
#     the greedy pass would miss." This implements exactly the workbook's own
#     documented (greedy, ordered-criteria) algorithm -- Stage 3's 5 rules, verbatim --
#     not the hypothetical fuller OR-Tools formulation it gestures at but never specifies.
#   - potential_weight_i (Stage 2): the workbook sources this from "Focus Product
#     Cohort relevance, via the 2A->2B->3 cohort API chain" -- that chain
#     (planning.product_cohort) is opt-in per PlanRun and not always available. Defaults
#     to a neutral 1.0 when no Focus Product Cohort data was supplied for this run,
#     exactly like build_route_balanced's alpha_used falling back to
#     global_default_alpha when no per-node travel data exists -- flagged, not
#     pretended to be live.
# =====================================================================================

PLAN_B_MAX_DAILY_DISTANCE_KM = 80.0    # Section 5 Constraints -- standard-case budget; Stage 3's conditional ceiling check (Sheets 3/5/10) reclassifies a cluster exceeding this as Exceptional (BO Rule) instead of dropping it
PLAN_B_MAX_DAILY_TRAVEL_MINUTES = 180.0  # Section 5 -- hard ceiling for the standard case only; soft/informational once a cluster is reclassified Exceptional (Sheet 10)
PLAN_B_MAX_INTRA_CLUSTER_DISTANCE_KM = PLAN_B_MAX_DAILY_DISTANCE_KM * 0.45  # Section 3.2: "no more than ~40-50% of the 80km budget just to traverse internally" -- midpoint of that stated range
PLAN_B_TARGET_CLUSTER_SIZE = 6          # not numerically specified by the workbook ("comparable count of BOs per km2, until BO-count-per-cluster converges within a target band") -- a mid-sized daily-beat count, flagged as a chosen default, not a confirmed figure
PLAN_B_RECENCY_DECAY_RATE = 1.0 / 30.0  # Edge Case #2/#11: bounded, smooth decay, full cycle within ~30 days
PLAN_B_RECENCY_DECAY_CAP = 2.0          # Edge Case #2: caps decay so a cluster can't be weighted away forever
PLAN_B_NEW_BO_RECENCY_WEIGHT = PLAN_B_RECENCY_DECAY_CAP  # Edge Case #6: a BO with no visit history gets max recency urgency, not zero


def _cluster_candidates_by_density(
    candidates: List[Dict[str, Any]],
    max_intra_cluster_km: float = PLAN_B_MAX_INTRA_CLUSTER_DISTANCE_KM,
    target_size: int = PLAN_B_TARGET_CLUSTER_SIZE,
) -> List[List[Dict[str, Any]]]:
    """Stage 1 (3.1-3.3). Greedy nearest-neighbor agglomeration: repeatedly seeds a new
    cluster from the unclustered candidate farthest from every existing cluster centroid
    (spreads seeds out rather than always starting in the same dense pocket), then grows
    it by absorbing its nearest remaining neighbor, one at a time, stopping when either
    the cluster reaches target_size (density normalization -- comparable BO count per
    cluster) or the next absorption would push the cluster's own max-pairwise intra-
    cluster distance past max_intra_cluster_km (distance normalization -- Section 3.2).
    A candidate with no usable coordinates becomes its own singleton cluster (Edge Case
    "Isolated / outlier BO" is handled one level up, at Stage 3's conditional ceiling
    check -- this function only partitions, it doesn't judge whether a cluster is worth
    visiting or standard-vs-Exceptional)."""
    with_coords = [c for c in candidates if None not in _candidate_coords(c)]
    without_coords = [c for c in candidates if None in _candidate_coords(c)]

    remaining = list(with_coords)
    clusters: List[List[Dict[str, Any]]] = []

    def centroid(cluster: List[Dict[str, Any]]) -> Tuple[float, float]:
        lats = [_candidate_coords(c)[0] for c in cluster]
        lons = [_candidate_coords(c)[1] for c in cluster]
        return sum(lats) / len(lats), sum(lons) / len(lons)

    def max_pairwise_km(cluster: List[Dict[str, Any]]) -> float:
        if len(cluster) < 2:
            return 0.0
        worst = 0.0
        for i in range(len(cluster)):
            for j in range(i + 1, len(cluster)):
                d = circuity_distance_km(*_candidate_coords(cluster[i]), *_candidate_coords(cluster[j])) or 0.0
                worst = max(worst, d)
        return worst

    while remaining:
        if not clusters:
            seed = remaining.pop(0)
        else:
            existing_centroids = [centroid(cl) for cl in clusters]
            seed = max(
                remaining,
                key=lambda c: min(circuity_distance_km(*_candidate_coords(c), *ec) or 0.0 for ec in existing_centroids),
            )
            remaining.remove(seed)
        cluster = [seed]

        while remaining and len(cluster) < target_size:
            c_lat, c_lon = centroid(cluster)
            nearest = min(remaining, key=lambda c: circuity_distance_km(c_lat, c_lon, *_candidate_coords(c)) or 1e9)
            dist_to_nearest = circuity_distance_km(c_lat, c_lon, *_candidate_coords(nearest)) or 0.0
            # Density-boundary stop: target_size alone would happily bridge a genuine
            # gap between two natural pockets just to hit the target count (confirmed
            # bug -- a 4+4 two-pocket synthetic test produced one 6-member cluster
            # straddling both pockets before this check existed). If the next nearest
            # candidate is markedly farther than the cluster's own current spread, this
            # is a real density boundary (Section 3.1's "dense pockets vs. sparse
            # pockets" distinction) -- stop growing even though target_size isn't met
            # yet; the 2.0km floor keeps this from firing on a tiny, tight cluster's
            # very first few (near-zero-spread) neighbors.
            current_spread = max_pairwise_km(cluster) if len(cluster) >= 2 else 0.0
            if len(cluster) >= 2 and dist_to_nearest > max(current_spread * 1.75, 2.0):
                break
            trial = cluster + [nearest]
            if max_pairwise_km(trial) > max_intra_cluster_km:
                break
            cluster = trial
            remaining.remove(nearest)

        clusters.append(cluster)

    clusters.extend([[c] for c in without_coords])
    return clusters


def _cumulative_bo_score(cluster: List[Dict[str, Any]], plan_date: str, potential_weight_by_dc: Optional[Dict[str, float]] = None) -> float:
    """Stage 2. CBS(C) = sum over BOs i in cluster of BO_score_i * recency_weight_i *
    potential_weight_i (workbook Section 2, verbatim formula).

    BO_score_i: the candidate's own priority_score -- already the confirmed 0.40/0.35/
    0.25-weighted BO1-BO5 tier score computed upstream (Section 7.2), reused as-is
    rather than recomputed, since the workbook defines it as "the base tier score from
    the BO Scoring Agent," which is exactly what priority_score already is.

    recency_weight_i: smooth bounded decay (Edge Case #11 -- "not a hard on/off"), 1.0
    baseline rising by PLAN_B_RECENCY_DECAY_RATE per day since last visit, capped at
    PLAN_B_RECENCY_DECAY_CAP; a never-visited BO (Days_Since_Last_Visit is None) gets
    the cap directly (Edge Case #6).

    potential_weight_i: Focus Product Cohort relevance when supplied for this run
    (potential_weight_by_dc, keyed by DC_ID); neutral 1.0 otherwise -- see the module
    docstring above on why this isn't always live."""
    total = 0.0
    for c in cluster:
        dc = c["dc"]
        bo_score = c["priority_score"]
        days_since = dc.get("Days_Since_Last_Visit")
        if days_since is None:
            recency_weight = PLAN_B_NEW_BO_RECENCY_WEIGHT
        else:
            recency_weight = min(1.0 + days_since * PLAN_B_RECENCY_DECAY_RATE, PLAN_B_RECENCY_DECAY_CAP)
        potential_weight = 1.0
        if potential_weight_by_dc is not None:
            potential_weight = potential_weight_by_dc.get(dc["DC_ID"], 1.0)
        total += bo_score * recency_weight * potential_weight
    return total


PLAN_B_RANKING_CRITERIA = ("efficiency", "score_max", "distance_min")


def build_route_cluster_based(
    candidates: List[Dict[str, Any]], origin: Tuple[float, float], constants: "BusinessConstants",
    avg_speed_kmph: float = R3_2_DEFAULT_AVG_SPEED_KMPH,
    potential_weight_by_dc: Optional[Dict[str, float]] = None,
    ranking_criterion: str = "efficiency",
    exclude_stop_sets: Optional[List[Tuple[str, ...]]] = None,
) -> Dict[str, Any]:
    """Plan B, main entry point -- Stages 1-3 of Beat_Planning_Routing_Agent_Cluster_
    Model.xlsx, verbatim. Returns the same result-dict shape as the 3 Plan A builders
    (stops/dropped/total_distance_km/total_travel_min/total_visit_min/
    priority_score_captured/feasible/infeasibility_reason) so planning/routing.py's
    persistence code needs no branching to handle this alongside Models 1-3, plus
    "clusters_evaluated" (count) for an honest audit trail of what Stage 3 actually
    chose between.

    ranking_criterion: the workbook's Illustrative Example sheet (its "3-Route
    Comparison Summary", confirmed 2026-08-31 -- initially missed on first read, then
    caught and fixed) runs the SAME greedy, budget-constrained selection logic 3 times,
    each time ranking candidate clusters by a different criterion, mirroring Plan A's
    own 3-model structure (Priority-Max/Distance-Min/Balanced) rather than offering
    only one route:
      "efficiency" (default, Route 1 -- Efficiency-Balanced): Score-per-km, descending.
        The workbook's own recommended default.
      "score_max" (Route 2 -- Score-Maximizing): raw Cumulative BO Score, descending --
        chases the single highest-value clusters regardless of distance efficiency; may
        under-use the budget or crowd out smaller high-efficiency clusters.
      "distance_min" (Route 3 -- Distance-Minimizing): raw closed-tour distance,
        ascending -- covers the most ground for the least travel, at the cost of
        leaving higher-value-but-farther clusters unvisited.
    All 3 share every other rule below unchanged; only Step 2's ranking key (and the
    matching key used by Step 5's individual-BO fallback) changes.

    Stage 3's ordered criteria, applied exactly as the workbook specifies (its own
    Section 3, "Route Selection Criteria (Applied in Order)", updated 2026-08-31 to the
    Conditional Ceiling Model -- Sheets 3/5/9/10):
      1. Conditional ceiling check -- a cluster whose own closed-tour distance/time
         alone exceeds 80km/180min is reclassified Exceptional and routed via the BO
         Rule (Sheet 10, see below) in place of Steps 2-4, never dropped.
      2. Rank by ranking_criterion (see above), descending for efficiency/score_max,
         ascending for distance_min -- cluster's own closed-tour distance/time computed
         via Clarke-Wright + 2-opt + or-opt from Origin_Point, same sequencing
         heuristics Models 2/3 already use.
      3. Greedy accumulation -- add clusters in rank order until the next one would
         push cumulative distance past 80km or cumulative time past 180min.
      4. Tie-break on an equal primary rank -- (a) lower total distance, (b) higher
         recency urgency (proxied by the cluster's own mean Days_Since_Last_Visit,
         None treated as maximally urgent), (c) higher mean potential_weight.
      5. Partial-budget fallback -- if budget remains but no further whole cluster
         fits, fall back to individual-BO selection (ranked by the same
         ranking_criterion) from the unselected remainder, within whatever budget is
         left.
      6. max_daily_tasks cap (2026-08-31, 8.10 reused as Sheet 10 Step 3's own "defined
         maximum operational cap"): after budget-based selection, the stop count is
         trimmed to constants.max_daily_tasks (5 by default, floored at 1), keeping the
         highest-ranked (by ranking_criterion, i.e. distance-and-BO-scoring) candidates
         across the whole merged set. Applies identically to the Exceptional/BO Rule
         path above (there, _bo_rule_key's tier-first order decides who's kept).

    exclude_stop_sets (2026-09-01, forced route distinctness -- explicit user request:
    "force 3 different routes even if 2 are worse"): stop-sets (as DC_ID tuples, in
    sequence) this call must avoid reproducing exactly, passed by the caller from
    earlier ranking_criterion calls' own results. Confirmed live root cause: an
    Exceptional-DC cluster's BO Rule route (see Step 1 above) never depended on
    ranking_criterion at all, so all 3 routes collapsed to the same one whenever a
    single Exceptional cluster dominated the pool -- the majority real cause of
    non-distinct Plan B routes in production data. When the top pick would duplicate an
    excluded set, tries the next-best Exceptional cluster by CBS, then falls through to
    the standard-cluster path if one exists (tagged Exceptional_DC_Deferred_Route_
    Diversity, distinct from the ordinary Exceptional_DC_Deferred_Lower_Priority) rather
    than force-duplicating. Only reduces distinctness-of-last-resort, never invents a
    route from zero candidates -- when every alternative is exhausted, still returns the
    best available (duplicate) result, exactly as before this parameter existed."""
    if ranking_criterion not in PLAN_B_RANKING_CRITERIA:
        raise ValueError(f"ranking_criterion must be one of {PLAN_B_RANKING_CRITERIA}, got {ranking_criterion!r}")
    # 8.10's max_daily_tasks (5, config-driven) reused here as Sheet 10 Step 3's own
    # "defined maximum operational cap" (previously unimplemented -- the BO Rule would
    # visit every BO in the chosen cluster regardless of count) -- floored at 1 so a
    # misconfigured 0 can never zero out a whole day's route.
    task_cap = max(1, constants.max_daily_tasks)
    dropped: List[Dict[str, str]] = []
    with_coords = [c for c in candidates if None not in _candidate_coords(c)]
    without_coords = [c for c in candidates if None in _candidate_coords(c)]
    dropped += [{"dc_id": c["dc"]["DC_ID"], "reason": "Geo_Incomplete"} for c in without_coords]

    if not with_coords:
        return {
            "stops": [], "dropped": dropped, "total_distance_km": 0.0, "total_travel_min": 0.0,
            "total_visit_min": 0.0, "priority_score_captured": 0.0, "feasible": True,
            "infeasibility_reason": "" if not candidates else "No candidate had usable geo-coordinates",
            "clusters_evaluated": 0,
        }

    clusters = _cluster_candidates_by_density(with_coords)

    # Sequence each cluster once (Clarke-Wright + 2-opt + or-opt, same as Models 2/3)
    # so its own closed-tour distance/time/CBS are all known before Stage 3 ranks them.
    scored_clusters = []
    for cluster in clusters:
        order = _or_opt(_two_opt(_clarke_wright_order(cluster, origin), origin, avg_speed_kmph), origin, avg_speed_kmph)
        metrics = _route_metrics(order, origin, avg_speed_kmph)
        cbs = _cumulative_bo_score(cluster, "", potential_weight_by_dc)
        score_per_km = cbs / metrics["total_distance_km"] if metrics["total_distance_km"] > 1e-6 else cbs
        days_since_values = [c["dc"].get("Days_Since_Last_Visit") for c in cluster]
        mean_recency_urgency = (
            PLAN_B_NEW_BO_RECENCY_WEIGHT if any(d is None for d in days_since_values)
            else sum(min(1.0 + d * PLAN_B_RECENCY_DECAY_RATE, PLAN_B_RECENCY_DECAY_CAP) for d in days_since_values) / len(days_since_values)
        )
        mean_potential = (
            sum((potential_weight_by_dc or {}).get(c["dc"]["DC_ID"], 1.0) for c in cluster) / len(cluster)
        )
        scored_clusters.append({
            "cluster": cluster, "order": order, "metrics": metrics, "cbs": cbs,
            "score_per_km": score_per_km, "mean_recency_urgency": mean_recency_urgency, "mean_potential": mean_potential,
        })

    # Step 1 (rewritten 2026-08-31, updated workbook -- Sheets 3/5/9/10, "Conditional
    # Ceiling Model"): a cluster whose own closed-tour distance/time alone exceeds the
    # 80km/180min ceiling is no longer dropped -- it's reclassified "Exceptional" and
    # STILL gets a mandatory route, sequenced by the BO Rule (Sheet 10) instead of
    # Score-per-km efficiency. The workbook's own "DC" (with multiple "BOs" inside it)
    # maps directly onto this function's "cluster" (with multiple candidate DCs inside
    # it) -- every other reference in this codebase already uses DC_ID/DC_Name for the
    # individual outlet, matching the workbook's "BO," so Sheet 10's hierarchy is simply
    # this function's cluster/candidate hierarchy under different names.
    standard, exceptional = [], []
    for sc in scored_clusters:
        m = sc["metrics"]
        if m["total_distance_km"] > PLAN_B_MAX_DAILY_DISTANCE_KM or m["total_travel_min"] > PLAN_B_MAX_DAILY_TRAVEL_MINUTES:
            exceptional.append(sc)
        else:
            standard.append(sc)

    if exceptional:
        # BO Rule (Sheet 10, Steps 1-4): strict priority-tier sequencing, not distance
        # minimization -- Clarke-Wright/2-opt/or-opt are Score-per-km's tools and are
        # deliberately NOT used here, since they'd optimize away exactly the
        # tier-first guarantee the BO Rule exists to provide. This codebase has no
        # separate discrete "BO1..BO5" field per DC (the workbook's own tier concept) --
        # priority_score IS already "the confirmed 0.40/0.35/0.25-weighted BO1-BO5 tier
        # score" (see _cumulative_bo_score's docstring), so it's reused directly as the
        # tier-ordering signal rather than inventing a second, redundant tiering scheme.
        # Recency/distance tie-breaks per Sheet 10 Step 2; distance tie-break uses each
        # candidate's straight-line distance from Origin_Point as a stable static proxy
        # for "nearest," since true nearest-to-previous-stop depends on a sequence this
        # sort is still deciding.
        def _bo_rule_key(c: Dict[str, Any]) -> Tuple[float, float, float]:
            days_since = c["dc"].get("Days_Since_Last_Visit")
            recency = PLAN_B_NEW_BO_RECENCY_WEIGHT if days_since is None else min(1.0 + days_since * PLAN_B_RECENCY_DECAY_RATE, PLAN_B_RECENCY_DECAY_CAP)
            dist_from_origin = circuity_distance_km(origin[0], origin[1], *_candidate_coords(c)) or 0.0
            return (-c["priority_score"], -recency, dist_from_origin)

        # Multiple exceptional clusters can't all fit one SE's one working day (each
        # already exceeds the FULL daily budget alone) -- cover the single most
        # valuable one (highest CBS) this cycle; the rest defer to a future cycle,
        # same "recency decay eventually forces a visit" logic Edge Case #2 already
        # relies on for starvation prevention, not a new mechanism.
        #
        # Forced distinctness (2026-09-01): try every exceptional cluster in CBS order,
        # skipping any whose own (task-cap-trimmed) BO Rule stop-set exactly matches one
        # this call must avoid (exclude_stop_sets, from an earlier ranking_criterion
        # call's result) -- so ranking_criterion calls that would otherwise all collapse
        # onto the same dominant Exceptional cluster instead surface the next-best one as
        # a genuine alternative.
        candidates_by_cbs = sorted(exceptional, key=lambda sc: -sc["cbs"])
        chosen = None
        chosen_bo_rule_order = None
        for candidate in candidates_by_cbs:
            trial_order = sorted(candidate["cluster"], key=_bo_rule_key)[:task_cap]
            trial_stop_set = tuple(c["dc"]["DC_ID"] for c in trial_order)
            if not exclude_stop_sets or trial_stop_set not in exclude_stop_sets:
                chosen, chosen_bo_rule_order = candidate, trial_order
                break

        if chosen is None and standard:
            # Every exceptional cluster's own route is already claimed by an earlier
            # route this cycle, but standard clusters exist -- offer the standard-combo
            # path as this route's genuinely different alternative instead of duplicating
            # an already-shown Exceptional route. Falls through to Step 2 below.
            for sc in exceptional:
                dropped += [{"dc_id": c["dc"]["DC_ID"], "reason": "Exceptional_DC_Deferred_Route_Diversity"} for c in sc["cluster"]]
        else:
            if chosen is None:
                # No standard alternative either -- every exceptional route is claimed
                # and there's nothing else to offer; fall back to the single most
                # valuable one anyway (unavoidable duplicate, same as before this
                # parameter existed).
                chosen = candidates_by_cbs[0]
                chosen_bo_rule_order = sorted(chosen["cluster"], key=_bo_rule_key)[:task_cap]
            for sc in exceptional:
                if sc is not chosen:
                    dropped += [{"dc_id": c["dc"]["DC_ID"], "reason": "Exceptional_DC_Deferred_Lower_Priority"} for c in sc["cluster"]]
            for sc in standard:
                dropped += [{"dc_id": c["dc"]["DC_ID"], "reason": "Standard_Cluster_Deferred_Exceptional_DC_Prioritized"} for c in sc["cluster"]]

            # Sheet 10 Step 3's "defined maximum operational cap" -- _bo_rule_key already
            # sorted tier-first (then recency, then nearest-distance), so chosen_bo_rule_
            # order (computed in the selection loop above) already kept exactly the
            # highest-tier/most-urgent/closest BOs; record what the cap dropped.
            full_bo_rule_order = sorted(chosen["cluster"], key=_bo_rule_key)
            if len(full_bo_rule_order) > task_cap:
                dropped += [{"dc_id": c["dc"]["DC_ID"], "reason": "Exceptional_DC_Daily_Task_Cap"} for c in full_bo_rule_order[task_cap:]]
            metrics = _route_metrics(chosen_bo_rule_order, origin, avg_speed_kmph)
            exceeds = metrics["total_distance_km"] > PLAN_B_MAX_DAILY_DISTANCE_KM or metrics["total_travel_min"] > PLAN_B_MAX_DAILY_TRAVEL_MINUTES
            return {
                "stops": metrics["stops"], "dropped": dropped,
                "total_distance_km": metrics["total_distance_km"], "total_travel_min": metrics["total_travel_min"],
                "total_visit_min": metrics["total_visit_min"], "priority_score_captured": metrics["priority_score_captured"],
                "feasible": True,  # BO Rule routes are never infeasible by design (Sheet 10) -- the ceiling is informational only here
                "infeasibility_reason": (
                    f"Exceptional_DC_BO_Rule: this cluster's own travel ({metrics['total_distance_km']}km/{metrics['total_travel_min']}min) "
                    f"exceeds the standard 80km/180min ceiling -- sequenced by strict BO-tier priority (BO Rule, Beat Planning Sheet 10) "
                    f"instead of Score-per-km efficiency; ceiling shown for visibility only, not enforced" if exceeds else ""
                ),
                "clusters_evaluated": len(scored_clusters),
                "is_exceptional_dc": True,
            }

    if not standard:
        return {
            "stops": [], "dropped": dropped, "total_distance_km": 0.0, "total_travel_min": 0.0,
            "total_visit_min": 0.0, "priority_score_captured": 0.0, "feasible": True,
            "infeasibility_reason": "" if not candidates else "No candidate had usable geo-coordinates",
            "clusters_evaluated": len(scored_clusters),
        }

    # Step 2: rank by ranking_criterion. "efficiency" (Route 1, default) ranks by
    # Score-per-km descending; "score_max" (Route 2) by raw CBS descending, ignoring
    # distance efficiency; "distance_min" (Route 3) by raw closed-tour distance
    # ascending, ignoring value. Step 4's tie-break (lower distance -> higher recency
    # urgency -> higher potential) applies identically after whichever primary key.
    if ranking_criterion == "score_max":
        primary_key = lambda sc: -sc["cbs"]
    elif ranking_criterion == "distance_min":
        primary_key = lambda sc: sc["metrics"]["total_distance_km"]
    else:
        primary_key = lambda sc: -sc["score_per_km"]
    ranked = sorted(
        standard,
        key=lambda sc: (primary_key(sc), sc["metrics"]["total_distance_km"], -sc["mean_recency_urgency"], -sc["mean_potential"]),
    )

    # Per-candidate version of the same ranking_criterion -- shared by Step 5's
    # individual-BO fallback and the final max_daily_tasks cap below, so "according to
    # distance and bo scoring" means the identical thing in both places.
    def _candidate_rank_key(c: Dict[str, Any]) -> float:
        leg_km = circuity_distance_km(origin[0], origin[1], *_candidate_coords(c)) or 0.0
        round_trip_km = leg_km * 2
        if ranking_criterion == "score_max":
            return -c["priority_score"]
        if ranking_criterion == "distance_min":
            return round_trip_km
        return -(c["priority_score"] / round_trip_km) if round_trip_km > 1e-6 else -c["priority_score"]

    # Step 3: greedy accumulation under the 80km / 180min budget.
    selected: List[Dict[str, Any]] = []
    remaining_km = PLAN_B_MAX_DAILY_DISTANCE_KM
    remaining_min = PLAN_B_MAX_DAILY_TRAVEL_MINUTES
    leftover: List[Dict[str, Any]] = []
    for sc in ranked:
        if sc["metrics"]["total_distance_km"] <= remaining_km and sc["metrics"]["total_travel_min"] <= remaining_min:
            selected.append(sc)
            remaining_km -= sc["metrics"]["total_distance_km"]
            remaining_min -= sc["metrics"]["total_travel_min"]
        else:
            leftover.append(sc)

    # Step 5: partial-budget fallback -- individual-BO selection from whatever didn't
    # make it in as a whole cluster, ranked by the SAME ranking_criterion as Step 2,
    # squeezed into whatever budget is left.
    fallback_stops: List[Dict[str, Any]] = []
    loose_candidates = [c for sc in leftover for c in sc["cluster"]]
    if loose_candidates and (remaining_km > 1e-6 or remaining_min > 1e-6):
        for c in sorted(loose_candidates, key=_candidate_rank_key):
            trial_order = fallback_stops + [c]
            trial_metrics = _route_metrics(trial_order, origin, avg_speed_kmph)
            if trial_metrics["total_distance_km"] <= (PLAN_B_MAX_DAILY_DISTANCE_KM - sum(sc["metrics"]["total_distance_km"] for sc in selected)) and \
               trial_metrics["total_travel_min"] <= (PLAN_B_MAX_DAILY_TRAVEL_MINUTES - sum(sc["metrics"]["total_travel_min"] for sc in selected)):
                fallback_stops = trial_order
            else:
                dropped.append({"dc_id": c["dc"]["DC_ID"], "reason": "Capacity_Exceeded"})

    # Every loose_candidates entry is accounted for exactly once above: either folded
    # into fallback_stops, or dropped inside that same loop. When no budget remains at
    # all (the `if loose_candidates and (...)` guard is false), none of them were
    # touched by that loop -- catch those here, keyed by id so a candidate already
    # dropped above is never appended twice (confirmed bug: an earlier version of this
    # function re-walked `leftover` unconditionally afterward and double-counted every
    # candidate the loop above had already dropped).
    already_accounted = {c["dc"]["DC_ID"] for c in fallback_stops} | {d["dc_id"] for d in dropped}
    for c in loose_candidates:
        if c["dc"]["DC_ID"] not in already_accounted:
            dropped.append({"dc_id": c["dc"]["DC_ID"], "reason": "Capacity_Exceeded"})
            already_accounted.add(c["dc"]["DC_ID"])

    # Merge selected whole clusters (each already sequenced) with the fallback partial
    # selection, then re-sequence the combined stop set once so the final route is one
    # coherent tour, not clusters awkwardly concatenated in selection order.
    all_stops_candidates = [c for sc in selected for c in sc["order"]] + fallback_stops
    if not all_stops_candidates:
        return {
            "stops": [], "dropped": dropped, "total_distance_km": 0.0, "total_travel_min": 0.0,
            "total_visit_min": 0.0, "priority_score_captured": 0.0, "feasible": True,
            "infeasibility_reason": "No cluster or individual BO fit inside the 80km/180min daily budget",
            "clusters_evaluated": len(scored_clusters),
        }
    # max_daily_tasks cap (8.10) -- the budget alone doesn't bound stop COUNT, only
    # distance/time, so a tight cluster of many nearby/small-visit-duration DCs could
    # otherwise exceed it. Trims to the task_cap highest-ranked (by the same
    # ranking_criterion, i.e. distance-and-BO-scoring) candidates across the whole
    # merged set, not per-cluster, so a candidate's own priority always wins over which
    # whole cluster it happened to be selected as part of.
    if len(all_stops_candidates) > task_cap:
        all_stops_candidates = sorted(all_stops_candidates, key=_candidate_rank_key)
        capped = all_stops_candidates[:task_cap]
        capped_stop_set = tuple(c["dc"]["DC_ID"] for c in capped)
        # Forced distinctness (2026-09-02) -- extends the 2026-09-01 "force 3 different
        # routes even if 2 are worse" fix past the Exceptional/BO Rule branch, which was
        # the only place exclude_stop_sets was ever consulted. Confirmed live: of 51 real
        # Plan B 3-route runs, 12 converged via THIS path with exclude_stop_sets silently
        # ignored (2 of those -- vatsya.krishnav, rohit.singh1 (recurring across 5 daily
        # runs) -- were genuine GR-R10 cases, pool had room to differ but never did).
        # When the cap-trimmed set duplicates an earlier route, try swapping the single
        # lowest-ranked included candidate for each just-past-the-cap alternative in rank
        # order, keeping the first swap that both fits the 80km/180min budget and yields
        # a genuinely new stop-set. Only ever swaps one seat, never a full re-optimize;
        # falls through to the ordinary (possibly duplicate) result if no swap both fits
        # and differs -- same "never invents a route from zero candidates, still returns
        # the best available" contract as the Exceptional branch's own version of this.
        if exclude_stop_sets and capped_stop_set in exclude_stop_sets:
            for swap_in in all_stops_candidates[task_cap:]:
                trial = capped[:-1] + [swap_in]
                trial_order = _or_opt(_two_opt(_clarke_wright_order(trial, origin), origin, avg_speed_kmph), origin, avg_speed_kmph)
                trial_metrics = _route_metrics(trial_order, origin, avg_speed_kmph)
                trial_stop_set = tuple(s["row"].DC_ID for s in trial_metrics["stops"])
                trial_feasible = (
                    trial_metrics["total_distance_km"] <= PLAN_B_MAX_DAILY_DISTANCE_KM
                    and trial_metrics["total_travel_min"] <= PLAN_B_MAX_DAILY_TRAVEL_MINUTES
                )
                if trial_feasible and trial_stop_set not in exclude_stop_sets:
                    kept_ids = {c["dc"]["DC_ID"] for c in trial}
                    for c in all_stops_candidates:
                        if c["dc"]["DC_ID"] not in kept_ids:
                            reason = "Standard_Cluster_Swap_Route_Diversity" if c is capped[-1] else "Daily_Task_Cap_Exceeded"
                            dropped.append({"dc_id": c["dc"]["DC_ID"], "reason": reason})
                    return {
                        "stops": trial_metrics["stops"], "dropped": dropped,
                        "total_distance_km": trial_metrics["total_distance_km"], "total_travel_min": trial_metrics["total_travel_min"],
                        "total_visit_min": trial_metrics["total_visit_min"], "priority_score_captured": trial_metrics["priority_score_captured"],
                        "feasible": True, "infeasibility_reason": "",
                        "clusters_evaluated": len(scored_clusters),
                    }
        dropped += [{"dc_id": c["dc"]["DC_ID"], "reason": "Daily_Task_Cap_Exceeded"} for c in all_stops_candidates[task_cap:]]
        all_stops_candidates = capped
    final_order = _or_opt(_two_opt(_clarke_wright_order(all_stops_candidates, origin), origin, avg_speed_kmph), origin, avg_speed_kmph)
    final_metrics = _route_metrics(final_order, origin, avg_speed_kmph)
    feasible = final_metrics["total_distance_km"] <= PLAN_B_MAX_DAILY_DISTANCE_KM and final_metrics["total_travel_min"] <= PLAN_B_MAX_DAILY_TRAVEL_MINUTES

    return {
        "stops": final_metrics["stops"], "dropped": dropped,
        "total_distance_km": final_metrics["total_distance_km"], "total_travel_min": final_metrics["total_travel_min"],
        "total_visit_min": final_metrics["total_visit_min"], "priority_score_captured": final_metrics["priority_score_captured"],
        "feasible": feasible,
        "infeasibility_reason": "" if feasible else "Daily_Budget_Exceeded: re-sequenced combined route exceeds 80km/180min after merging clusters -- see Stage 3 docstring",
        "clusters_evaluated": len(scored_clusters),
    }


def _bo_composite_score(bo_scores: Optional[Dict[str, Dict[str, Any]]]) -> Optional[float]:
    """DailyTaskRow.BO_Composite_Score docstring has the full caveat -- unweighted
    average of whatever score_pct/ratio/coverage_pct value each scored objective has,
    None entries skipped. None (not 0.0) when nothing usable exists, so a DC with no
    real BO data doesn't silently look identical to a genuinely perfect score."""
    if not bo_scores:
        return None
    values = []
    for data in bo_scores.values():
        v = data.get("score_pct")
        if v is None:
            v = data.get("ratio")
        if v is None:
            v = data.get("coverage_pct")
        if v is not None:
            values.append(v)
    return (sum(values) / len(values)) if values else None


def _assign_bo_ranks(rows: List["DailyTaskRow"]) -> None:
    """Mutates BO_Rank in place across one SE's full day's task list (2026-09-03,
    explicit user request for a numerical BO-score-based ranking) -- 1 = lowest
    BO_Composite_Score (worst-performing/most in need of attention) among THIS list
    only, dense-ranked (equal composite scores share a rank, no arbitrary tiebreak).
    Rows with no composite score (BO_Composite_Score is None) are left unranked."""
    ranked = sorted((r for r in rows if r.BO_Composite_Score is not None), key=lambda r: r.BO_Composite_Score)
    rank, prev_score = 0, None
    for r in ranked:
        if prev_score is None or r.BO_Composite_Score != prev_score:
            rank += 1
        r.BO_Rank = rank
        prev_score = r.BO_Composite_Score


def generate_se_daily_plan(
    se_id: str,
    se_name: Optional[str],
    plan_date: str,
    dc_candidates: Table,
    bo_scores_by_objective: Dict[str, Dict[str, Any]],
    dynamic_params: Dict[str, Any],
    constants: BusinessConstants,
    attendance_gate_ok: Optional[bool],
    outstanding_by_dc: Optional[Dict[str, float]] = None,
    recent_attempts_by_dc: Optional[Dict[str, int]] = None,
    dc_financials: Optional[Dict[str, Dict[str, Any]]] = None,
    last_payment_by_dc: Optional[Dict[str, str]] = None,
    dc_club_by_id: Optional[Dict[str, Dict[str, Any]]] = None,
    ytd_pl_by_dc: Optional[Dict[str, float]] = None,
    punch_in_coords: Optional[Tuple[float, float]] = None,
    farmer_meeting_scheduled_today: bool = False,
    dc_bo_scores: Optional[Dict[str, Dict[str, Dict[str, Any]]]] = None,
    route_selector: Optional[Callable[..., Dict[str, Any]]] = None,
    consecutive_misses_by_dc: Optional[Dict[str, int]] = None,
    promise_by_dc: Optional[Dict[str, Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    """Section 6 process flow + Section 7/8/10: rank objectives, respect capacity/travel
    caps, apply the confirmed tie-break/override rules, and shape the output exactly as
    the doc's Section 10 (updated) specifies: ONE ROW PER DC PER DAY across 15 columns
    (Sr.No, DC Name, DC ID, Distance, Task Type, Purpose, Reason, Last Visit + days,
    Present Outstanding, Present Overdue, Last Order date/value, Last Payment date,
    YTD PL, Club participation) -- not one row per (DC, objective) task as in the prior
    version. A DC selected under more than one objective is deduplicated to a single row
    tagged with its highest-ranked objective. Attendance gates whether a plan is generated
    at all (Section 3a) -- if the gate can't be evaluated, the plan is still produced but
    tagged Provisional, never silently as if the gate passed.

    route_selector: optional, wired 2026-08-12 for the Routing Agent (planning/routing.py).
    When None (the default -- Step 1's legacy run_pipeline() path), behavior is
    byte-for-byte unchanged from before this parameter existed: Layer 3 greedily fills
    ranked_pool into a single Tasks list capped at max_daily_tasks, then
    sequence_with_distance() sequences it. When provided (the real Django path via
    planning/services.py), Layer 3 instead builds a fully-formed candidate row for
    EVERY ranked_pool entry (no capacity trimming) and hands the whole list to
    route_selector(candidates, se_id, plan_date, punch_in_coords, constants, dc_by_id),
    which returns the shape {"Tasks": [...], "Sequencing_Basis": str,
    "Travel_Cap_Exceeded": bool} -- this is the seam the Routing Agent's Models 1-3
    plug into, without touching anything downstream of Layer 3 (capacity/safety-flag
    reporting, GR-14 checks, the return dict) here."""
    outstanding_by_dc = outstanding_by_dc or {}
    recent_attempts_by_dc = recent_attempts_by_dc or {}
    dc_financials = dc_financials or {}
    last_payment_by_dc = last_payment_by_dc or {}
    dc_club_by_id = dc_club_by_id or {}
    ytd_pl_by_dc = ytd_pl_by_dc or {}
    dc_bo_scores = dc_bo_scores or {}
    consecutive_misses_by_dc = consecutive_misses_by_dc or {}
    promise_by_dc = promise_by_dc or {}

    def _promise_status(dc_id: str) -> Optional[str]:
        """2026-09-04, explicit user request (SE_DC_Data_Normalization_Agent_Prompt.docx
        update, Source 3j -- task_management_visitpurposedetails, visit_purpose_id=4):
        only the DC's MOST RECENT Promise To Pay record matters (per direct instruction
        -- older, superseded promises are ignored). "Kept"/"Broken" only apply once the
        committed promise_date has actually passed relative to plan_date -- a promise
        not yet due is "Pending" and changes nothing (there's no way to judge it yet).
        "Paid" (Kept) is confirmed live-SQL-side (see SQL_PROMISE_TO_PAY_3J) as ANY real
        SUCCESS payment landing between when the promise was logged and its committed
        date -- not required to cover the full promised amount (per direct instruction),
        and a 0-amount promise still counts (per direct instruction -- the doc confirms
        32% of real promise records carry no amount, a genuine pattern, not junk data).
        None when this DC has no promise on record at all."""
        p = promise_by_dc.get(dc_id)
        if not p or not p.get("Promise_Date"):
            return None
        if p["Promise_Date"] >= plan_date:
            return "Pending"
        return "Kept" if p.get("Paid_On_Time") else "Broken"

    header = {
        "SE_ID": se_id, "SE_Name": se_name, "Plan_Date": plan_date,
        "Total_Capacity_Minutes": constants.total_capacity_min,
    }

    if attendance_gate_ok is False:
        return {
            **header, "Tasks": [],
            "Skipped_Reason": "No punch-in recorded for this SE today (Section 3a gating signal)",
            "Data_Confidence": "Live",
        }

    # Layer 0/1 (8.11) -- farmer-meeting day exclusivity. Layer 0's FM_Urgency (monthly
    # meeting pacing) needs a Farmer_Meetings data source that doesn't exist anywhere in
    # this pipeline yet, so it can only be honestly computed as False, never guessed --
    # this gate only fires when the caller explicitly passes
    # farmer_meeting_scheduled_today=True (e.g. from a real scheduling system, once one
    # exists). GR-12 requires that when it fires, the meeting is the ONLY task that day.
    if farmer_meeting_scheduled_today:
        fm_row = DailyTaskRow(
            Sr_No=1, DC_Name=None, DC_ID=None, Distance_Km=None,
            Recommended_Task_Type="Farmer Meeting", Purpose_Of_Visit="Farmer Meeting",
            Reason_Of_Visit="8.11 exclusivity: a farmer meeting is scheduled today -- no other task assigned",
            Last_Visit_Date=None, Days_Since_Last_Visit=None, Present_Outstanding=None,
            Present_Overdue=None, Last_Order_Date=None, Last_Order_Value=None,
            Last_Payment_Date=None, YTD_Private_Label=None, DC_Club_Participation="N/A",
            Objective="Long-Term", Estimated_Duration=constants.meeting_duration_min,
        )
        return {
            **header,
            "Ranked_Objectives": ["Long-Term"],
            "Tasks": [fm_row.__dict__],
            "Capacity_Check": {
                "Call_Minutes_Used": 0, "Call_Minutes_Budget": constants.call_minutes_per_day,
                "Field_Minutes_Used": constants.meeting_duration_min, "Field_Minutes_Budget": constants.field_minutes_per_day,
                "Total_Minutes_Used": constants.meeting_duration_min, "Total_Capacity_Minutes": constants.total_capacity_min,
                "Objectives_Used": 1, "Max_Objectives": constants.max_objectives_per_day,
                "Long_Cycle_Used": 1, "Max_Long_Cycle": constants.max_long_cycle_per_day,
            },
            "Safety_Flags": {"FM_Exclusivity_Applied": True, "GR_12_Note": "Farmer_Meeting_Task is the sole item in this list, per 8.11/GR-12"},
            "Travel": {"Sequencing_Basis": "not_applicable_fm_exclusivity", "Daily_Cap_Km": constants.daily_travel_cap_km, "Cap_Exceeded": False},
            "Data_Confidence": "Live" if attendance_gate_ok is True else "Provisional_No_Attendance_Gate",
        }

    # 7.4 rule (b), confirmed: if EVERY BO is graded D, switch to the all-D override order.
    all_d = all(v.get("grade") == "D" for v in bo_scores_by_objective.values())
    tie_break_order = constants.all_d_override_order if all_d else dynamic_params["7.3_tie_break_order"]["order"]

    def _objective_gap(v: Dict[str, Any]) -> float:
        """7.1's 'Objective_Score DESC' isn't given a literal cross-BO formula in the doc,
        so this reads it as achievement-gap-vs-target (1 - pct achieved): a D-graded
        Outstanding DC ranks ahead of a healthy one, matching the doc's own Section 10
        worked example. Unscored objectives sort last rather than being guessed at."""
        magnitude = v.get("score_pct") if v.get("score_pct") is not None else v.get("coverage_pct") if v.get("coverage_pct") is not None else v.get("ratio") if v.get("ratio") is not None else v.get("score")
        return (1.0 - magnitude) if magnitude is not None else -1.0

    ranked = sorted(
        bo_scores_by_objective.items(),
        key=lambda kv: (-_objective_gap(kv[1]), tie_break_order.index(kv[0]) if kv[0] in tie_break_order else 99),
    )
    top3 = ranked[: constants.max_objectives_per_day]

    # 7.4 rule (a): Outstanding=D caps Overall Sales at B (default-applied, flagged pending confirmation)
    outstanding_entry = bo_scores_by_objective.get("Outstanding", {})
    rule_a_applied = outstanding_entry.get("grade") == "D"

    # 3.7 order-block (No_New_Orders): now driven by the CONFIRMED current_od (actual
    # overdue amount, Source 3d full schema) when dc_financials is supplied -- any real
    # overdue balance triggers the flag. Falls back to the older top-quartile-of-current_os
    # proxy only when dc_financials isn't available, for backward compatibility.
    os_values = sorted((v for v in outstanding_by_dc.values() if v is not None), reverse=True)
    os_top_quartile_cut = os_values[len(os_values) // 4] if os_values else None

    def _no_new_orders(dc_id: str) -> bool:
        fin = dc_financials.get(dc_id)
        if fin and fin.get("Current_Overdue") is not None:
            return fin["Current_Overdue"] > 0
        current_os = outstanding_by_dc.get(dc_id)
        return os_top_quartile_cut is not None and current_os is not None and current_os >= os_top_quartile_cut

    dc_by_id = {dc["DC_ID"]: dc for dc in dc_candidates}
    ranked_dcs = sorted(
        (dc for dc in dc_candidates if dc.get("In_Scope_Flag")),
        key=lambda d: (
            d.get("Cohort") != "Strategic", d.get("Cohort") != "Growth",
            -(d.get("Total_Score") or 0),
            recent_attempts_by_dc.get(d["DC_ID"], 0) >= constants.contact_fatigue_max_attempts,
        ),
    )

    # Layer 2 (8.5/8.12) -- Candidate_DCs = union of the 4 objective qualify-lists,
    # restricted to {BO1,BO2,BO3,BO5} per 8.12's own definition (BO4 Overall Sales and
    # Liquidation are NOT part of this candidate pool under the new formula -- a real
    # behavior change from the old engine, which also generated separate Call-type tasks
    # for Sales). Each qualifying function reads real per-DC data already threaded into
    # this call; where the doc's exact condition needs data this pipeline doesn't have,
    # it degrades honestly (see each docstring) rather than guessing.
    def _qualify_visits(dc: Dict[str, Any]) -> bool:
        d = dc.get("Days_Since_Last_Visit")
        return d is None or d > constants.qualify_visits_days_since

    def _qualify_outstanding(dc: Dict[str, Any]) -> bool:
        """8.5: balance >=20,000 AND overdue >=15 days. The days-overdue leg has no
        confirmed per-DC field in dc_datamart's schema (only aggregate aging buckets /
        weighted_avg_repayment_days) -- balance leg only; Safety_Flags notes this.
        FIXED 2026-08-06: was reading Current_Overdue here (checking overdue amount),
        not Current_Outstanding (the actual balance the docstring/qualify_outstanding_
        balance constant name both describe) -- a DC with a large balance but currently
        $0 overdue could never qualify, defeating the point of a balance leg that's
        independent of overdue status."""
        # Promise To Pay override (2026-09-04, explicit user request) -- checked BEFORE
        # the balance threshold: a Broken promise (committed date passed, no qualifying
        # payment) force-qualifies this DC for collection regardless of current balance
        # -- a broken promise is itself the signal, not the balance. A Kept promise
        # suppresses this objective for the run even if the current balance still
        # crosses the threshold (per direct instruction) -- NOT silently hidden: the
        # DC's real Current_Outstanding, Promise_Status="Kept", and the promise's own
        # date/amount are all still visible in this run's outcome fields, so a reviewer
        # can see a kept-promise-but-still-above-threshold DC for themselves rather than
        # this function silently deciding it's fine. Pending/None promises don't change
        # anything -- falls through to the ordinary balance check.
        status = _promise_status(dc["DC_ID"])
        if status == "Broken":
            return True
        outstanding = dc_financials.get(dc["DC_ID"], {}).get("Current_Outstanding")
        qualifies = outstanding is not None and outstanding >= constants.qualify_outstanding_balance
        if status == "Kept":
            return False
        return qualifies

    def _qualify_pl(dc: Dict[str, Any]) -> bool:
        """8.5's literal '<3 PL orders in 30 days' still isn't computable -- needs a
        per-DC PL order COUNT, and coupon_analysis (the only live PL-transaction source)
        has no confirmed DC-level join key (see SQL_SALES_TRANSACTIONS_3D). Live proxy
        wired 2026-08-06 instead: a DC qualifies if its live PL_Ratio (score_bo1_
        private_label, computed from pathik_report.pl_billed_amount -- recent-30d actual
        vs a trailing-90d baseline, since 1.2's PL_Expected combination-with-AOP method
        is itself still TBD in Source 5) grades C or D -- i.e. genuinely underperforming
        on PL, a more direct read of "does this DC need a PL push" than a raw order
        count would have been anyway. Requires dc_bo_scores["PL"] to be supplied by the
        caller (planning/services.py); returns False, the same fail-safe default as
        before, when no PL score exists for this DC at all."""
        grade = dc_bo_scores.get(dc["DC_ID"], {}).get("PL", {}).get("grade")
        return grade in ("C", "D")

    # _qualify_longterm() removed 2026-08-06. Long-Term (BO5) is deliberately excluded
    # from this DC Visit candidate pool -- 8.11
    # was rewritten in the 2026-08-06 re-sync to make DC Visit and Farmer Meeting
    # mutually exclusive BOTH directions ("if the system proposes a DC Visit day, no
    # Farmer Meeting exists in that day's plan"), and GR-12 was rewritten to match:
    # "Daily_Task_List may never mix Day_Types". Day_Type is decided once by Layer 0/1
    # (farmer_meeting_scheduled_today), not per-DC -- so Long-Term/BO5 activity only
    # ever shows up via the standalone Farmer_Meeting_Task path above, never bundled
    # into a DC Visit task alongside Visits/Outstanding/PL. Note: the sheet's own Layer
    # 2 formula text still lists "Qualify_LongTerm" in the Candidate_DCs union, but its
    # Matched_Purposes set (Promise_To_Bill/Promise_To_Pay_Collection/Query_Resolution/
    # Sale/Stock_at_DC -- all real DC Visit purposes) never draws from Long-Term either
    # way, and the Final row's "Matched_Objectives" language wasn't updated to match
    # Layer 2's rename to "Matched_Purposes" -- read as an incomplete edit in the source
    # sheet, not a deliberate instruction to violate the GR-12 exclusivity it just added.
    QUALIFIERS = {"Visits": _qualify_visits, "Outstanding": _qualify_outstanding, "PL": _qualify_pl}

    # Layer 3 -- Priority_Score(DC) = 0.40/0.35/0.25 of the DC's own top-3 matched
    # objectives' grades. True per-DC BO1/BO3 grading needs data this pipeline doesn't
    # have yet (historical Outstanding for 3.1's Expected_Outstanding; 1.2's PL_Expected
    # combination method is itself still TBD in Source 5) -- so this substitutes the
    # SE-level objective score (_objective_gap) as this DC's proxy grade per matched
    # objective, same honest-degrade signal the old engine used, just applied per-DC.
    weights = (constants.rank1_weight, constants.rank2_weight, constants.rank3_weight)
    pool: List[Tuple[Dict[str, Any], List[str], float]] = []
    # Recorded for every in-scope DC that fails ALL 3 qualifiers -- discarded below
    # (never attached to the returned plan) unless ranked_pool ends up completely empty,
    # in which case this is exactly the detail behind the generic "No in-scope DC
    # qualified for any objective this run" skip reason (see Skipped_Qualification_Detail
    # in the return dict, and planning.services' skipped_ses handling). Free to compute
    # either way -- it's the same QUALIFIERS check already running per DC, just kept
    # instead of discarded.
    unqualified_detail: List[Dict[str, Any]] = []
    for dc in ranked_dcs:
        matched = [obj for obj, fn in QUALIFIERS.items() if fn(dc)]
        if not matched:
            d = dc.get("Days_Since_Last_Visit")
            outstanding = dc_financials.get(dc["DC_ID"], {}).get("Current_Outstanding")
            pl_grade = dc_bo_scores.get(dc["DC_ID"], {}).get("PL", {}).get("grade")
            # Promise To Pay override note (2026-09-04) -- a "Kept" promise can suppress
            # Outstanding even when the raw balance itself would have qualified; say so
            # explicitly here instead of leaving the balance-only message looking like
            # the balance check simply failed on its own terms.
            outstanding_detail = f"balance={outstanding!r} (need >={constants.qualify_outstanding_balance:,.0f})"
            if _promise_status(dc["DC_ID"]) == "Kept":
                outstanding_detail += " -- suppressed: most recent Promise To Pay was Kept"
            unqualified_detail.append({
                "DC_ID": dc["DC_ID"], "DC_Name": dc.get("DC_Name"),
                "Visits": f"last visit {d}d ago (need >{constants.qualify_visits_days_since}d/never)",
                "Outstanding": outstanding_detail,
                "PL": f"grade={pl_grade!r} (need C/D)",
            })
            continue
        # Prefer a real per-DC grade (dc_bo_scores) over the SE-level proxy when one is
        # supplied -- e.g. score_bo3_outstanding_live_proxy() gives Outstanding a real,
        # DC-specific severity instead of every matched DC sharing one SE-wide floor
        # score. Falls back to the old SE-level _objective_gap() for objectives with no
        # per-DC score wired yet (PL/Long-Term/Sales/Liquidation, as of 2026-08-06).
        per_dc = dc_bo_scores.get(dc["DC_ID"], {})
        gap_by_obj = sorted(
            ((_objective_gap(per_dc[obj]) if obj in per_dc else _objective_gap(bo_scores_by_objective.get(obj, {})), obj) for obj in matched),
            reverse=True,
        )
        # Bug fixed 2026-08-26: contact-fatigue (8.7 -- >=2 attempts in the rolling
        # window cuts priority by contact_fatigue_priority_cut, 30%) used to be computed
        # ONLY inside _build_candidate_row(), after ranked_pool's order was already
        # final -- it was stored as Priority_Multiplier and shown in Reason_Of_Visit
        # ("contact-fatigue -30% ...") but never actually multiplied into the
        # priority_score that determines ranking/selection (legacy greedy-fill's
        # capacity-limited top-N, and the Routing Agent's priority_score_captured sum).
        # A DC hitting the fatigue threshold could keep topping the ranking with its
        # full, undiscounted score -- the exact "same DC hammered every day" outcome
        # 8.7 exists to prevent. Computed once here, applied to priority_score, and
        # threaded through as the single source of truth so _build_candidate_row's
        # displayed multiplier can never drift from what actually affected ranking.
        attempts = recent_attempts_by_dc.get(dc["DC_ID"], 0)
        fatigue_multiplier = (1 - constants.contact_fatigue_priority_cut) if attempts >= constants.contact_fatigue_max_attempts else 1.0
        priority_score = sum(w * gap for w, (gap, _) in zip(weights, gap_by_obj)) * fatigue_multiplier
        # 90+ day aged overdue queue-jump (explicit user request 2026-09-04) -- same
        # current_overdue>0 + os_90_plus>0 gate _build_candidate_row uses for Overdue_
        # Aging_Bucket, so this never fires on a genuine Rs0-overdue/aging-mismatch case.
        # Applied AFTER fatigue_multiplier on purpose -- a queue-jump-worthy DC should
        # never be knocked back down by the contact-fatigue discount.
        fin_for_boost = dc_financials.get(dc["DC_ID"], {})
        current_overdue_for_boost = fin_for_boost.get("Current_Overdue")
        os_90_plus_for_boost = fin_for_boost.get("OS_90_Plus")
        if (
            "Outstanding" in matched
            and current_overdue_for_boost and current_overdue_for_boost > 0
            and os_90_plus_for_boost and os_90_plus_for_boost > 0
        ):
            priority_score += constants.overdue_90_plus_priority_boost
        pool.append((dc, [o for _, o in gap_by_obj], priority_score, fatigue_multiplier))

    ranked_pool = sorted(
        pool,
        key=lambda t: (-t[2], min(tie_break_order.index(o) if o in tie_break_order else 99 for o in t[1])),
    )

    # Per-DC row construction -- shared by both the legacy greedy-fill path and the
    # Routing Agent path below, so the two never drift apart on what a candidate's
    # Purpose/Reason/financials/overdue-aging actually say. Pure per-DC transform, no
    # capacity state -- safe to call for every ranked_pool entry, not just selected ones.
    def _build_candidate_row(dc: Dict[str, Any], matched: List[str], multiplier: float) -> DailyTaskRow:
        dc_id = dc["DC_ID"]
        attempts = recent_attempts_by_dc.get(dc_id, 0)
        fin = dc_financials.get(dc_id, {})
        club = dc_club_by_id.get(dc_id)
        credit_on_hold = bool(fin.get("Credit_On_Hold"))

        # Real aging bucket, from dc_datamart's own os_1_to_90/os_90_plus split -- see
        # DailyTaskRow.Overdue_Aging_Bucket docstring for why this isn't an exact date.
        # Gated on Current_Overdue > 0 first (fixed live 2026-08-06, caught in testing):
        # os_1_to_90/os_90_plus measure aging of the OUTSTANDING balance overall (likely
        # since invoice date), not specifically the OVERDUE/past-due portion --
        # dc_datamart's own is_mismatch flag exists precisely because these breakdowns
        # don't always reconcile. Attaching an aging label like "(1-90 days)" next to a
        # genuine Rs0 overdue figure is self-contradictory and misleading, so no bucket
        # is shown at all unless there's a real overdue amount to attach it to.
        current_overdue = fin.get("Current_Overdue")
        os_90_plus, os_1_to_90 = fin.get("OS_90_Plus"), fin.get("OS_1_To_90")
        if not current_overdue or current_overdue <= 0:
            overdue_aging = None
        elif os_90_plus and os_90_plus > 0:
            overdue_aging = "90+ days"
        elif os_1_to_90 and os_1_to_90 > 0:
            overdue_aging = "1-90 days"
        else:
            overdue_aging = "Current month"
        # Same gate as overdue_aging above, same reason: Weighted_Avg_Repayment_Days is
        # an average across the whole outstanding balance, not specifically the overdue
        # portion -- showing a days figure next to a genuine Rs0 overdue amount would be
        # just as self-contradictory as an aging-bucket label would be there.
        avg_repayment_days = fin.get("Weighted_Avg_Repayment_Days") if (current_overdue and current_overdue > 0) else None

        # Confirmed 2026-08-18 -- cross-cutting "cover this one first" signal, computed
        # from data already gathered above (no new query): chronic miss escalation,
        # 90+ day aged overdue, or credit-on-hold. See DailyTaskRow.Critical docstring.
        misses = consecutive_misses_by_dc.get(dc_id, 0)
        critical_reasons = []
        if misses >= DC_VISIT_ESCALATION_THRESHOLD:
            critical_reasons.append(f"Escalated -- missed {misses}x running")
        if overdue_aging == "90+ days":
            critical_reasons.append(f"₹{current_overdue:,.0f} overdue 90+ days")
        if credit_on_hold:
            critical_reasons.append("Credit on hold")
        promise_status = _promise_status(dc_id)
        promise = promise_by_dc.get(dc_id) or {}
        if promise_status == "Broken":
            critical_reasons.append(
                f"Broke promise to pay ₹{promise.get('Promise_Amount') or 0:,.0f} by {promise.get('Promise_Date')}"
            )

        # 8.12 bundling: one visit-task covers every matched objective for this DC.
        purpose = " + ".join(dict.fromkeys(PURPOSE_BY_OBJECTIVE.get(o, o) for o in matched))
        per_dc_scores = dc_bo_scores.get(dc_id, {})
        grade_notes = [f"{o} Grade {per_dc_scores[o]['grade']} ({per_dc_scores[o].get('reason', '')})" for o in matched if o in per_dc_scores and per_dc_scores[o].get("grade")]
        overdue_90_plus_boosted = overdue_aging == "90+ days" and "Outstanding" in matched
        reason = (
            f"Matched {', '.join(matched)} -- {dc.get('Cohort')} cohort, rank {dc.get('Rank')}"
            + (f" -- {'; '.join(grade_notes)}" if grade_notes else "")
            + (f", contact-fatigue -{int(constants.contact_fatigue_priority_cut*100)}% ({attempts} attempts in {constants.contact_fatigue_window_days}d)" if multiplier < 1.0 else "")
            + (f", queue-jumped for 90+ day aged overdue (+{constants.overdue_90_plus_priority_boost:.0f} priority)" if overdue_90_plus_boosted else "")
        )
        return DailyTaskRow(
            Sr_No=0, DC_Name=dc.get("DC_Name"), DC_ID=dc_id, Distance_Km=None,
            Recommended_Task_Type="DC Visit", Purpose_Of_Visit=purpose,
            Reason_Of_Visit=reason,
            Last_Visit_Date=dc.get("Last_Visit_Date"), Days_Since_Last_Visit=dc.get("Days_Since_Last_Visit"),
            Present_Outstanding=fin.get("Current_Outstanding"), Present_Overdue=fin.get("Current_Overdue"),
            Last_Order_Date=fin.get("Last_Order_Date"), Last_Order_Value=fin.get("Last_Order_Value"),
            Last_Payment_Date=last_payment_by_dc.get(dc_id), YTD_Private_Label=ytd_pl_by_dc.get(dc_id),
            DC_Club_Participation=dc_club_participation_text(club) if dc_club_by_id else "Config_Ambiguous -- DC club data not supplied",
            Club_Detail=club if dc_club_by_id else None,
            Objective=",".join(matched), No_New_Orders=_no_new_orders(dc_id),
            Credit_On_Hold=credit_on_hold, Credit_On_Hold_Reason=fin.get("Credit_On_Hold_Reason"),
            Estimated_Duration=constants.visit_duration_min, Priority_Multiplier=multiplier,
            Finance_Status=fin.get("Partner_Finance_Status"),
            Promise_To_Pay_Date=promise.get("Promise_Date"),
            Promise_To_Pay_Amount=promise.get("Promise_Amount"),
            Promise_Status=promise_status,
            BO_Scores=per_dc_scores or None,
            BO_Composite_Score=_bo_composite_score(per_dc_scores),
            Last_Payment_Join_Key_Unconfirmed=False,
            Overdue_Aging_Bucket=overdue_aging,
            Avg_Repayment_Days=avg_repayment_days,
            Critical=bool(critical_reasons),
            Critical_Reasons="; ".join(critical_reasons),
        )

    if route_selector is None:
        # FINAL (8.10/GR-11) -- fill up to max_daily_tasks (5) within the field-time
        # budget. No Call-type tasks in this formula -- Candidate_DCs excludes BO4/Sales
        # entirely (see Layer 2 docstring above), so call_minutes_used stays 0 by
        # construction. Unchanged since before route_selector existed -- Step 1's
        # legacy run_pipeline() path (route_selector=None) always takes this branch.
        rows: List[DailyTaskRow] = []
        selected_dc_ids: set = set()
        call_minutes_used = field_minutes_used = 0
        credit_blocked_count = 0

        for dc, matched, priority_score, fatigue_multiplier in ranked_pool:
            if len(rows) >= constants.max_daily_tasks:
                break
            dc_id = dc["DC_ID"]
            if dc_id in selected_dc_ids:
                continue
            duration = constants.visit_duration_min
            if field_minutes_used + duration > constants.field_minutes_per_day:
                break
            if call_minutes_used + field_minutes_used + duration > constants.total_capacity_min:
                break
            row = _build_candidate_row(dc, matched, fatigue_multiplier)
            if row.Credit_On_Hold:
                credit_blocked_count += 1
            rows.append(row)
            selected_dc_ids.add(dc_id)
            field_minutes_used += duration

        # GR-11 -- post-generation safety net, redundant with the loop's own break
        # above but kept as an explicit second check per the Guardrails sheet's pattern.
        if len(rows) > constants.max_daily_tasks:
            rows = rows[: constants.max_daily_tasks]

        sequenced, travel_cap_warning, sequencing_basis = sequence_with_distance(rows, dc_by_id, constants, punch_in_coords)
    else:
        # Routing Agent path (planning/routing.py) -- every ranked_pool candidate gets a
        # fully-formed row (no capacity trimming here; that's the Routing Agent's job,
        # against the real 420-min-total/180-min-travel-floor/5-task caps per Models
        # 1-3), tagged with its Priority_Score so the router can rank/select/sequence
        # for real.
        candidates = [
            {"row": _build_candidate_row(dc, matched, fatigue_multiplier), "dc": dc, "priority_score": priority_score, "matched": matched}
            for dc, matched, priority_score, fatigue_multiplier in ranked_pool
        ]
        selector_result = route_selector(candidates, se_id, plan_date, punch_in_coords, constants, dc_by_id)
        sequenced = selector_result["Tasks"]
        travel_cap_warning = selector_result.get("Travel_Cap_Exceeded", False)
        sequencing_basis = selector_result.get("Sequencing_Basis", "routing_agent")
        field_minutes_used = sum(r.Estimated_Duration for r in sequenced)
        call_minutes_used = 0
        credit_blocked_count = sum(1 for r in sequenced if r.Credit_On_Hold)

    # GR-14 -- independent second-pass exclusion check (redundant with
    # apply_dc_exclusion_rules() upstream by design, not treated as a duplicate). Legal
    # Hold is only ever recorded as DC_Status == "Legal_Hold" (see apply_dc_exclusion_
    # rules()'s `legal_hold = dc.get("DC_Status") == "Legal_Hold"`) -- no dc dict ever
    # carries a separate "Legal_Hold" key, so checking .get("Legal_Hold") here always
    # returned None/False and silently never caught anything; this half of the
    # redundant check was dead since it was written.
    gr14_violations = [
        r.DC_ID for r in sequenced
        if dc_by_id.get(r.DC_ID, {}).get("DC_Status") in ("Inactive", "Closed", "Legal_Hold")
    ]

    objectives_used = sorted(
        {o for r in sequenced for o in r.Objective.split(",") if o},
        key=lambda o: tie_break_order.index(o) if o in tie_break_order else 99,
    )

    _assign_bo_ranks(sequenced)

    return {
        **header,
        "Ranked_Objectives": objectives_used,
        "Tasks": [r.__dict__ for r in sequenced],
        # Only present when ranked_pool ended up completely empty (every in-scope DC
        # failed all 3 qualifiers) -- the detail behind the generic "No in-scope DC
        # qualified for any objective this run" skip reason. planning.services folds
        # this into skipped_ses alongside the separate Not_In_Scope tier (DCs excluded
        # before ever reaching this function at all -- see apply_dc_exclusion_rules()).
        "Skipped_Qualification_Detail": unqualified_detail if (not ranked_pool and unqualified_detail) else None,
        "Capacity_Check": {
            "Call_Minutes_Used": call_minutes_used, "Call_Minutes_Budget": constants.call_minutes_per_day,
            "Field_Minutes_Used": field_minutes_used, "Field_Minutes_Budget": constants.field_minutes_per_day,
            "Total_Minutes_Used": call_minutes_used + field_minutes_used, "Total_Capacity_Minutes": constants.total_capacity_min,
            "Tasks_Used": len(sequenced), "Max_Daily_Tasks": constants.max_daily_tasks,
        },
        "Safety_Flags": {
            "Rule_7_4a_Applied": rule_a_applied,
            "Rule_7_4a_Note": "Outstanding=D caps Overall Sales at B -- no longer affects task selection under the new formula, since BO4/Sales isn't part of Candidate_DCs (8.12); kept for informational/audit purposes only" if rule_a_applied else None,
            "Rule_7_4b_All_D_Reorder_Applied": all_d,
            "No_New_Orders_DC_Count": sum(1 for r in sequenced if r.No_New_Orders),
            "Legal_Hold_Exclusions": "applied upstream in apply_dc_exclusion_rules() -- excluded DCs never reach dc_candidates",
            "Credit_Blocked_DC_Count": credit_blocked_count,
            "Credit_Blocked_Basis": "sale_orderrequest.credit_on_hold (Source 3d, confirmed queryable) -- flagged per 6.4, not auto-blocked" if dc_financials else "not evaluated -- dc_financials not supplied this run",
            "GR_14_Second_Pass_Violations": gr14_violations,
            "Qualify_Outstanding_Days_Overdue_Leg": "not enforced -- no confirmed per-DC days-overdue field in dc_datamart; balance-only qualification applied (see Layer 2 docstring)",
            # The literal '<3 PL orders in 30 days' (8.5) is never enforced either way --
            # coupon_analysis has no confirmed DC-level join key. What DOES vary is whether
            # this call's dc_bo_scores carries a real per-DC PL grade (planning/services.py
            # wires this in for every Django-generated plan; the bare CLI run_pipeline()
            # path never does) -- _qualify_pl() uses that live grade C/D as a proxy
            # whenever it's present, so the flag must say which case this run actually is,
            # not a single static claim that was only ever true for the CLI path.
            "Qualify_PL_Order_Count": (
                "literal order-count rule not enforced (no confirmed DC-level join key) -- live proxy applied instead: "
                "a DC qualifies for PL when its live PL_Ratio grade is C or D (see _qualify_pl docstring)"
                if any((v or {}).get("PL", {}).get("grade") is not None for v in dc_bo_scores.values())
                else
                "not enforced -- coupon_analysis has no confirmed DC-level join key, and no per-DC PL grade "
                "(dc_bo_scores) was supplied this run, so PL objective never qualifies a DC into Candidate_DCs "
                "on this path (see _qualify_pl docstring)"
            ),
        },
        "Travel": {"Sequencing_Basis": sequencing_basis, "Daily_Cap_Km": constants.daily_travel_cap_km, "Cap_Exceeded": travel_cap_warning},
        "Data_Confidence": "Live" if attendance_gate_ok is True else "Provisional_No_Attendance_Gate",
    }


# =====================================================================================
# 12. PIPELINE ORCHESTRATION (Section 6 process flow)
# =====================================================================================

def run_pipeline(output_dir: Path, plan_date: Optional[str] = None) -> Dict[str, Any]:
    run_ts = utc_now_iso()
    plan_date = plan_date or datetime.now(timezone.utc).date().isoformat()
    all_exceptions: Table = []
    check_summary: Dict[str, Dict[str, int]] = {}

    def merge(exc: Exceptions) -> None:
        all_exceptions.extend(exc.rows)
        for code, counts in exc.summary().items():
            dst = check_summary.setdefault(code, {"pass": 0, "fail": 0})
            dst["pass"] += counts["pass"]
            dst["fail"] += counts["fail"]

    logger.info("1/6 loading Source 5 (Config Master)...")
    config_rows, exc = load_config()
    merge(exc)
    config_index = {r["Param_Key"]: r["Configured_Value"] for r in config_rows}

    logger.info("2/6 loading Source 2 (DC Master + Rank)...")
    dc_master, exc = load_dc_master()
    merge(exc)

    logger.info("3/6 loading Source 6 (AOP & Target)...")
    aop_targets, exc = load_aop_targets()
    merge(exc)

    logger.info("4/6 loading Sources 1/3/4 (live via Metabase)...")
    client = get_client()
    try:
        live, exc = load_live_sources(client)
        merge(exc)

        logger.info("  4/6 supplementing DC Master with live-active DCs missing from DC_RAnk.csv...")
        supplement_exc = Exceptions(run_ts)
        supplement_dc_master_from_live(dc_master, client, supplement_exc)
        merge(supplement_exc)
    finally:
        client.close()

    logger.info("5/6 normalizing + cross-checking...")
    attendance_exc = Exceptions(run_ts)
    attendance = normalize_attendance(live["Attendance_3a"], attendance_exc)
    merge(attendance_exc)
    visits_exc = Exceptions(run_ts)
    visits = normalize_visits(live["Active_Roster_4"], live["Task_Nodes_1b"], config_index, visits_exc)
    merge(visits_exc)
    sales_exc = Exceptions(run_ts)
    sales, dc_financials = normalize_sales_transactions(
        live["Sales_Transactions_3d"], live["Outstanding_3d"], live["Orders_3d"], sales_exc
    )
    merge(sales_exc)
    payments_exc = Exceptions(run_ts)
    payments, last_payment_by_dc = normalize_payments(live["Payments_3f"], payments_exc)
    merge(payments_exc)
    promise_exc = Exceptions(run_ts)
    promise_by_dc = normalize_promise_to_pay(live["Promise_To_Pay_3j"], promise_exc)
    merge(promise_exc)
    club_exc = Exceptions(run_ts)
    dc_club = normalize_dc_club(
        live["DC_Club_Mapping_3g"], live["DC_Club_Slabs_3g"], live["DC_Club_Qualifying_Turnover_3g"],
        dc_financials, club_exc,
    )
    merge(club_exc)
    dc_club_by_id = {row["DC_ID"]: row for row in dc_club}
    if dc_club_scheme_window_expired(plan_date):
        expiry_exc = Exceptions(run_ts)
        expiry_exc.flag(
            "ALL", "Source3g", "Club_Scheme_Window_Expired",
            f"plan_date {plan_date} is past the DC Club scheme's confirmed validity window "
            f"({DC_CLUB_SCHEME_WINDOW_START} to {DC_CLUB_SCHEME_WINDOW_END}, exclusive) -- "
            "SQL_DC_CLUB_QUALIFYING_TURNOVER_3G now silently returns no rows; every DC's "
            "Qualifying_Turnover/Club_Tier will read None until the scheme's next-year window "
            "is confirmed and DC_CLUB_SCHEME_WINDOW_START/END are updated.",
        )
        merge(expiry_exc)

    liquidation_exc = Exceptions(run_ts)
    liquidation = normalize_liquidation(live["Liquidation_3d"], liquidation_exc)
    merge(liquidation_exc)

    geo_check_exc = Exceptions(run_ts)
    run_cross_source_checks(dc_master, live["Geo_Mapping_1c"], geo_check_exc)
    merge(geo_check_exc)

    constants = BusinessConstants()
    drift_exc = Exceptions(run_ts)
    check_business_constants_against_config(constants, config_rows, drift_exc)
    merge(drift_exc)
    last_visit_by_dc: Dict[str, str] = {}
    for v in visits:
        if v["DC_ID"] and v["Date"]:
            if v["DC_ID"] not in last_visit_by_dc or v["Date"] > last_visit_by_dc[v["DC_ID"]]:
                last_visit_by_dc[v["DC_ID"]] = v["Date"]
    top_dc_allowlist, top_dc_exc = load_top_dc_allowlist()
    merge(top_dc_exc)
    excl_exc = Exceptions(run_ts)
    apply_dc_exclusion_rules(dc_master, excl_exc, constants, last_visit_by_dc, plan_date, top_dc_allowlist=top_dc_allowlist)
    merge(excl_exc)

    ref_exc = Exceptions(run_ts)
    dc_ids = {dc["DC_ID"] for dc in dc_master}
    se_ids = {row.get("se_user_id") for row in live["Active_Roster_4"]}
    referential_integrity_check(visits, sales, dc_ids, se_ids, ref_exc)
    merge(ref_exc)

    compliance_exc = Exceptions(run_ts)
    valid_visit_compliance(visits, compliance_exc)
    merge(compliance_exc)

    logger.info("6/6 scoring + generating SE Daily Plans...")
    se_visit_history: Dict[str, List[float]] = defaultdict(list)
    for v in visits:
        if v["SE_ID"]:
            se_visit_history[v["SE_ID"]].append(1.0)
    dynamic_params = resolve_dynamic_parameters(dc_master, aop_targets, se_visit_history, config_rows, constants)

    attendance_by_se_today = {a["SE_ID"] for a in attendance if a["Date"] == plan_date and a["Attendance_Status"] != "Invalid"}

    # 8.7 contact-fatigue: attempts per (SE, DC) within the rolling fatigue window, counted
    # from every Visits_Normalized row (any status) up to plan_date -- not just valid visits.
    fatigue_window_start = (datetime.fromisoformat(plan_date) - timedelta(days=constants.contact_fatigue_window_days)).date().isoformat()
    recent_attempts_all: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for v in visits:
        if v.get("SE_ID") and v.get("DC_ID") and v.get("Date") and fatigue_window_start <= v["Date"] < plan_date:
            recent_attempts_all[v["SE_ID"]][v["DC_ID"]] += 1

    se_names = {row.get("se_user_id"): row.get("se_email") for row in live["Task_Nodes_1b"]}
    punch_in_by_se: Dict[str, Tuple[float, float]] = {}
    for a in attendance:
        if a["Date"] == plan_date and a.get("Punch_In_Lat") is not None and a.get("Punch_In_Long") is not None:
            punch_in_by_se[a["SE_ID"]] = (a["Punch_In_Lat"], a["Punch_In_Long"])

    daily_plans = []
    for se_id in sorted(se_ids) if se_ids else []:
        se_dc_candidates = [dc for dc in dc_master if dc.get("Assigned_SE_Email")]  # scoped further by real join in production
        bo_scores = {
            "Visits": score_bo2_visits(sum(1 for v in visits if v["SE_ID"] == se_id and v["Valid_Visit_Flag"]), len(se_dc_candidates), constants),
            "PL": {"score_pct": None, "grade": None, "reason": "PL_Value/PL_Expected need Sales_Transactions_Normalized joined by DC -- wire in once live"},
            "Outstanding": {"ratio": None, "grade": None},
            "Sales": {"score_pct": None, "grade": None},
            "Liquidation": {"score_pct": None, "grade": None, "reason": "no confirmed scoring formula exists (Source 3d Provisional)"},
            "Long-Term": score_bo5_long_term(0, 0, constants),
        }
        attendance_gate_ok = (se_id in attendance_by_se_today) if client.configured else None
        plan = generate_se_daily_plan(
            se_id, se_names.get(se_id), plan_date, se_dc_candidates, bo_scores, dynamic_params, constants,
            attendance_gate_ok, recent_attempts_by_dc=dict(recent_attempts_all.get(se_id, {})),
            dc_financials=dc_financials, last_payment_by_dc=last_payment_by_dc, dc_club_by_id=dc_club_by_id,
            punch_in_coords=punch_in_by_se.get(se_id), promise_by_dc=promise_by_dc,
        )
        daily_plans.append(plan)

    output_dir.mkdir(parents=True, exist_ok=True)
    tables = {
        "DC_Master_Normalized": dc_master,
        "Config_Normalized": config_rows,
        "AOP_Target_Normalized": aop_targets,
        "Attendance_Normalized": attendance,
        "Visits_Normalized": visits,
        "Sales_Transactions_Normalized": sales,
        "Payments_Normalized": payments,
        "DC_Club_Normalized": dc_club,
        "Liquidation_Normalized": liquidation,
        "Geo_Mapping_Normalized": live["Geo_Mapping_1c"],
        "Task_Nodes_Normalized": live["Task_Nodes_1b"],
        "Exceptions_Report": all_exceptions,
        "SE_Daily_Plan": daily_plans,
    }
    for name, data in tables.items():
        with (output_dir / f"{name}.json").open("w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str)

    run_summary = {
        "Run_Timestamp": run_ts,
        "Plan_Date": plan_date,
        "Metabase_Configured": client.configured,
        "Row_Counts": {name: len(data) for name, data in tables.items()},
        "Check_Summary": check_summary,
        "Dynamic_Parameters_Resolved": dynamic_params,
        "Note": (
            "Sources 1/3/4 were skipped (Metabase not configured) -- normalized output "
            "reflects Sources 2/5/6 only; SE_Daily_Plan tasks are Provisional."
            if not client.configured else
            "All 6 sources pulled this run."
        ),
    }
    with (output_dir / "Run_Summary.json").open("w", encoding="utf-8") as f:
        json.dump(run_summary, f, indent=2, default=str)

    logger.info("Done. Wrote %d tables + run summary to %s", len(tables), output_dir)
    return run_summary


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", default=str(BASE_DIR / "output"), help="Where normalized tables + plan are written")
    parser.add_argument("--date", default=None, help="Plan date YYYY-MM-DD (default: today)")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO, format="%(levelname)s %(message)s")
    summary = run_pipeline(Path(args.output_dir), plan_date=args.date)
    print(json.dumps(summary, indent=2, default=str))


if __name__ == "__main__":
    main()
